import os
import sys
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import event
from sqlalchemy.orm import Session

def resource_path(relative_path):
    """Résout les chemins des ressources pour les exécutables PyInstaller"""
    if getattr(sys, 'frozen', False):
        # Si l'application est gelée (compilée)
        base_path = sys._MEIPASS
    else:
        # Si l'application est en mode développement
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    return os.path.join(base_path, relative_path)

# Initialisation de l'application Flask
app = Flask(__name__, 
            template_folder=resource_path('templates'),
            static_folder=resource_path('static'))
from sqlalchemy import func, case
from flask_wtf.csrf import CSRFProtect, generate_csrf, CSRFError
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, date, timezone
import os
import re
import unicodedata
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from werkzeug.utils import secure_filename
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT, TA_JUSTIFY
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import arabic_reshaper
from bidi.algorithm import get_display
# from weasyprint import HTML

def format_relative_time(value):
    if not value:
        return "Jamais"
    
    now = datetime.now(timezone.utc)
    if isinstance(value, str):
        value = datetime.fromisoformat(value.replace('Z', '+00:00'))
    
    diff = now - value.replace(tzinfo=timezone.utc) if value.tzinfo is None else now - value
    
    if diff < timedelta(minutes=1):
        return "À l'instant"
    elif diff < timedelta(hours=1):
        minutes = int(diff.seconds / 60)
        return f"Il y a {minutes} minute{'s' if minutes > 1 else ''}"
    elif diff < timedelta(days=1):
        hours = int(diff.seconds / 3600)
        return f"Il y a {hours} heure{'s' if hours > 1 else ''}"
    elif diff < timedelta(days=30):
        days = diff.days
        return f"Il y a {days} jour{'s' if days > 1 else ''}"
    elif diff < timedelta(days=365):
        months = int(diff.days / 30)
        return f"Il y a {months} mois"
    else:
        years = int(diff.days / 365)
        return f"Il y a {years} an{'s' if years > 1 else ''}"

app.jinja_env.filters['relative_time'] = format_relative_time
app.config['SECRET_KEY'] = 'change_this_secret_key_in_production'
app.config['WTF_CSRF_ENABLED'] = True
app.config['WTF_CSRF_SECRET_KEY'] = 'a_random_secret_key_for_csrf'  # Changez cette clé en production
app.config['WTF_CSRF_TIME_LIMIT'] = None  # Pas d'expiration du token CSRF
app.config['WTF_CSRF_CHECK_DEFAULT'] = False  # Désactiver la vérification automatique pour gérer manuellement

# Initialisation de la protection CSRF
csrf = CSRFProtect(app)

# Custom Jinja2 finalize function pour gérer automatiquement les dates
# DÉSACTIVÉ car peut causer des problèmes avec certaines valeurs
# def finalize_date(value):
#     """Convertit automatiquement les strings de dates en datetime objects"""
#     if isinstance(value, str) and value:
#         # Vérifier si c'est une date
#         if ('T' in value or (' ' in value and ':' in value)) and '-' in value:
#             # C'est probablement une date, la convertir
#             try:
#                 # Format ISO avec T
#                 if 'T' in value:
#                     if '.' in value:
#                         parts = value.split('.')
#                         if len(parts) == 2:
#                             base_date = parts[0]
#                             microseconds = parts[1][:6]
#                             value = f"{base_date}.{microseconds}"
#                         return datetime.strptime(value, '%Y-%m-%dT%H:%M:%S.%f')
#                     else:
#                         return datetime.strptime(value, '%Y-%m-%dT%H:%M:%S')
#                 # Format avec espace
#                 elif ' ' in value and ':' in value:
#                     if '.' in value:
#                         return datetime.strptime(value, '%Y-%m-%d %H:%M:%S.%f')
#                     else:
#                         return datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
#             except:
#                 pass
#     return value

# Appliquer la fonction finalize à Jinja2
# DÉSACTIVÉ - peut causer des problèmes
# app.jinja_env.finalize = finalize_date

# Gestionnaire d'erreur CSRF
@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    if request.is_json or request.headers.get('Content-Type') == 'application/json':
        return jsonify({
            'success': False,
            'message': 'Token CSRF invalide ou expiré. Veuillez rafraîchir la page.',
            'error': 'csrf_error'
        }), 400
    flash('Token CSRF invalide ou expiré. Veuillez réessayer.', 'error')
    return redirect(request.referrer or url_for('dashboard'))

# Utiliser un chemin absolu pour la base de données (évite les problèmes de répertoire d'exécution)
basedir = os.path.abspath(os.path.dirname(__file__))

# Helper function pour parser les dates ISO avec gestion des erreurs
def safe_parse_datetime(date_str):
    """
    Parse une chaîne de date en gérant différents formats
    
    Args:
        date_str: Chaîne de date à parser
        
    Returns:
        datetime object ou None si erreur
    """
    if not date_str:
        return None
    
    # Si c'est déjà un objet datetime, le retourner tel quel
    if isinstance(date_str, datetime):
        return date_str
    
    try:
        # Convertir en string si ce n'est pas déjà le cas
        date_str = str(date_str).strip()
        
        # Format ISO avec T (PRIORITÉ 1 - format en base maintenant): 2025-11-11T18:56:03
        if 'T' in date_str:
            if '.' in date_str:
                # Avec microsecondes: 2025-11-11T18:56:03.000000
                parts = date_str.split('.')
                if len(parts) == 2:
                    base_date = parts[0]
                    microseconds = parts[1][:6]  # Limiter à 6 chiffres
                    date_str = f"{base_date}.{microseconds}"
                return datetime.strptime(date_str, '%Y-%m-%dT%H:%M:%S.%f')
            else:
                # Sans microsecondes: 2025-11-11T18:56:03
                return datetime.strptime(date_str, '%Y-%m-%dT%H:%M:%S')
        
        # Format SQLite avec espace (PRIORITÉ 2): 2025-11-11 18:56:03
        elif ' ' in date_str and ':' in date_str:
            if '.' in date_str:
                # Avec microsecondes: 2025-11-11 18:56:03.000000
                return datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S.%f')
            else:
                # Sans microsecondes: 2025-11-11 18:56:03
                return datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
        
        # Format date seule: 2025-11-11
        elif '-' in date_str and ':' not in date_str:
            return datetime.strptime(date_str, '%Y-%m-%d')
        
        else:
            # Dernier recours: retourner None
            return None
            
    except (ValueError, AttributeError) as e:
        print(f"⚠️ Erreur de parsing de date: {date_str} - {str(e)}")
        return None

# Helper function pour convertir TOUS les attributs de date d'un objet
def convert_object_dates(obj):
    """Convertit tous les attributs de date d'un objet en datetime"""
    if not obj:
        return obj
        
    date_attrs = ['created_at', 'updated_at', 'deleted_at', 'date_dotation', 'date_reception',
                  'date_engagement', 'date_service_fait', 'date_reglement', 'date_visa',
                  'date_visa_ordo', 'date_creation', 'date_depart', 'date_retour',
                  'date_limite_reception']
    
    for attr in date_attrs:
        if hasattr(obj, attr):
            try:
                value = getattr(obj, attr)
                if isinstance(value, str) and value:
                    converted = safe_parse_datetime(value)
                    if converted:
                        setattr(obj, attr, converted)
            except:
                pass
    return obj

# Configuration de la base de données
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(basedir, "stock_management.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# TypeDecorator personnalisé pour gérer automatiquement les dates
from sqlalchemy.types import TypeDecorator, DateTime as SQLDateTime, String, Text
from sqlalchemy import event, text
from sqlalchemy.orm import Session
from sqlalchemy.engine import Engine

class AutoDateTime(TypeDecorator):
    """Type personnalisé qui convertit automatiquement les strings en datetime"""
    impl = SQLDateTime
    cache_ok = True
    
    def process_result_value(self, value, dialect):
        """Appelé quand on lit depuis la DB"""
        if value is None:
            return None
        if isinstance(value, str):
            return safe_parse_datetime(value)
        return value
    
    def process_bind_param(self, value, dialect):
        """Appelé quand on écrit dans la DB"""
        if value is None:
            return None
        if isinstance(value, str):
            parsed = safe_parse_datetime(value)
            return parsed if parsed else value
        return value

class AutoDate(TypeDecorator):
    """Type personnalisé pour gérer des colonnes DATE robustes.
    - Accepte en lecture les formats 'YYYY-MM-DD', 'YYYY-MM-DD HH:MM:SS', 'YYYY-MM-DDTHH:MM:SS[.ffffff]'
    - Retourne un objet date Python
    - Sérialise en écriture au format 'YYYY-MM-DD' (string) pour compatibilité SQLite
    """
    impl = String
    cache_ok = True

    def process_result_value(self, value, dialect):
        if value is None:
            return None
        # Déjà une date
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        # Datetime -> date
        if isinstance(value, datetime):
            return value.date()
        # String -> essayer plusieurs formats via safe_parse_datetime
        if isinstance(value, str):
            # Formats directs si possible
            v = value.strip()
            # Cas simple 'YYYY-MM-DD'
            try:
                if len(v) >= 10 and v[4] == '-' and v[7] == '-':
                    # Tronquer tout ce qui suit la date
                    only_date = v[:10]
                    return datetime.strptime(only_date, '%Y-%m-%d').date()
            except Exception:
                pass
            # Utiliser le parser robuste
            dt = safe_parse_datetime(v)
            if dt:
                return dt.date()
        return None

    def process_bind_param(self, value, dialect):
        if value is None:
            return None
        # Accepter date, datetime ou string
        if isinstance(value, date) and not isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')
        if isinstance(value, datetime):
            return value.date().strftime('%Y-%m-%d')
        if isinstance(value, str):
            v = value.strip()
            # Si déjà au bon format 'YYYY-MM-DD'
            try:
                if len(v) >= 10 and v[4] == '-' and v[7] == '-':
                    only_date = v[:10]
                    # Valider
                    datetime.strptime(only_date, '%Y-%m-%d')
                    return only_date
            except Exception:
                pass
            dt = safe_parse_datetime(v)
            if dt:
                return dt.date().strftime('%Y-%m-%d')
        # Dernier recours: convertir en string
        return str(value)

# Event listener GLOBAL pour convertir TOUTES les dates string en datetime
# DÉSACTIVÉ - ne fonctionne pas correctement avec SQLAlchemy
# @event.listens_for(Session, 'loaded_as_persistent')
# def convert_dates_on_load(session, instance):
#     """Convertit automatiquement TOUTES les dates string en datetime après chargement"""
#     convert_object_dates(instance)

# Décorateur pour convertir automatiquement les dates dans les routes
from functools import wraps

def convert_dates_in_response(f):
    """Décorateur qui convertit automatiquement les dates avant le render"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Appeler la fonction originale
        result = f(*args, **kwargs)
        
        # Si c'est un render_template, on ne peut pas intercepter
        # Donc on doit convertir AVANT dans la route
        return result
    
    return decorated_function

# Models
class Region(db.Model):
    __tablename__ = 'regions'
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), unique=True, nullable=False)
    
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    name = db.Column(db.String(100))
    role = db.Column(db.String(50), default='admin')
    is_super_admin = db.Column(db.Boolean, default=False)  # Super Admin du projet
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)  # ID de l'admin qui a créé l'utilisateur
    created_at = db.Column(AutoDateTime, default=datetime.utcnow)
    updated_at = db.Column(AutoDateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    last_login = db.Column(AutoDateTime, nullable=True)  # Dernière connexion de l'utilisateur
    permissions = db.Column(db.Text, nullable=True)  # Permissions avancées (liste séparée par des virgules)
    is_active = db.Column(db.Boolean, default=True)  # Activer/Désactiver l'utilisateur
    
    # Soft Delete
    is_deleted = db.Column(db.Boolean, default=False)  # Marqueur de suppression
    deleted_at = db.Column(AutoDateTime, nullable=True)  # Date de suppression
    deleted_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)  # Qui a supprimé
    
    # Domaine de compétence (Région/Unité/Centre)
    region_id = db.Column(db.Integer, db.ForeignKey('regions.id'), nullable=True)
    unite_id = db.Column(db.Integer, db.ForeignKey('unites.id'), nullable=True)
    
    region_competence = db.relationship('Region', foreign_keys=[region_id], backref='users')
    unite_competence = db.relationship('Unite', foreign_keys=[unite_id], backref='gestionnaires')
    
    def get_permissions(self):
        """Retourne la liste des permissions associées à l'utilisateur"""
        if not self.permissions:
            return []
        return [p.strip() for p in self.permissions.split(',') if p.strip()]

    def set_permissions(self, permissions_list):
        """Enregistre les permissions à partir d'une liste de chaînes"""
        if not permissions_list:
            self.permissions = None
            return
        cleaned = sorted({p.strip() for p in permissions_list if p and p.strip()})
        self.permissions = ','.join(cleaned)

    def has_permission(self, permission):
        """Vérifie si l'utilisateur possède une permission donnée"""
        if self.is_super_admin:
            return True
        perms = self.get_permissions()
        if perms:
            return permission in perms
        # Fallback basé sur le rôle si aucune permission explicite n'est définie
        if self.role == 'admin':
            return True
        if self.role == 'gestionnaire':
            return permission in ['stock', 'receptions', 'dotations', 'reports']
        if self.role == 'user':
            return permission in ['reports']
        return False

    def set_password(self, password):
        """Définit le mot de passe hashé"""
        self.password_hash = generate_password_hash(password)
    
    def verify_password(self, password):
        """Vérifie le mot de passe"""
        return check_password_hash(self.password_hash, password)

class Item(db.Model):
    __tablename__ = 'items'
    id = db.Column(db.Integer, primary_key=True)
    sku = db.Column(db.String(100), unique=True, nullable=False)
    name = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    category = db.Column(db.String(100))
    unit = db.Column(db.String(50), default='Unité')
    reorder_level = db.Column(db.Integer, default=0)
    quantity = db.Column(db.Integer, default=0)
    created_at = db.Column(AutoDateTime, default=datetime.utcnow)
    updated_at = db.Column(AutoDateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Soft Delete articles
    is_deleted = db.Column(db.Boolean, default=False)
    deleted_at = db.Column(AutoDateTime, nullable=True)
    deleted_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    
    # Relations supprimées - les réceptions sont gérées séparément
    
    @staticmethod
    def reorganize_all_skus():
        """Réorganise tous les SKU pour qu'ils soient séquentiels sans trous
        SKU-001, SKU-002, SKU-003, etc.
        """
        # Récupère tous les articles triés par leur SKU actuel
        all_items = Item.query.filter(Item.sku.like('SKU-%')).order_by(Item.id).all()
        
        if not all_items:
            return
        
        # Renumérote tous les articles de manière séquentielle
        for index, item in enumerate(all_items, start=1):
            new_sku = f"SKU-{index:03d}"
            if item.sku != new_sku:
                item.sku = new_sku
        
        db.session.commit()
    
    @staticmethod
    def generate_next_sku():
        """Génère le prochain SKU unique en s'incrémentant automatiquement
        Commence toujours à SKU-001 et suit l'ordre séquentiel
        """
        # Récupère tous les SKU existants
        all_items = Item.query.filter(Item.sku.like('SKU-%')).all()
        
        if not all_items:
            # Aucun article : commence à SKU-001
            return "SKU-001"
        
        # Compte simplement le nombre d'articles + 1
        next_number = len(all_items) + 1
        
        # Génère le nouveau SKU avec format SKU-XXX (3 chiffres avec zéros)
        return f"SKU-{next_number:03d}"

class Unite(db.Model):
    __tablename__ = 'unites'
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    created_at = db.Column(AutoDateTime, default=datetime.utcnow)
    
    # Nouveaux champs pour la hiérarchie
    region_id = db.Column(db.Integer, db.ForeignKey('regions.id'), nullable=True)
    parent_id = db.Column(db.Integer, db.ForeignKey('unites.id'), nullable=True)
    type = db.Column(db.String(50), default='centre')  # 'unite' (Provinciale) ou 'centre' (Secours)
    
    region = db.relationship('Region', foreign_keys=[region_id], backref='unites')
    services = db.relationship('Service', backref='unite', lazy=True, cascade='all, delete-orphan')
    children = db.relationship('Unite', backref=db.backref('parent', remote_side=[id]), lazy=True)
    # Relations mises à jour vers Dotation

class Service(db.Model):
    __tablename__ = 'services'
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(200), nullable=False)
    unite_id = db.Column(db.Integer, db.ForeignKey('unites.id'), nullable=False)
    description = db.Column(db.Text)
    created_at = db.Column(AutoDateTime, default=datetime.utcnow)
    
    # Relations mises à jour vers Dotation

class Dotation(db.Model):
    __tablename__ = 'dotations'
    id = db.Column(db.Integer, primary_key=True)
    numero_dotation = db.Column(db.String(50), unique=True, nullable=False)  # DOT-YYYY-XXXX
    unite_id = db.Column(db.Integer, db.ForeignKey('unites.id'), nullable=False)
    service_id = db.Column(db.Integer, db.ForeignKey('services.id'), nullable=True)  # Optionnel
    categorie = db.Column(db.String(100), nullable=False)  # Catégorie des articles
    date_dotation = db.Column(AutoDateTime, default=datetime.utcnow)
    statut = db.Column(db.String(50), default='en_cours', nullable=False)
    notes = db.Column(db.Text)
    
    dotee_par = db.Column(db.Integer, db.ForeignKey('users.id'))
    
    # NOUVEAU: Champ pour le fichier PDF de décharge signée
    decharge_signee_path = db.Column(db.String(500), nullable=True)  # Chemin vers le PDF signé
    decharge_signee_filename = db.Column(db.String(200), nullable=True)  # Nom original du fichier
    decharge_signee_date = db.Column(db.DateTime, nullable=True)  # Date d'import du PDF signé
    
    # Soft Delete dotations
    is_deleted = db.Column(db.Boolean, default=False)
    deleted_at = db.Column(AutoDateTime, nullable=True)
    deleted_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    
    # Relations
    unite = db.relationship('Unite', backref='dotations')
    service = db.relationship('Service', backref='dotations')
    user = db.relationship('User', backref='dotations', foreign_keys=[dotee_par])
    deleter = db.relationship('User', foreign_keys=[deleted_by])
    items = db.relationship('DotationItem', backref='dotation', cascade='all, delete-orphan')
    
    @property
    def cout_total_dotation(self):
        """Calcule le coût total de la dotation (somme des coûts de tous les articles)"""
        return sum(item.cout_total for item in self.items)
    
    @staticmethod
    def generate_numero_dotation():
        """Génère un numéro de dotation unique (DOT-2025-001)"""
        year = datetime.now().year
        last_dotation = Dotation.query.filter(
            Dotation.numero_dotation.like(f'DOT-{year}-%')
        ).order_by(Dotation.numero_dotation.desc()).first()
        
        if last_dotation:
            try:
                last_number = int(last_dotation.numero_dotation.split('-')[2])
                next_number = last_number + 1
            except (ValueError, IndexError):
                next_number = 1
        else:
            next_number = 1
        
        return f"DOT-{year}-{next_number:03d}"

class DotationItem(db.Model):
    __tablename__ = 'dotation_items'
    id = db.Column(db.Integer, primary_key=True)
    dotation_id = db.Column(db.Integer, db.ForeignKey('dotations.id'), nullable=False)
    item_id = db.Column(db.Integer, db.ForeignKey('items.id'), nullable=False)
    quantite_dotee = db.Column(db.Integer, nullable=False)
    prix_unitaire = db.Column(db.Float, default=0.0)  # Prix unitaire de l'article (HT ou TTC selon config)
    
    # NOUVEAUX CHAMPS POUR FIFO
    prix_unitaire_ht = db.Column(db.Float, default=0.0)  # Prix unitaire HT
    prix_unitaire_ttc = db.Column(db.Float, default=0.0)  # Prix unitaire TTC
    taux_tva = db.Column(db.Float, default=20.0)  # Taux TVA appliqué
    montant_total_ht = db.Column(db.Float, default=0.0)  # Montant total HT (qté × PU HT)
    montant_total_ttc = db.Column(db.Float, default=0.0)  # Montant total TTC (qté × PU TTC)
    reception_stock_id = db.Column(db.Integer, db.ForeignKey('reception_stocks.id'), nullable=True)  # Lien vers le stock source FIFO
    
    # Relations
    item = db.relationship('Item', backref='dotation_items')
    reception_stock = db.relationship('ReceptionStock', backref='dotation_items')
    
    @property
    def cout_total(self):
        """Calcule le coût total (quantité × prix unitaire) - Rétrocompatibilité"""
        return self.quantite_dotee * (self.prix_unitaire or 0.0)
    
    @property
    def cout_total_ht(self):
        """Calcule le coût total HT"""
        return self.montant_total_ht or (self.quantite_dotee * (self.prix_unitaire_ht or 0.0))
    
    @property
    def cout_total_ttc(self):
        """Calcule le coût total TTC"""
        return self.montant_total_ttc or (self.quantite_dotee * (self.prix_unitaire_ttc or 0.0))

# Ancienne classe Movement supprimée - remplacée par Reception

class Reception(db.Model):
    __tablename__ = 'receptions'
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('items.id'), nullable=False)
    type = db.Column(db.String(15), nullable=False)  # 'reception', 'ajustement'
    quantity = db.Column(db.Integer, nullable=False)
    reason = db.Column(db.String(200))
    date_reception = db.Column(AutoDate, nullable=True)  # Date de la réception
    created_at = db.Column(AutoDateTime, default=datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    
    # NOUVEAUX CHAMPS BUDGÉTAIRES
    prix_unitaire_ht = db.Column(db.Float, nullable=False, default=0.0)  # Prix unitaire HT
    taux_tva = db.Column(db.Float, nullable=False, default=20.0)  # Taux TVA en %
    prix_unitaire_ttc = db.Column(db.Float, nullable=False, default=0.0)  # Prix unitaire TTC (calculé)
    prix_total_ht = db.Column(db.Float, nullable=False, default=0.0)  # Prix total HT (PU HT × Qté)
    prix_total_ttc = db.Column(db.Float, nullable=False, default=0.0)  # Prix total TTC (calculé)
    
    # Soft Delete réceptions
    is_deleted = db.Column(db.Boolean, default=False)
    deleted_at = db.Column(AutoDateTime, nullable=True)
    deleted_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    
    item = db.relationship('Item', backref=db.backref('receptions', lazy=True))
    user = db.relationship('User', backref=db.backref('receptions', lazy=True), foreign_keys=[user_id])
    deleter = db.relationship('User', foreign_keys=[deleted_by])

# Modèle pour gérer les stocks FIFO (First In, First Out)
class ReceptionStock(db.Model):
    __tablename__ = 'reception_stocks'
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('items.id'), nullable=False)
    reception_id = db.Column(db.Integer, db.ForeignKey('receptions.id'), nullable=False)
    quantite_initiale = db.Column(db.Float, nullable=False)  # Quantité reçue initialement
    quantite_restante = db.Column(db.Float, nullable=False)  # Quantité non encore dotée
    prix_unitaire_ht = db.Column(db.Float, nullable=False)  # Prix HT de cette réception
    prix_unitaire_ttc = db.Column(db.Float, nullable=False)  # Prix TTC de cette réception
    taux_tva = db.Column(db.Float, nullable=False)  # Taux TVA
    date_reception = db.Column(AutoDateTime, nullable=False)  # Date pour ordre FIFO
    created_at = db.Column(AutoDateTime, default=datetime.utcnow)
    
    # Relations
    item = db.relationship('Item', backref=db.backref('reception_stocks', lazy=True))
    reception = db.relationship('Reception', backref=db.backref('stocks', lazy=True))
    
    @property
    def quantite_utilisee(self):
        """Calcule la quantité déjà utilisée"""
        return self.quantite_initiale - self.quantite_restante
    
    @property
    def taux_utilisation(self):
        """Calcule le taux d'utilisation en %"""
        if self.quantite_initiale > 0:
            return (self.quantite_utilisee / self.quantite_initiale) * 100
        return 0

# Modèle pour les avis d'achat
class AvisAchat(db.Model):
    __tablename__ = 'avis_achats'
    id = db.Column(db.Integer, primary_key=True)
    numero_avis = db.Column(db.String(50), unique=True, nullable=False)  # /2025/PC
    nature_prestation = db.Column(db.String(100), nullable=False)  # Nature de prestation (catégorie)
    objet_prestation = db.Column(db.Text)  # Objet de la prestation
    lieu_execution = db.Column(db.String(200), default='Caserne de la Protection Civile de Sidi Kacem')
    delai_execution = db.Column(db.String(100))  # Délai en jours
    date_limite_reception = db.Column(db.String(100))  # Date et heure limites
    date_creation = db.Column(AutoDateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    statut = db.Column(db.String(20), default='brouillon')  # brouillon, publie, archive
    
    # Soft Delete avis
    is_deleted = db.Column(db.Boolean, default=False)
    deleted_at = db.Column(AutoDateTime, nullable=True)
    deleted_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    
    # Relations
    user = db.relationship('User', backref='avis_achats', foreign_keys=[created_by])
    deleter = db.relationship('User', foreign_keys=[deleted_by])
    items = db.relationship('AvisAchatItem', backref='avis_achat', cascade='all, delete-orphan')
    
    @staticmethod
    def generate_numero_avis():
        """Génère un numéro d'avis unique (/2025/PC)"""
        year = datetime.now().year
        last_avis = AvisAchat.query.filter(
            AvisAchat.numero_avis.like(f'%/{year}/PC')
        ).order_by(AvisAchat.numero_avis.desc()).first()
        
        if last_avis:
            try:
                # Extraire le numéro avant /2025/PC
                numero_part = last_avis.numero_avis.split('/')[0]
                last_number = int(numero_part)
                next_number = last_number + 1
            except (ValueError, IndexError):
                next_number = 1
        else:
            next_number = 1
        
        return f"{next_number:03d}/{year}/PC"

class AvisAchatItem(db.Model):
    __tablename__ = 'avis_achat_items'
    id = db.Column(db.Integer, primary_key=True)
    avis_achat_id = db.Column(db.Integer, db.ForeignKey('avis_achats.id'), nullable=False)
    item_id = db.Column(db.Integer, db.ForeignKey('items.id'), nullable=False)
    quantite = db.Column(db.Integer, nullable=False)
    caracteristiques = db.Column(db.Text)  # Caractéristiques et spécifications
    garanties = db.Column(db.String(100))  # Garanties exigées
    
    # Relations
    item = db.relationship('Item', backref='avis_achat_items')

# Modèle pour le log d'audit
class AuditLog(db.Model):
    __tablename__ = 'audit_logs'
    id = db.Column(db.Integer, primary_key=True)
    action = db.Column(db.String(50), nullable=False)  # 'create', 'update', 'delete'
    entity_type = db.Column(db.String(50), nullable=False)  # 'item', 'reception', 'dotation', etc.
    entity_id = db.Column(db.Integer)
    entity_name = db.Column(db.String(200))  # Nom de l'entité pour référence
    description = db.Column(db.Text)  # Description de l'action
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    user_name = db.Column(db.String(100))  # Nom de l'utilisateur pour référence
    created_at = db.Column(AutoDateTime, default=datetime.utcnow)
    
    user = db.relationship('User', backref=db.backref('audit_logs', lazy=True))

# Fonction helper pour nettoyer les logs d'audit de plus de 5 minutes
def cleanup_old_audit_logs():
    """Supprime automatiquement les logs d'audit de plus de 5 minutes"""
    try:
        # Calculer le seuil de temps (5 minutes avant maintenant)
        time_threshold = datetime.utcnow() - timedelta(minutes=5)
        
        # Supprimer tous les logs plus anciens que 5 minutes
        deleted_count = AuditLog.query.filter(AuditLog.created_at < time_threshold).delete()
        
        if deleted_count > 0:
            db.session.commit()
            print(f"🗑️ {deleted_count} log(s) d'audit supprimé(s) (plus de 5 minutes)")
    except Exception as e:
        db.session.rollback()
        print(f"⚠️ Erreur lors du nettoyage des logs d'audit: {e}")

# Fonction helper pour créer un log d'audit
def create_audit_log(action, entity_type, entity_id=None, entity_name=None, description=None):
    """Crée une entrée dans le log d'audit"""
    user_id = session.get('user_id')
    user_name = session.get('user_name', 'Système')
    
    log = AuditLog(
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        entity_name=entity_name,
        description=description,
        user_id=user_id,
        user_name=user_name
    )
    db.session.add(log)
    db.session.commit()
    
    # Nettoyer automatiquement les anciens logs après chaque création
    cleanup_old_audit_logs()

# Modèles pour le suivi des engagements budgétaires
class BudgetNature(db.Model):
    __tablename__ = 'budget_natures'
    id = db.Column(db.Integer, primary_key=True)
    nature = db.Column(db.String(200), nullable=False)
    # Nouveau champ pour le budget spécifique à une unité/centre
    montant_ttc = db.Column(db.Float, nullable=False, default=0.0)
    
    # Gardés temporairement pour rétrocompatibilité et migration
    budget_centre_ttc = db.Column(db.Float, nullable=False, default=0.0)
    budget_unite_ttc = db.Column(db.Float, nullable=False, default=0.0)
    
    annee = db.Column(db.Integer, nullable=False)  # Année budgétaire
    unite_id = db.Column(db.Integer, db.ForeignKey('unites.id'), nullable=True)  # Lié à une unité OU un centre
    
    created_at = db.Column(AutoDateTime, default=datetime.utcnow)
    updated_at = db.Column(AutoDateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relation
    unite = db.relationship('Unite', backref=db.backref('budget_natures', lazy=True))
    
    @property
    def budget_total(self):
        """Retourne le montant alloué pour cette structure spécifique"""
        return self.montant_ttc
    
    @property
    def budget_consomme(self):
        """Calcule le budget déjà consommé pour cette nature, année et unité spécifique"""
        if not self.unite_id:
            # Fallback legacy si pas d'unite_id : consommation globale (pourquoi pas)
            return db.session.query(db.func.sum(ConsommationBudget.montant_ttc)).filter_by(
                nature=self.nature, annee=self.annee
            ).scalar() or 0.0
            
        return db.session.query(db.func.sum(ConsommationBudget.montant_ttc)).filter(
            ConsommationBudget.nature == self.nature,
            ConsommationBudget.annee == self.annee,
            db.or_(
                ConsommationBudget.centre_id == self.unite_id,
                ConsommationBudget.centre_nom == (self.unite.nom if self.unite else None)
            )
        ).scalar() or 0.0

    @property
    def budget_restant(self):
        """Calcule le budget restant pour cette nature et année"""
        return self.budget_total - self.budget_consomme
    
    @property
    def est_verrouille(self):
        """Vérifie si le budget est épuisé"""
        return self.budget_restant <= 0

    def __repr__(self):
        return f'<BudgetNature {self.nature} - {self.annee} - Unite {self.unite_id}>'

# Modèle pour tracker les consommations budgétaires
class ConsommationBudget(db.Model):
    __tablename__ = 'consommations_budget'
    id = db.Column(db.Integer, primary_key=True)
    nature = db.Column(db.String(200), nullable=False)  # Nature de prestation
    centre_id = db.Column(db.Integer, nullable=True)  # ID du centre (NULL = Unité principale)
    centre_nom = db.Column(db.String(200), nullable=True)  # Nom du centre pour référence
    montant_ttc = db.Column(db.Float, nullable=False)  # Montant consommé TTC
    dotation_id = db.Column(db.Integer, db.ForeignKey('dotations.id'), nullable=True)  # Lié à une dotation
    reception_id = db.Column(db.Integer, db.ForeignKey('receptions.id'), nullable=True)  # Lié à une réception
    annee = db.Column(db.Integer, nullable=False)  # Année budgétaire
    created_at = db.Column(AutoDateTime, default=datetime.utcnow)
    
    # Relations
    dotation = db.relationship('Dotation', backref=db.backref('consommations', lazy=True))
    reception = db.relationship('Reception', backref=db.backref('consommations', lazy=True))
    
    def __repr__(self):
        return f'<ConsommationBudget {self.nature} - {self.montant_ttc} DH>'

class Fournisseur(db.Model):
    __tablename__ = 'fournisseurs'
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(200), nullable=False)
    contact = db.Column(db.String(100))
    email = db.Column(db.String(120))
    telephone = db.Column(db.String(30))
    fax = db.Column(db.String(30))
    adresse = db.Column(db.Text)
    ville = db.Column(db.String(100))
    registre_commerce = db.Column(db.String(100))
    identification_fonciere = db.Column(db.String(100))
    compte_bancaire_ribe = db.Column(db.String(100))
    statut = db.Column(db.String(50), default='En attente', nullable=False)
    delai_moyen_livraison_jours = db.Column(db.Integer, default=0)
    score_performance = db.Column(db.Float, default=0.0)
    notes = db.Column(db.Text)
    created_at = db.Column(AutoDateTime, default=datetime.utcnow)
    updated_at = db.Column(AutoDateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relations
    contacts = db.relationship('Contact', backref='fournisseur', lazy=True, cascade="all, delete-orphan")
    categories = db.relationship('Categorie', secondary='categories_fournisseurs', lazy='subquery',
        backref=db.backref('fournisseurs', lazy=True))

# 1. La table de liaison
categories_fournisseurs = db.Table('categories_fournisseurs',
    db.Column('fournisseur_id', db.Integer, db.ForeignKey('fournisseurs.id'), primary_key=True),
    db.Column('categorie_id', db.Integer, db.ForeignKey('categorie.id'), primary_key=True)
)

# 2. Le nouveau modèle Categorie
class Categorie(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), unique=True, nullable=False)
    parent_id = db.Column(db.Integer, db.ForeignKey('categorie.id'))
    children = db.relationship('Categorie', backref=db.backref('parent', remote_side=[id]), lazy=True)

    def __repr__(self):
        return f'<Categorie {self.nom}>'

# Nouveau modèle
class Contact(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)
    poste = db.Column(db.String(100))
    email = db.Column(db.String(120))
    telephone = db.Column(db.String(20))
    
    # C'est la "clé étrangère" qui relie ce contact au fournisseur
    fournisseur_id = db.Column(db.Integer, db.ForeignKey('fournisseurs.id'), nullable=False)

# Migration pour ajouter les colonnes manquantes
def ensure_extra_columns():
    try:
        # 1. Soft Delete columns
        tables_soft_delete = ['items', 'dotations', 'receptions', 'avis_achats', 'users']
        for table in tables_soft_delete:
            res = db.session.execute(text(f"PRAGMA table_info('{table}')"))
            existing_cols = {row[1] for row in res}
            
            if 'is_deleted' not in existing_cols:
                db.session.execute(text(f"ALTER TABLE {table} ADD COLUMN is_deleted BOOLEAN DEFAULT 0"))
            if 'deleted_at' not in existing_cols:
                db.session.execute(text(f"ALTER TABLE {table} ADD COLUMN deleted_at DATETIME"))
            if 'deleted_by' not in existing_cols:
                db.session.execute(text(f"ALTER TABLE {table} ADD COLUMN deleted_by INTEGER"))

        # 2. unite_id for Users (Domaine de compétence)
        res = db.session.execute(text("PRAGMA table_info('users')"))
        existing_cols = {row[1] for row in res}
        if 'unite_id' not in existing_cols:
            db.session.execute(text("ALTER TABLE users ADD COLUMN unite_id INTEGER REFERENCES unites(id)"))
            
        # 3. Regions Table creation & Column additions
        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS regions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nom VARCHAR(100) UNIQUE NOT NULL
            )
        """))
        
        # Add region_id to users and unites
        for table in ['users', 'unites']:
            res = db.session.execute(text(f"PRAGMA table_info('{table}')"))
            existing_cols = {row[1] for row in res}
            if 'region_id' not in existing_cols:
                db.session.execute(text(f"ALTER TABLE {table} ADD COLUMN region_id INTEGER REFERENCES regions(id)"))
        
        # Populate regions if empty
        res = db.session.execute(text("SELECT COUNT(*) FROM regions"))
        if res.fetchone()[0] == 0:
            regions_morocco = [
                "Tanger-Tétouan-Al Hoceïma", "L'Oriental", "Fès-Meknès", "Rabat-Salé-Kénitra",
                "Béni Mellal-Khénifra", "Casablanca-Settat", "Marrakech-Safi", "Drâa-Tafilalet",
                "Souss-Massa", "Guelmim-Oued Noun", "Laâyoune-Sakia El Hamra", "Dakhla-Oued Ed Dahab"
            ]
            for reg in regions_morocco:
                db.session.execute(text("INSERT INTO regions (nom) VALUES (:nom)"), {"nom": reg})
                
        db.session.commit()
    except Exception as e:
        print(f"⚠️ Erreur lors de la migration des colonnes: {e}")
        db.session.rollback()

# Remplacer ensure_soft_delete_columns par ensure_extra_columns dans le flux
def ensure_soft_delete_columns():
    return ensure_extra_columns()

# Fin de la section migration


# Migration légère pour ajouter les nouvelles colonnes fournisseurs si manquantes (SQLite)
def ensure_supplier_extra_columns():
    try:
        res = db.session.execute(text("PRAGMA table_info('fournisseurs')"))
        existing_cols = {row[1] for row in res}
        alter_statements = []
        if 'fax' not in existing_cols:
            alter_statements.append("ALTER TABLE fournisseurs ADD COLUMN fax VARCHAR(30)")
        if 'ville' not in existing_cols:
            alter_statements.append("ALTER TABLE fournisseurs ADD COLUMN ville VARCHAR(100)")
        if 'registre_commerce' not in existing_cols:
            alter_statements.append("ALTER TABLE fournisseurs ADD COLUMN registre_commerce VARCHAR(100)")
        if 'identification_fonciere' not in existing_cols:
            alter_statements.append("ALTER TABLE fournisseurs ADD COLUMN identification_fonciere VARCHAR(100)")
        if 'compte_bancaire_ribe' not in existing_cols:
            alter_statements.append("ALTER TABLE fournisseurs ADD COLUMN compte_bancaire_ribe VARCHAR(100)")
        for stmt in alter_statements:
            db.session.execute(text(stmt))
        if alter_statements:
            db.session.commit()
    except Exception as e:
        print(f"⚠️ Migration fournisseurs ignorée: {e}")

def ensure_unite_hierarchy_columns():
    """S'assure que les colonnes hierarchy existent dans la table 'unites'"""
    try:
        from sqlalchemy import text
        res = db.session.execute(text("PRAGMA table_info('unites')"))
        existing_cols = {row[1] for row in res}
        
        if 'parent_id' not in existing_cols:
            db.session.execute(text("ALTER TABLE unites ADD COLUMN parent_id INTEGER REFERENCES unites(id)"))
            db.session.commit()
            print("✅ Colonne 'parent_id' ajoutée à la table 'unites'.")
            
        if 'type' not in existing_cols:
            # S'assurer que VARCHAR(50) fonctionne bien avec SQLite
            db.session.execute(text("ALTER TABLE unites ADD COLUMN type VARCHAR(50) DEFAULT 'centre'"))
            db.session.commit()
            print("✅ Colonne 'type' ajoutée à la table 'unites'.")
            
        # Initialisation du type pour les unités existantes
        # On suppose que l'ID 1 (Sidi Kacem) est l'unité provinciale par défaut si non défini
        db.session.execute(text("UPDATE unites SET type = 'unite' WHERE id = 1 AND (type IS NULL OR type = 'centre')"))
        db.session.execute(text("UPDATE unites SET parent_id = 1 WHERE id > 1 AND parent_id IS NULL"))
        db.session.commit()
    except Exception as e:
        print(f"⚠️ Erreur migration hierarchy unites: {e}")
        db.session.rollback()

def ensure_budget_unite_column():
    """S'assure que la colonne unite_id existe dans la table 'budget_natures'"""
    try:
        from sqlalchemy import text
        res = db.session.execute(text("PRAGMA table_info('budget_natures')"))
        existing_cols = {row[1] for row in res}
        if 'unite_id' not in existing_cols:
            db.session.execute(text("ALTER TABLE budget_natures ADD COLUMN unite_id INTEGER REFERENCES unites(id)"))
            db.session.commit()
            print("✅ Colonne 'unite_id' ajoutée à la table 'budget_natures'.")
            
            # Lier les budgets existants à l'unité 1 (Sidi Kacem) par défaut
            db.session.execute(text("UPDATE budget_natures SET unite_id = 1 WHERE unite_id IS NULL"))
            db.session.commit()
    except Exception as e:
        print(f"⚠️ Erreur migration budget_natures unite_id: {e}")
        db.session.rollback()

# ============================================
# FONCTIONS HELPER POUR LA GESTION BUDGÉTAIRE
# ============================================

def calculer_prix_reception(pu_ht, quantite, taux_tva):
    """
    Calcule automatiquement tous les prix d'une réception
    
    Args:
        pu_ht (float): Prix unitaire hors taxe
        quantite (int): Quantité reçue
        taux_tva (float): Taux de TVA en pourcentage (ex: 20)
    
    Returns:
        dict: Dictionnaire contenant pt_ht, pu_ttc, pt_ttc
    """
    pt_ht = pu_ht * quantite
    pu_ttc = pu_ht * (1 + taux_tva / 100)
    pt_ttc = pt_ht * (1 + taux_tva / 100)
    
    return {
        'pt_ht': round(pt_ht, 2),
        'pu_ttc': round(pu_ttc, 2),
        'pt_ttc': round(pt_ttc, 2)
    }

def verifier_budget_disponible(nature, centre_id, montant_ttc, annee):
    """
    Vérifie si le budget est suffisant pour une opération
    
    Args:
        nature (str): Nature de prestation
        centre_id (int): ID du centre (None pour unité principale)
        montant_ttc (float): Montant à consommer
        annee (int): Année budgétaire
    
    Returns:
        dict: {disponible: bool, reste: float, budget_alloue: float, consommation: float, message: str}
    """
    # 1. Récupérer le budget de la nature
    budget = BudgetNature.query.filter_by(nature=nature, annee=annee).first()
    
    if not budget:
        return {
            'disponible': False,
            'reste': 0.0,
            'budget_alloue': 0.0,
            'consommation': 0.0,
            'message': f'⚠️ Aucun budget défini pour "{nature}" en {annee}'
        }
    
    # 2. Calculer la consommation actuelle
    consommation = db.session.query(
        db.func.sum(ConsommationBudget.montant_ttc)
    ).filter_by(
        nature=nature,
        centre_id=centre_id,
        annee=annee
    ).scalar() or 0.0
    
    # 3. Déterminer le budget alloué pour cet ID spécifique
    # On cherche le budget lié à cette unité/centre
    budget_unite = BudgetNature.query.filter_by(
        nature=nature, 
        annee=annee, 
        unite_id=centre_id
    ).first()
    
    if budget_unite:
        budget_alloue = budget_unite.montant_ttc
        type_dest = "centre/unité"
    else:
        # Fallback : Si aucun budget par unité n'est trouvé, on utilise le budget global si unite_id est None dans BudgetNature
        budget_global = BudgetNature.query.filter_by(nature=nature, annee=annee, unite_id=None).first()
        if budget_global:
            # Ancienne logique pour compatibilité temporaire
            unite = Unite.query.get(centre_id) if centre_id else None
            budget_alloue = budget_global.budget_centre_ttc if (unite and unite.type == 'centre') else budget_global.budget_unite_ttc
            type_dest = "global"
        else:
            budget_alloue = 0.0
            type_dest = "inconnu"
    
    # 4. Calculer le reste
    reste = budget_alloue - consommation
    disponible = reste >= montant_ttc
    
    # 5. Générer le message
    if disponible:
        message = f'✅ Budget suffisant pour {type_dest}'
    else:
        message = f'⚠️ BUDGET INSUFFISANT ! Reste: {reste:.2f} DH, Demandé: {montant_ttc:.2f} DH'
    
    return {
        'disponible': disponible,
        'reste': round(reste, 2),
        'budget_alloue': round(budget_alloue, 2),
        'consommation': round(consommation, 2),
        'message': message
    }

def calculer_montant_dotation_ttc(articles_list):
    """
    Calcule le montant TTC total d'une dotation basé sur les prix des réceptions
    
    Args:
        articles_list (list): Liste de tuples (item_id, quantite)
    
    Returns:
        float: Montant total TTC
    """
    montant_total = 0.0
    
    for item_id, quantite in articles_list:
        # Récupérer la dernière réception de cet article pour avoir le prix
        derniere_reception = Reception.query.filter_by(
            item_id=item_id
        ).order_by(Reception.created_at.desc()).first()
        
        if derniere_reception and derniere_reception.prix_unitaire_ttc > 0:
            montant_total += derniere_reception.prix_unitaire_ttc * quantite
    
    return round(montant_total, 2)

class BonCommande(db.Model):
    __tablename__ = 'bons_commande'
    id = db.Column(db.Integer, primary_key=True)
    numero_bc = db.Column(db.String(100), unique=True, nullable=False)
    loi_finance = db.Column(db.String(100), nullable=False)
    nature_prestation = db.Column(db.String(200), nullable=False)
    sous_categorie = db.Column(db.String(200))  # Pour "Entretien et réparation"
    date_engagement = db.Column(AutoDate, nullable=False)
    numero_fiche_navette = db.Column(db.String(100), nullable=False)
    montant_engage = db.Column(db.Float, nullable=False)
    beneficiaire = db.Column(db.String(200), nullable=False)
    
    # Décision contrôle
    decision_controle = db.Column(db.String(50), nullable=False)  # 'Visa' ou 'Retour avec observations'
    observations_controle = db.Column(db.Text)  # Si 'Retour avec observations'
    
    # Champs si Visa
    numero_bordereau_emission = db.Column(db.String(100))
    date_service_fait = db.Column(AutoDate)
    controle_tresorier = db.Column(db.String(50))  # 'Visa' ou 'Retour avec observations'
    observations_tresorier = db.Column(db.Text)  # Si 'Retour avec observations'
    date_reglement = db.Column(AutoDate)
    mode_reglement = db.Column(db.String(50), default='Virement')  # Toujours 'Virement'
    
    # Relations
    budget_nature_id = db.Column(db.Integer, db.ForeignKey('budget_natures.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(AutoDateTime, default=datetime.utcnow)
    updated_at = db.Column(AutoDateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    user = db.relationship('User', backref=db.backref('bons_commande', lazy=True))
    # Fournit l'attribut bon.budget_nature (corrige l'AttributeError)
    budget_nature = db.relationship('BudgetNature', backref=db.backref('bons_commande', lazy=True))
    
    def __repr__(self):
        return f'<BonCommande {self.numero_bc}>'

class IndemnitesDeplacement(db.Model):
    __tablename__ = 'indemnites_deplacement'
    id = db.Column(db.Integer, primary_key=True)
    objet = db.Column(db.String(500), nullable=False)
    periode_deplacement = db.Column(db.String(50), nullable=False)  # '1er Trimestre', '2ème Trimestre', etc.
    loi_finance = db.Column(db.String(100), nullable=False)
    montant_engage = db.Column(db.Float, nullable=False)
    type_engagement = db.Column(db.String(100), nullable=False)  # 'Fiche de navette' ou 'Engagement global'
    numero_engagement = db.Column(db.String(100), nullable=False)  # N° Fiche ou N° Engagement
    
    # Relations
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(AutoDateTime, default=datetime.utcnow)
    updated_at = db.Column(AutoDateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    user = db.relationship('User', backref=db.backref('indemnites_deplacement', lazy=True))
    beneficiaires = db.relationship('BeneficiaireIndemnite', backref='indemnite', lazy=True, cascade='all, delete-orphan')
    
    def __repr__(self):
        return f'<IndemnitesDeplacement {self.objet}>'

class BeneficiaireIndemnite(db.Model):
    __tablename__ = 'beneficiaires_indemnite'
    id = db.Column(db.Integer, primary_key=True)
    indemnite_id = db.Column(db.Integer, db.ForeignKey('indemnites_deplacement.id'), nullable=False)
    nom = db.Column(db.String(100), nullable=False)
    prenom = db.Column(db.String(100), nullable=False)
    grade = db.Column(db.String(100), nullable=False)
    montant = db.Column(db.Float, nullable=False)
    
    def __repr__(self):
        return f'<BeneficiaireIndemnite {self.nom} {self.prenom}>'

# Authentication decorator
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Permission decorator - Vérifie que l'utilisateur n'est PAS un simple "user"
def modification_required(f):
    """Décorateur pour les routes nécessitant des droits de modification (admin ou gestionnaire)"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        
        user = User.query.get(session['user_id'])
        if not user or user.role == 'user':
            flash('❌ Accès refusé. Vous n\'avez que des droits de consultation.', 'error')
            return redirect(url_for('dashboard'))
        
        return f(*args, **kwargs)
    return decorated_function

# Vérifie si l'utilisateur peut gérer un autre utilisateur
def can_manage_user(f):
    """Décorateur pour vérifier si l'utilisateur peut gérer un autre utilisateur"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Vérifier si c'est une requête AJAX
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
                  request.headers.get('Content-Type') == 'application/json'
        
        if 'user_id' not in session:
            if is_ajax:
                return jsonify({'success': False, 'message': 'Non authentifié'}), 401
            return redirect(url_for('login'))
        
        current_user = User.query.get(session['user_id'])
        if not current_user:
            if is_ajax:
                return jsonify({'success': False, 'message': 'Utilisateur non trouvé'}), 404
            flash('❌ Utilisateur non trouvé.', 'error')
            return redirect(url_for('dashboard'))
            
        # Seul le Super Admin peut tout gérer, les Admins et Gestionnaires gèrent ce qu'ils ont créé ou leur domaine
        if not current_user.is_super_admin:
            if current_user.role not in ['admin', 'gestionnaire']:
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Accès refusé.'}), 403
                flash('❌ Accès refusé.', 'error')
                return redirect(url_for('admin_users_list'))
            
            # Récupérer l'ID de l'utilisateur à gérer (peut s'appeler user_id ou user_id_param)
            target_id = kwargs.get('user_id') or kwargs.get('user_id_param')
            user_to_manage = User.query.get(target_id)
            if user_to_manage:
                # On peut toujours se modifier soi-même ou un utilisateur qu'on a créé
                if user_to_manage.id == current_user.id or user_to_manage.created_by == current_user.id:
                    return f(*args, **kwargs)
                
                # Un admin ou gestionnaire peut gérer les gens de son unité/région
                if current_user.role in ['admin', 'gestionnaire']:
                    if current_user.unite_id and user_to_manage.unite_id == current_user.unite_id:
                        return f(*args, **kwargs)
                    if current_user.region_id and user_to_manage.region_id == current_user.region_id:
                        return f(*args, **kwargs)

                if is_ajax:
                    return jsonify({'success': False, 'message': 'Accès refusé à cet utilisateur.'}), 403
                flash('❌ Vous n\'avez pas l\'autorisation de gérer cet utilisateur.', 'error')
                return redirect(url_for('admin_users_list'))
            
        return f(*args, **kwargs)
    return decorated_function

# Admin only decorator - Vérifie que l'utilisateur est admin ou gestionnaire
def admin_or_manager_required(f):
    """Décorateur pour les routes réservées aux admins et gestionnaires"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        
        user = User.query.get(session['user_id'])
        if not user or user.role not in ['admin', 'gestionnaire']:
            flash('❌ Accès refusé. Droits administrateur ou gestionnaire requis.', 'error')
            return redirect(url_for('dashboard'))
        
        return f(*args, **kwargs)
    return decorated_function

# Filtre pour échapper les chaînes JavaScript
def escapejs_filter(value):
    """Filtre pour échapper les caractères spéciaux JavaScript"""
    if value is None:
        return ''
    return (str(value)
            .replace('\\', '\\\\')
            .replace('\"', '\\"')
            .replace("\'", "\\'")
            .replace('\n', '\\n')
            .replace('\r', '\\r')
            .replace('\t', '\\t')
            .replace('</', '<\\/'))

# Ajout du filtre à l'application
app.jinja_env.filters['escapejs'] = escapejs_filter

# Context processor pour rendre des variables disponibles dans tous les templates
@app.context_processor
def inject_global_vars():
    # Récupérer la dernière modification de la base de données
    last_log = AuditLog.query.order_by(AuditLog.created_at.desc()).first()
    last_update = last_log.created_at if last_log else datetime.now()
    
    # Fonction pour compatibilité avec les templates qui utilisent now()
    def get_now():
        return last_update
    
    return {
        'now': get_now,  # Fonction callable pour {{ now() }}
        'current_datetime': last_update,  # Objet datetime direct pour {{ current_datetime }}
        'last_update_time': last_update,
        'csrf_token': generate_csrf  # Fonction pour générer le token CSRF
    }

# Filtre Jinja2 personnalisé pour formater les dates
@app.template_filter('format_date')
def format_date_filter(date_value, format_string='%d/%m/%Y'):
    """Filtre Jinja2 pour formater les dates en gérant les strings"""
    if not date_value:
        return ''
    
    # Si c'est une string, la convertir en datetime
    if isinstance(date_value, str):
        date_value = safe_parse_datetime(date_value)
        if not date_value:
            return ''
    
    # Si c'est un objet datetime, le formater
    if isinstance(date_value, datetime):
        return date_value.strftime(format_string)
    
    return str(date_value)

# Helper function pour récupérer un utilisateur par ID
@app.context_processor
def utility_processor():
    def get_user_by_id(user_id):
        if user_id:
            return User.query.get(user_id)
        return None
    
    # Fonction helper pour formater les dates dans les templates
    def format_datetime(date_value, format_string='%d/%m/%Y %H:%M'):
        """Helper pour formater les dates en gérant les strings"""
        if not date_value:
            return ''
        
        # Si c'est une string, la convertir en datetime
        if isinstance(date_value, str):
            date_value = safe_parse_datetime(date_value)
            if not date_value:
                return ''
        
        # Si c'est un objet datetime, le formater
        if isinstance(date_value, datetime):
            return date_value.strftime(format_string)
        
        return str(date_value)
    
    return dict(
        get_user_by_id=get_user_by_id, 
        safe_parse_datetime=safe_parse_datetime,
        format_datetime=format_datetime
    )

# Compteurs globaux pour la sidebar
@app.context_processor
def inject_sidebar_counts():
    try:
        low_stock_count = Item.query.filter(Item.quantity <= Item.reorder_level).count()
    except Exception:
        low_stock_count = 0
    
    try:
        dotations_en_cours_count = Dotation.query.filter(Dotation.statut == 'en_cours').count()
    except Exception:
        dotations_en_cours_count = 0
    
    # Réceptions des 7 derniers jours
    try:
        seven_days_ago = datetime.utcnow() - timedelta(days=7)
        receptions_7d_count = Reception.query.filter(
            Reception.type == 'reception',
            Reception.created_at >= seven_days_ago
        ).count()
    except Exception:
        receptions_7d_count = 0
    
    # Engagements ouverts (bons de commande non réglés)
    try:
        engagements_ouverts_count = BonCommande.query.filter(
            (BonCommande.etat_reglement.is_(None)) | (BonCommande.etat_reglement != 'Réglé')
        ).count()
    except Exception:
        engagements_ouverts_count = 0
    
    return dict(sidebar_counts={
        'low_stock': low_stock_count,
        'dotations_en_cours': dotations_en_cours_count,
        'receptions_7d': receptions_7d_count,
        'engagements_ouverts': engagements_ouverts_count,
    })


@app.context_processor
def inject_budget_natures():
    """Expose la liste des natures de prestation budgétaires à tous les templates."""
    try:
        rows = db.session.query(BudgetNature.nature).distinct().order_by(BudgetNature.nature).all()
        budget_natures = [r[0] for r in rows if r[0]]
    except Exception:
        budget_natures = []
    return dict(all_budget_natures=budget_natures)

# ============================================================================
# FONCTIONS HELPER POUR LA GESTION FIFO (First In, First Out)
# ============================================================================

def get_available_stock_fifo(item_id):
    """
    Récupère les stocks disponibles pour un article selon l'ordre FIFO.
    Retourne une liste de ReceptionStock triés par date (plus ancien en premier).
    """
    stocks = ReceptionStock.query.filter(
        ReceptionStock.item_id == item_id,
        ReceptionStock.quantite_restante > 0
    ).order_by(ReceptionStock.date_reception.asc()).all()
    
    return stocks

def calculate_dotation_cost_fifo(item_id, quantite_demandee):
    """
    Calcule le coût d'une dotation selon la méthode FIFO.
    Retourne un dictionnaire avec les détails de coût et les stocks utilisés.
    
    Returns:
        {
            'possible': bool,  # Si la dotation est possible
            'stocks_utilises': list,  # Liste des stocks FIFO utilisés
            'montant_total_ht': float,
            'montant_total_ttc': float,
            'quantite_manquante': float  # Si stock insuffisant
        }
    """
    stocks_disponibles = get_available_stock_fifo(item_id)
    
    quantite_restante = quantite_demandee
    montant_total_ht = 0.0
    montant_total_ttc = 0.0
    stocks_utilises = []
    
    for stock in stocks_disponibles:
        if quantite_restante <= 0:
            break
        
        # Quantité à prendre de ce stock
        qte_a_prendre = min(quantite_restante, stock.quantite_restante)
        
        # Calculer le coût
        cout_ht = qte_a_prendre * stock.prix_unitaire_ht
        cout_ttc = qte_a_prendre * stock.prix_unitaire_ttc
        
        stocks_utilises.append({
            'stock_id': stock.id,
            'reception_id': stock.reception_id,
            'quantite': qte_a_prendre,
            'prix_unitaire_ht': stock.prix_unitaire_ht,
            'prix_unitaire_ttc': stock.prix_unitaire_ttc,
            'taux_tva': stock.taux_tva,
            'cout_ht': cout_ht,
            'cout_ttc': cout_ttc,
            'date_reception': stock.date_reception
        })
        
        montant_total_ht += cout_ht
        montant_total_ttc += cout_ttc
        quantite_restante -= qte_a_prendre
    
    return {
        'possible': quantite_restante <= 0,
        'stocks_utilises': stocks_utilises,
        'montant_total_ht': round(montant_total_ht, 2),
        'montant_total_ttc': round(montant_total_ttc, 2),
        'quantite_manquante': max(0, quantite_restante)
    }

def apply_dotation_fifo(item_id, quantite_demandee, dotation_id):
    """
    Applique une dotation selon la méthode FIFO.
    Crée les DotationItem et met à jour les stocks.
    
    Returns:
        {
            'success': bool,
            'dotation_items': list,  # Liste des DotationItem créés
            'montant_total_ht': float,
            'montant_total_ttc': float,
            'message': str
        }
    """
    # Calculer d'abord si c'est possible
    calcul = calculate_dotation_cost_fifo(item_id, quantite_demandee)
    
    if not calcul['possible']:
        return {
            'success': False,
            'dotation_items': [],
            'montant_total_ht': 0.0,
            'montant_total_ttc': 0.0,
            'message': f"Stock insuffisant. Manque: {calcul['quantite_manquante']} unités"
        }
    
    dotation_items = []
    
    for stock_info in calcul['stocks_utilises']:
        # Créer le DotationItem
        dotation_item = DotationItem(
            dotation_id=dotation_id,
            item_id=item_id,
            quantite_dotee=stock_info['quantite'],
            prix_unitaire=stock_info['prix_unitaire_ttc'],  # Pour rétrocompatibilité
            prix_unitaire_ht=stock_info['prix_unitaire_ht'],
            prix_unitaire_ttc=stock_info['prix_unitaire_ttc'],
            taux_tva=stock_info['taux_tva'],
            montant_total_ht=stock_info['cout_ht'],
            montant_total_ttc=stock_info['cout_ttc'],
            reception_stock_id=stock_info['stock_id']
        )
        db.session.add(dotation_item)
        dotation_items.append(dotation_item)
        
        # Mettre à jour le stock FIFO
        stock = ReceptionStock.query.get(stock_info['stock_id'])
        if stock:
            stock.quantite_restante -= stock_info['quantite']
    
    return {
        'success': True,
        'dotation_items': dotation_items,
        'montant_total_ht': calcul['montant_total_ht'],
        'montant_total_ttc': calcul['montant_total_ttc'],
        'message': 'Dotation appliquée avec succès selon FIFO'
    }

def get_item_stock_value_fifo(item_id):
    """
    Calcule la valeur totale du stock d'un article selon FIFO.
    
    Returns:
        {
            'quantite_totale': float,
            'valeur_ht': float,
            'valeur_ttc': float,
            'prix_moyen_ht': float,
            'prix_moyen_ttc': float
        }
    """
    stocks = ReceptionStock.query.filter_by(item_id=item_id).all()
    
    quantite_totale = sum(s.quantite_restante for s in stocks)
    valeur_ht = sum(s.quantite_restante * s.prix_unitaire_ht for s in stocks)
    valeur_ttc = sum(s.quantite_restante * s.prix_unitaire_ttc for s in stocks)
    
    prix_moyen_ht = valeur_ht / quantite_totale if quantite_totale > 0 else 0
    prix_moyen_ttc = valeur_ttc / quantite_totale if quantite_totale > 0 else 0
    
    return {
        'quantite_totale': quantite_totale,
        'valeur_ht': round(valeur_ht, 2),
        'valeur_ttc': round(valeur_ttc, 2),
        'prix_moyen_ht': round(prix_moyen_ht, 2),
        'prix_moyen_ttc': round(prix_moyen_ttc, 2)
    }

# PDF Generation Functions
def create_decharge_pdf(dotation_category, dotation_data, unite_destinataire="", numero_dotation="", service_destinataire=""):
    """
    Generate a 'Bon de sortie' PDF for dotation.
    dotation_category: str -> e.g. 'Fournitures de Bureau'
    dotation_data: list of lists -> table data (header + rows)
    unite_destinataire: str -> Nom de l'unité destinataire
    numero_dotation: str -> Numéro de référence de la dotation
    service_destinataire: str -> Nom du service destinataire
    """

    from datetime import datetime
    from reportlab.platypus import PageTemplate, Frame
    from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate

    # File name
    file_name = f"decharge_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    # Créer un document personnalisé avec footer
    class DechargeDocTemplate(BaseDocTemplate):
        def __init__(self, filename, numero_dotation="", **kwargs):
            BaseDocTemplate.__init__(self, filename, **kwargs)
            self.numero_dotation = numero_dotation
            
        def afterPage(self):
            # Footer avec référence - respecter les marges du layout
            self.canv.saveState()
            self.canv.setFont('Helvetica', 10)
            reference_text = f"Référence : {self.numero_dotation}" if self.numero_dotation else "Référence : DOT-0001-2025"
            self.canv.drawString(10, 15, reference_text)  # Position alignée avec les marges
            self.canv.restoreState()
    
    # Ajuster les marges pour une impression visible et lisible
    doc = DechargeDocTemplate(file_name, numero_dotation=numero_dotation,
                             pagesize=A4, rightMargin=42.52, leftMargin=10, 
                             topMargin=10, bottomMargin=10)  # Plus de marge en bas pour le footer
    
    # Créer un frame pour le contenu principal
    frame = Frame(10, 10, A4[0]-52.52, A4[1]-20, id='normal')  # Marge droite 1,5cm (42.52pt)
    template = PageTemplate(id='main', frames=frame)
    doc.addPageTemplates([template])

    # Styles améliorés
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='ObjectTitle', alignment=TA_CENTER, fontSize=12, spaceAfter=2, leading=12, fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='Beneficiary', alignment=TA_CENTER, fontSize=11, spaceAfter=15, leading=13))
    styles.add(ParagraphStyle(name='RightTextBold', alignment=TA_RIGHT, fontSize=10, spaceAfter=5, fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='RightText', alignment=TA_RIGHT, fontSize=10, spaceAfter=15))
    styles.add(ParagraphStyle(name='LeftText', alignment=TA_LEFT, fontSize=10, spaceAfter=5))
    styles.add(ParagraphStyle(name='CenterText', alignment=TA_CENTER, fontSize=12, spaceAfter=20, fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='WrappedCell', alignment=TA_LEFT, fontSize=9, leading=11, wordWrap='CJK'))
    styles.add(ParagraphStyle(name='WrappedCellHeader', alignment=TA_CENTER, fontSize=10, leading=12, wordWrap='CJK', fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='WrappedCellCenter', alignment=TA_CENTER, fontSize=9, leading=11, wordWrap='CJK'))
    styles.add(ParagraphStyle(name='CenterTextBold', alignment=TA_CENTER, fontSize=10, spaceAfter=5, fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='CenterTextNormal', alignment=TA_CENTER, fontSize=10, spaceAfter=15))

    elements = []

    # --- HEADER (three images side by side: enteteFR.png, logo.png, enteteAR.png) ---
    # Positionnement symétrique : enteteFR (gauche) sans décalage, enteteAR (droite) avec compensation
    img1 = Image("images/enteteFR.png", width=200, height=70)
    img2 = Image("images/logo.png", width=1.5*cm, height=1.5*cm)
    img3 = Image("images/enteteAR.png", width=200, height=70)

    # Calculer la largeur totale disponible (largeur A4 - marges)
    total_width = A4[0] - 52.52  # Nouvelles marges : gauche (10) + droite (42.52)
    # Répartir la largeur : 40% - 20% - 40%
    col1_width = total_width * 0.4
    col2_width = total_width * 0.2  
    col3_width = total_width * 0.4
    
    header_table = Table([[img1, img2, img3]], colWidths=[col1_width, col2_width, col3_width])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (0, 0), 0*cm),   # Pas de décalage pour enteteFR (parfaitement visible)
        ('RIGHTPADDING', (2, 0), (2, 0), -1.15*cm), # Décalage négatif pour compenser la marge droite plus large
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 10))

    # --- "Modèle : MAT 32" aligné par rapport à la marge droite de la page ---
    model_paragraph = Paragraph("<b>Modèle : MAT 32</b>", styles["RightTextBold"])
    elements.append(model_paragraph)
    elements.append(Spacer(1, 10))

    # --- Objet centré en gras sur deux lignes ---
    object_line1 = "<b>Objet : BON DE SORTIE RELATIVE À LA DOTATION EN</b>"
    object_line2 = f"<b>{dotation_category.upper()}</b>"
    elements.append(Paragraph(object_line1, styles["ObjectTitle"]))
    elements.append(Paragraph(object_line2, styles["ObjectTitle"]))

    # --- Bénéficiaire centré ---
    if service_destinataire:
        beneficiary_text = f"Bénéficiaire : {service_destinataire}"
    else:
        beneficiary_text = f"Bénéficiaire : {unite_destinataire}"
    elements.append(Paragraph(beneficiary_text, styles["Beneficiary"]))

    # --- Tableau avec largeurs optimisées (uniquement l'en-tête) ---
    table_data_wrapped = []
    if dotation_data:
        header_row = [Paragraph(str(cell), styles["WrappedCellHeader"]) for cell in dotation_data[0]]
        table_data_wrapped.append(header_row)
        for row in dotation_data[1:]:
            row_cells = [
                Paragraph(str(row[0]), styles["WrappedCellCenter"]),
                Paragraph(str(row[1]), styles["WrappedCell"]),
                Paragraph(str(row[2]), styles["WrappedCellCenter"]),
                Paragraph(str(row[3]), styles["WrappedCellCenter"]),
            ]
            table_data_wrapped.append(row_cells)

    # Largeurs de colonnes redimensionnées : N°(1cm), Désignations (auto), Unité(2.5cm), Quantité(2cm)
    col_widths = [1*cm, 12*cm, 2.5*cm, 2*cm]
    
    table = Table(table_data_wrapped, repeatRows=1, hAlign='CENTER', colWidths=col_widths)
    table_styles = [
        ('GRID', (0, 0), (-1, -1), 0.7, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
    ]
    if len(table_data_wrapped) > 1:
        table_styles += [
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('VALIGN', (0, 1), (0, -1), 'MIDDLE'),
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),
            ('VALIGN', (2, 1), (2, -1), 'MIDDLE'),
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),
            ('VALIGN', (3, 1), (3, -1), 'MIDDLE'),
        ]
    table.setStyle(TableStyle(table_styles))

    elements.append(table)
    elements.append(Spacer(1, 15))

    # --- Date et lieu alignés par rapport à la marge droite de la page ---
    current_date = datetime.now().strftime('%d/%m/%Y')
    date_text = f"Sidi Kacem, le {current_date}"
    date_paragraph = Paragraph(date_text, styles["RightText"])
    elements.append(date_paragraph)
    elements.append(Spacer(1, 15))

    # --- Signature centrée ---
    elements.append(Paragraph("SIGNATURES", styles["CenterText"]))

    # --- Tableau des signatures ---
    signature_data = [['Le prenant :', 'Opérateur GID :', 'Visa du chef hiérarchique :']]
    signature_table = Table(signature_data, colWidths=[6*cm, 6*cm, 6*cm])
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 40),  # Espace pour cachet et signature
        ('TOPPADDING', (0, 0), (-1, -1), 10),
    ]))

    elements.append(signature_table)

    # --- Build PDF ---
    doc.build(elements)

    return file_name

# Routes
# ============================================================================
# INTERFACE D'ADMINISTRATION PERSONNALISÉE
# ============================================================================
# Flask-Admin supprimé - Utilisation de l'interface personnalisée /admin-custom


# ============================================================================
# ROUTES
# ============================================================================

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))

@app.route('/admin/users')
@app.route('/admin/users/')
@login_required
def admin_users_list():
    """Page de gestion des utilisateurs"""
    current_user = User.query.get(session['user_id'])
    
    if not current_user:
        flash('❌ Utilisateur non trouvé.', 'error')
        return redirect(url_for('dashboard'))
    
    # Paramètre pour afficher les utilisateurs supprimés (SUPER ADMIN uniquement)
    show_deleted = request.args.get('show_deleted', 'false') == 'true'
    
    # Super Admin voit tous les utilisateurs
    if current_user.is_super_admin:
        if show_deleted:
            # Afficher uniquement les utilisateurs supprimés
            users = User.query.filter_by(is_deleted=True).order_by(User.deleted_at.desc()).all()
        else:
            # Afficher uniquement les utilisateurs actifs
            users = User.query.filter_by(is_deleted=False).order_by(User.created_at.desc()).all()
    # Les admins voient les utilisateurs qu'ils ont créés OU ceux de leur domaine (si restreints)
    elif current_user.role == 'admin':
        if current_user.unite_id:
            users = User.query.filter(
                (User.unite_id == current_user.unite_id) | (User.created_by == current_user.id) | (User.id == current_user.id),
                User.is_deleted == False
            ).order_by(User.created_at.desc()).all()
        elif current_user.region_id:
            users = User.query.filter(
                (User.region_id == current_user.region_id) | (User.created_by == current_user.id) | (User.id == current_user.id),
                User.is_deleted == False
            ).order_by(User.created_at.desc()).all()
        else:
            users = User.query.filter(
                (User.created_by == current_user.id) | (User.id == current_user.id),
                User.is_deleted == False
            ).order_by(User.created_at.desc()).all()
    # Les gestionnaires voient les utilisateurs qu'ils ont créés OU ceux de leur domaine
    elif current_user.role == 'gestionnaire':
        if current_user.unite_id:
            users = User.query.filter(
                (User.unite_id == current_user.unite_id) | (User.created_by == current_user.id) | (User.id == current_user.id),
                User.is_deleted == False
            ).order_by(User.created_at.desc()).all()
        elif current_user.region_id:
            users = User.query.filter(
                (User.region_id == current_user.region_id) | (User.created_by == current_user.id) | (User.id == current_user.id),
                User.is_deleted == False
            ).order_by(User.created_at.desc()).all()
        else:
            users = User.query.filter(
                (User.created_by == current_user.id) | (User.id == current_user.id),
                User.is_deleted == False
            ).order_by(User.created_at.desc()).all()
    else:
        flash('❌ Accès refusé. Vous n\'avez pas les droits nécessaires.', 'error')
        return redirect(url_for('dashboard'))
    
    # Compter les utilisateurs supprimés (pour SUPER ADMIN)
    deleted_count = 0
    if current_user.is_super_admin:
        deleted_count = User.query.filter_by(is_deleted=True).count()
        
    # Calculer les statistiques basées sur les utilisateurs visibles pour l'administrateur
    stats = {
        'admin': sum(1 for u in users if u.role == 'admin'),
        'gestionnaire': sum(1 for u in users if u.role == 'gestionnaire'),
        'user': sum(1 for u in users if u.role == 'user'),
        'invite': sum(1 for u in users if u.role == 'invite')
    }
    
    return render_template('admin_users.html', 
                         users=users, 
                         current_user=current_user,
                         show_deleted=show_deleted,
                         deleted_count=deleted_count,
                         stats=stats)

@app.route('/admin/users/create', methods=['GET', 'POST'])
@login_required
def create_user():
    """Créer un nouvel utilisateur - Admin, Super Admin ou Gestionnaire"""
    current_user = User.query.get(session['user_id'])
    if not current_user or current_user.role not in ['admin', 'gestionnaire'] and not current_user.is_super_admin:
        flash('❌ Accès refusé. Vous n\'avez pas les droits nécessaires pour gérer les utilisateurs.', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        email = (request.form.get('email') or '').strip().lower()
        role = request.form.get('role', 'user')
        password = request.form.get('password')
        permissions = request.form.getlist('permissions')
        unite_id = request.form.get('unite_id')
        region_id = request.form.get('region_id')

        if not all([name, email, role, password]):
            flash('❌ Tous les champs sont requis.', 'error')
            unites_query = Unite.query.filter_by(type='unite')
            regions_query = Region.query
            if current_user.role == 'gestionnaire':
                if current_user.unite_id: unites_query = unites_query.filter_by(id=current_user.unite_id)
                elif current_user.region_id: unites_query = unites_query.filter_by(region_id=current_user.region_id)
                regions_query = regions_query.filter(Region.id == -1)
            unites = unites_query.order_by(Unite.nom).all()
            regions = regions_query.order_by(Region.nom).all()
            return render_template('admin_create_user.html', current_user=current_user, unites=unites, regions=regions, form_data=request.form)

        if len(password) < 8:
            flash('❌ Le mot de passe doit contenir au moins 8 caractères.', 'error')
            unites_query = Unite.query.filter_by(type='unite')
            regions_query = Region.query
            if current_user.role == 'gestionnaire':
                if current_user.unite_id: unites_query = unites_query.filter_by(id=current_user.unite_id)
                elif current_user.region_id: unites_query = unites_query.filter_by(region_id=current_user.region_id)
                regions_query = regions_query.filter(Region.id == -1)
            unites = unites_query.order_by(Unite.nom).all()
            regions = regions_query.order_by(Region.nom).all()
            return render_template('admin_create_user.html', current_user=current_user, unites=unites, regions=regions, form_data=request.form)

        if role not in ['admin', 'gestionnaire', 'user', 'invite']:
            flash('❌ Rôle invalide.', 'error')
            unites_query = Unite.query.filter_by(type='unite')
            regions_query = Region.query
            if current_user.role in ['admin', 'gestionnaire']:
                if current_user.unite_id: unites_query = unites_query.filter_by(id=current_user.unite_id)
                elif current_user.region_id: unites_query = unites_query.filter_by(region_id=current_user.region_id)
                regions_query = regions_query.filter(Region.id == -1)
            unites = unites_query.order_by(Unite.nom).all()
            regions = regions_query.order_by(Region.nom).all()
            return render_template('admin_create_user.html', current_user=current_user, unites=unites, regions=regions, form_data=request.form)

        # Restriction hiérarchique : seul un Admin ou Super Admin peut créer un autre Admin
        if role == 'admin' and current_user.role == 'gestionnaire':
            flash('❌ Accès refusé. Un gestionnaire ne peut pas créer de compte Administrateur.', 'error')
            unites_query = Unite.query.filter_by(type='unite')
            if current_user.unite_id: unites_query = unites_query.filter_by(id=current_user.unite_id)
            elif current_user.region_id: unites_query = unites_query.filter_by(region_id=current_user.region_id)
            unites = unites_query.order_by(Unite.nom).all()
            return render_template('admin_create_user.html', current_user=current_user, unites=unites, regions=[], form_data=request.form)

        if not email.endswith('@protectioncivile.ma') and not email.endswith('@protection.com'):
            flash('❌ L\'email doit être du domaine @protectioncivile.ma ou @protection.com', 'error')
            unites_query = Unite.query.filter_by(type='unite')
            regions_query = Region.query
            if current_user.role == 'gestionnaire':
                if current_user.unite_id: unites_query = unites_query.filter_by(id=current_user.unite_id)
                elif current_user.region_id: unites_query = unites_query.filter_by(region_id=current_user.region_id)
                regions_query = regions_query.filter(Region.id == -1)
            unites = unites_query.order_by(Unite.nom).all()
            regions = regions_query.order_by(Region.nom).all()
            return render_template('admin_create_user.html', current_user=current_user, unites=unites, regions=regions, form_data=request.form)

        existing_user = User.query.filter_by(email=email).first()
        if existing_user:
            if existing_user.is_deleted:
                msg = f'❌ Cet email ({email}) est déjà utilisé par un compte se trouvant dans la corbeille. Veuillez restaurer le compte ou contacter un Super Admin.'
            else:
                msg = f'❌ Cet email ({email}) est déjà utilisé dans le système. Si vous ne voyez pas cet utilisateur dans votre liste, c\'est qu\'il appartient à une autre unité ou région.'
            
            flash(msg, 'error')
            unites_query = Unite.query.filter_by(type='unite')
            regions_query = Region.query
            regions_query = Region.query
            if current_user.role in ['admin', 'gestionnaire']:
                if current_user.unite_id:
                    unites_query = unites_query.filter_by(id=current_user.unite_id)
                    regions_query = regions_query.filter(Region.id == -1)
                elif current_user.region_id:
                    unites_query = unites_query.filter_by(region_id=current_user.region_id)
                    regions_query = regions_query.filter(Region.id == current_user.region_id)
            unites = unites_query.order_by(Unite.nom).all()
            regions = regions_query.order_by(Region.nom).all()
            return render_template('admin_create_user.html', current_user=current_user, unites=unites, regions=regions, form_data=request.form)

        # Validation domaine de compétence selon l'utilisateur créateur
        if not current_user.is_super_admin and current_user.role in ['admin', 'gestionnaire']:
            # Si l'admin/gestionnaire tente d'assigner une région
            if region_id:
                if not current_user.region_id or str(region_id) != str(current_user.region_id):
                    if current_user.role == 'gestionnaire':
                        flash('❌ Accès refusé. Un gestionnaire ne peut pas attribuer une compétence régionale.', 'error')
                    else:
                        flash('❌ Accès refusé. Vous ne pouvez attribuer que votre propre région.', 'error')
                    
                    unites_query = Unite.query.filter_by(type='unite')
                    if current_user.unite_id: unites_query = unites_query.filter_by(id=current_user.unite_id)
                    elif current_user.region_id: unites_query = unites_query.filter_by(region_id=current_user.region_id)
                    unites = unites_query.order_by(Unite.nom).all()
                    
                    regions_query = Region.query.filter(Region.id == -1)
                    if current_user.role == 'admin' and current_user.region_id:
                        regions_query = Region.query.filter_by(id=current_user.region_id)
                    regions = regions_query.all()
                    
                    return render_template('admin_create_user.html', current_user=current_user, unites=unites, regions=regions, form_data=request.form)
            
            # Validation Unité
            if unite_id:
                if current_user.unite_id and str(unite_id) != str(current_user.unite_id):
                    flash('❌ Vous ne pouvez gérer que votre propre unité provinciale.', 'error')
                    unites_query = Unite.query.filter_by(id=current_user.unite_id)
                    unites = unites_query.all()
                    return render_template('admin_create_user.html', current_user=current_user, unites=unites, regions=[], form_data=request.form)
                elif current_user.region_id:
                    target_unite = Unite.query.get(unite_id)
                    if not target_unite or str(target_unite.region_id) != str(current_user.region_id):
                        flash('❌ Vous ne pouvez gérer que les unités de votre région.', 'error')
                        unites_query = Unite.query.filter_by(region_id=current_user.region_id)
                        unites = unites_query.order_by(Unite.nom).all()
                        return render_template('admin_create_user.html', current_user=current_user, unites=unites, regions=[], form_data=request.form)

        new_user = User(
            name=name,
            email=email,
            role=role,
            created_by=current_user.id,
            is_super_admin='is_super_admin' in request.form if current_user.is_super_admin else False,
            is_active=True,
            unite_id=unite_id if unite_id else None,
            region_id=region_id if region_id else None
        )
        new_user.set_password(password)
        new_user.set_permissions(permissions)

        try:
            if unite_id:
                target_unite = Unite.query.get(unite_id)
                if target_unite and target_unite.type == 'centre':
                    flash('❌ Erreur : Un utilisateur ne peut pas avoir une compétence de niveau Centre. Veuillez choisir une Unité Provinciale.', 'error')
                    return redirect(url_for('create_user'))

            db.session.add(new_user)
            db.session.commit()

            create_audit_log(
                action='create',
                entity_type='user',
                entity_id=new_user.id,
                entity_name=new_user.name,
                description=f"Création de l'utilisateur {new_user.name} ({email}) avec le rôle {role}"
            )

            flash(f'Utilisateur {new_user.name} créé avec succès.', 'success')
            return redirect(url_for('admin_users_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Erreur lors de la création: {str(e)}', 'error')
            unites_query = Unite.query.filter_by(type='unite')
            regions_query = Region.query
            regions_query = Region.query
            if current_user.role in ['admin', 'gestionnaire']:
                if current_user.unite_id:
                    unites_query = unites_query.filter_by(id=current_user.unite_id)
                    regions_query = regions_query.filter(Region.id == -1)
                elif current_user.region_id:
                    unites_query = unites_query.filter_by(region_id=current_user.region_id)
                    regions_query = regions_query.filter(Region.id == current_user.region_id)
            unites = unites_query.order_by(Unite.nom).all()
            regions = regions_query.order_by(Region.nom).all()
            return render_template('admin_create_user.html', current_user=current_user, unites=unites, regions=regions, form_data=request.form)
            
    unites_query = Unite.query.filter_by(type='unite')
    regions_query = Region.query

    if current_user.role in ['admin', 'gestionnaire']:
        if current_user.unite_id:
            unites_query = unites_query.filter_by(id=current_user.unite_id)
            if current_user.region_id:
                regions_query = regions_query.filter_by(id=current_user.region_id)
            else:
                regions_query = regions_query.filter(Region.id == -1)
        elif current_user.region_id:
            unites_query = unites_query.filter_by(region_id=current_user.region_id)
            if current_user.role == 'admin':
                regions_query = regions_query.filter_by(id=current_user.region_id)
            else:
                regions_query = regions_query.filter(Region.id == -1)
        # If no restriction, they see everything (global admins)

    unites = unites_query.order_by(Unite.nom).all()
    regions = regions_query.order_by(Region.nom).all()
    return render_template('admin_create_user.html', current_user=current_user, unites=unites, regions=regions)

@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@can_manage_user
def edit_user(user_id):
    """Éditer un utilisateur - Super Admin ou l'admin qui l'a créé"""
    current_user = User.query.get(session['user_id'])
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        unite_id = request.form.get('unite_id') or None
        region_id = request.form.get('region_id') or None

        # Validation domaine de compétence selon l'utilisateur modificateur
        if not current_user.is_super_admin and current_user.role in ['admin', 'gestionnaire']:
            # Si l'admin/gestionnaire tente d'assigner une région
            if region_id:
                if not current_user.region_id or str(region_id) != str(current_user.region_id):
                    if current_user.role == 'gestionnaire':
                        flash('❌ Accès refusé. Un gestionnaire ne peut pas attribuer une compétence régionale.', 'error')
                    else:
                        flash('❌ Accès refusé. Vous ne pouvez attribuer que votre propre région.', 'error')
                    return redirect(url_for('edit_user', user_id=user_id))
            
            # Validation Unité
            if unite_id:
                if current_user.unite_id and str(unite_id) != str(current_user.unite_id):
                    flash('❌ Vous ne pouvez gérer que votre propre unité provinciale.', 'error')
                    return redirect(url_for('edit_user', user_id=user_id))
                elif current_user.region_id:
                    target_unite = Unite.query.get(unite_id)
                    if not target_unite or str(target_unite.region_id) != str(current_user.region_id):
                        flash('❌ Vous ne pouvez gérer que les unités de votre région.', 'error')
                        return redirect(url_for('edit_user', user_id=user_id))

        user.name = request.form.get('name')
        user.email = (request.form.get('email') or '').strip().lower()
        user.role = request.form.get('role', 'user')
        user.unite_id = unite_id
        user.region_id = region_id

        if user.unite_id:
            target_unite = Unite.query.get(user.unite_id)
            if target_unite and target_unite.type == 'centre':
                flash('❌ Erreur : Un utilisateur ne peut pas avoir une compétence de niveau Centre.', 'error')
                return redirect(url_for('edit_user', user_id=user_id))
        
        # Super Admin only: toggle super admin and active status
        if current_user.is_super_admin:
            # Sécurité: on ne peut pas se retirer ses propres droits Super Admin ici 
            # (pour éviter de se bloquer, mieux vaut passer par un autre Super Admin)
            if user.id != current_user.id:
                user.is_super_admin = 'is_super_admin' in request.form
            
            user.is_active = 'is_active' in request.form
        
        # Email domain validation
        if not user.email.endswith('@protectioncivile.ma') and not user.email.endswith('@protection.com'):
             flash('❌ L\'email doit être du domaine @protectioncivile.ma ou @protection.com', 'error')
             unites_query = Unite.query.filter_by(type='unite')
             regions_query = Region.query
             if current_user.role in ['admin', 'gestionnaire']:
                 if current_user.unite_id:
                     unites_query = unites_query.filter_by(id=current_user.unite_id)
                     regions_query = regions_query.filter(Region.id == -1)
                 elif current_user.region_id:
                     unites_query = unites_query.filter_by(region_id=current_user.region_id)
                     if current_user.role == 'admin':
                        regions_query = regions_query.filter_by(id=current_user.region_id)
                     else:
                        regions_query = regions_query.filter(Region.id == -1)
             unites = unites_query.order_by(Unite.nom).all()
             regions = regions_query.order_by(Region.nom).all()
             return render_template('admin_edit_user.html', user=user, current_user=current_user, user_permissions=user.get_permissions(), unites=unites, regions=regions)
        
        permissions = request.form.getlist('permissions')
        user.set_permissions(permissions)
        
        new_password = request.form.get('password')
        if new_password:
            if len(new_password) < 8:
                flash('❌ Le mot de passe doit contenir au moins 8 caractères.', 'error')
                unites_query = Unite.query.filter_by(type='unite')
                regions_query = Region.query
                if current_user.role in ['admin', 'gestionnaire']:
                    if current_user.unite_id:
                        unites_query = unites_query.filter_by(id=current_user.unite_id)
                        regions_query = regions_query.filter(Region.id == -1)
                    elif current_user.region_id:
                        unites_query = unites_query.filter_by(region_id=current_user.region_id)
                        if current_user.role == 'admin':
                            regions_query = regions_query.filter_by(id=current_user.region_id)
                        else:
                            regions_query = regions_query.filter(Region.id == -1)
                unites = unites_query.order_by(Unite.nom).all()
                regions = regions_query.order_by(Region.nom).all()
                return render_template('admin_edit_user.html', user=user, current_user=current_user, user_permissions=user.get_permissions(), unites=unites, regions=regions)
            user.set_password(new_password)
        
        try:
            db.session.commit()
            create_audit_log(
                action='update',
                entity_type='user',
                entity_id=user.id,
                entity_name=user.name,
                description=f"Modification de l'utilisateur {user.name}"
            )
            flash(f'Utilisateur {user.name} modifié avec succès.', 'success')
            return redirect(url_for('admin_users_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Erreur lors de la modification: {str(e)}', 'error')
            unites_query = Unite.query.filter_by(type='unite')
            regions_query = Region.query
            if current_user.role in ['admin', 'gestionnaire']:
                if current_user.unite_id:
                    unites_query = unites_query.filter_by(id=current_user.unite_id)
                    regions_query = regions_query.filter(Region.id == -1)
                elif current_user.region_id:
                    unites_query = unites_query.filter_by(region_id=current_user.region_id)
                    if current_user.role == 'admin':
                        regions_query = regions_query.filter_by(id=current_user.region_id)
                    else:
                        regions_query = regions_query.filter(Region.id == -1)
            unites = unites_query.order_by(Unite.nom).all()
            regions = regions_query.order_by(Region.nom).all()
            return render_template('admin_edit_user.html', user=user, current_user=current_user, user_permissions=user.get_permissions(), unites=unites, regions=regions)
    
    unites_query = Unite.query.filter_by(type='unite')
    regions_query = Region.query

    if current_user.role in ['admin', 'gestionnaire']:
        if current_user.unite_id:
            unites_query = unites_query.filter_by(id=current_user.unite_id)
            if current_user.region_id:
                regions_query = regions_query.filter_by(id=current_user.region_id)
            else:
                regions_query = regions_query.filter(Region.id == -1)
        elif current_user.region_id:
            unites_query = unites_query.filter_by(region_id=current_user.region_id)
            if current_user.role == 'admin':
                regions_query = regions_query.filter_by(id=current_user.region_id)
            else:
                regions_query = regions_query.filter(Region.id == -1)

    unites = unites_query.order_by(Unite.nom).all()
    regions = regions_query.order_by(Region.nom).all()
    return render_template('admin_edit_user.html', user=user, current_user=current_user, user_permissions=user.get_permissions(), unites=unites, regions=regions)

@app.route('/admin/users/delete/<int:user_id_param>', methods=['POST'])
@login_required
@can_manage_user
def delete_user(user_id_param=None):
    """
    Supprimer un utilisateur (soft delete) - Super Admin ou l'admin qui l'a créé
    """
    # Récupérer l'utilisateur connecté
    current_user = User.query.get(session['user_id'])
    
    # Vérifier si c'est une requête AJAX
    is_ajax = request.headers.get('Content-Type') == 'application/json'
    
    if user_id_param is None:
        if is_ajax:
            return jsonify({'success': False, 'message': 'ID utilisateur manquant'}), 400
        return redirect(url_for('admin_users_list'))
        
    user = User.query.get_or_404(user_id_param)
    
    # Vérifier si l'utilisateur essaie de se supprimer lui-même
    if user.id == current_user.id:
        if is_ajax:
            return jsonify({'success': False, 'message': 'Vous ne pouvez pas supprimer votre propre compte'}), 403
        flash('❌ Vous ne pouvez pas supprimer votre propre compte', 'error')
        return redirect(url_for('admin_users_list'))
    
    try:
        # Soft Delete : marquer comme supprimé au lieu de supprimer définitivement
        user.is_deleted = True
        user.deleted_at = datetime.utcnow()
        user.deleted_by = current_user.id
        db.session.commit()
        
        # Créer un log d'audit
        create_audit_log(
            action='delete',
            entity_type='user',
            entity_id=user.id,
            entity_name=user.name,
            description=f"Suppression (soft delete) de l'utilisateur {user.name} ({user.email}) par {current_user.name}"
        )
        
        return jsonify({
            'success': True,
            'message': f'Utilisateur "{user.name}" supprimé avec succès !'
        })
    except Exception as e:
        db.session.rollback()
        print(f"ERREUR DE SUPPRESSION: {str(e)}")  # Log pour débogage
        import traceback
        traceback.print_exc()  # Afficher la trace complète
        return jsonify({
            'success': False,
            'message': f'Erreur lors de la suppression: {str(e)}'
        }), 500

@app.route('/admin/unites')
@login_required
@admin_or_manager_required
def admin_unites():
    """Liste des unités et centres"""
    user = User.query.get(session['user_id'])
    
    # Domaine de compétence (Admins restreints ou Gestionnaires)
    if user.role in ['admin', 'gestionnaire'] and (user.unite_id or user.region_id):
        if user.unite_id:
            competence = Unite.query.get(user.unite_id)
            if not competence or competence.type == 'centre':
                flash("Accès refusé. Le niveau centre ne peut pas gérer les structures.", 'warning')
                return redirect(url_for('dashboard'))
            
            unites_principales = [competence]
            toutes_les_unites = [competence] + list(competence.children)
        elif user.region_id:
            # Compétence régionale
            unites_principales = Unite.query.filter_by(region_id=user.region_id, type='unite').all()
            toutes_les_unites = []
            for u in unites_principales:
                toutes_les_unites.append(u)
                toutes_les_unites.extend(u.children)
    else:
        # Admin Global ou Gestionnaire sans restriction (si autorisé)
        unites_principales = Unite.query.filter_by(type='unite').all()
        toutes_les_unites = Unite.query.all()
        
    if user.role in ['admin', 'gestionnaire'] and (user.unite_id or user.region_id):
        if user.unite_id:
            regions_count = 0
        elif user.region_id:
            regions_count = 1
        else:
            regions_count = 0
    else:
        regions_count = db.session.query(func.count(Region.id)).scalar()
    return render_template('admin_unites.html', unites_principales=unites_principales, toutes_les_unites=toutes_les_unites, regions_count=regions_count)

@app.route('/admin/unites/ajouter', methods=['GET', 'POST'])
@login_required
@admin_or_manager_required
def add_unite():
    """Ajouter une unité ou un centre"""
    user = User.query.get(session['user_id'])
    
    # Vérification initiale pour utilisateurs restreints : le niveau centre ne gère rien
    if user.role in ['admin', 'gestionnaire'] and (user.unite_id or user.region_id):
        if user.unite_id:
            competence = Unite.query.get(user.unite_id)
            if not competence or competence.type == 'centre':
                flash("Accès refusé. Le niveau centre ne peut pas gérer les structures.", 'warning')
                return redirect(url_for('dashboard'))

    if request.method == 'POST':
        nom = request.form.get('nom')
        type_unite = request.form.get('type')
        parent_id = request.form.get('parent_id')
        region_id = request.form.get('region_id')
        description = request.form.get('description')
        services_liste = request.form.get('services_liste', '').split('\n')
        
        if not nom or not type_unite:
            flash('Le nom et le type sont obligatoires.', 'danger')
            return redirect(url_for('add_unite'))
            
        # Vérification domaine de compétence pour utilisateurs restreints
        if user.role in ['admin', 'gestionnaire'] and (user.unite_id or user.region_id):
            if user.unite_id:
                competence = Unite.query.get(user.unite_id)
                if not competence or competence.type == 'centre':
                    flash("Votre domaine de compétence ne vous permet pas de créer des sous-unités.", 'danger')
                    return redirect(url_for('admin_unites'))
                
                # Un gestionnaire/admin d'une unité provinciale ne peut créer que des centres sous CETTE unité
                if type_unite != 'centre' or not parent_id or int(parent_id) != user.unite_id:
                    flash(f"Vous ne pouvez créer que des centres rattachés à votre unité ({competence.nom}).", 'danger')
                    return redirect(url_for('admin_unites'))
            elif user.region_id:
                # Gestionnaire ou Admin régional
                if type_unite == 'unite' and user.role == 'gestionnaire':
                     # Un gestionnaire ne peut pas créer d'unités provinciales
                     flash("Seul un Administrateur peut créer de nouvelles Unités Provinciales.", 'danger')
                     return redirect(url_for('admin_unites'))
                
                # S'il crée un centre, il doit être rattaché à une unité de sa région
                if parent_id:
                    p_unit = Unite.query.get(parent_id)
                    if not p_unit or p_unit.region_id != user.region_id:
                        flash("Le centre doit être rattaché à une unité de votre région.", 'danger')
                        return redirect(url_for('admin_unites'))
                
                # S'il crée une unité (cas Admin régional), elle doit être forcée dans sa région
                if type_unite == 'unite':
                    region_id = user.region_id
            
        nouvelle_unite = Unite(
            nom=nom,
            type=type_unite,
            parent_id=parent_id if parent_id and type_unite == 'centre' else None,
            region_id=region_id if region_id and type_unite == 'unite' else None,
            description=description
        )
        
        # Si c'est un centre, il hérite automatiquement de la région de son parent
        if type_unite == 'centre' and nouvelle_unite.parent_id:
             parent = Unite.query.get(nouvelle_unite.parent_id)
             if parent:
                 nouvelle_unite.region_id = parent.region_id
        
        try:
            db.session.add(nouvelle_unite)
            db.session.flush() # Pour avoir l'ID
            
            # Ajouter les services si fournis
            for s_nom in services_liste:
                s_nom = s_nom.strip()
                if s_nom:
                    nouveau_service = Service(
                        nom=s_nom,
                        unite_id=nouvelle_unite.id
                    )
                    db.session.add(nouveau_service)
                    
            db.session.commit()
            flash(f'Unité "{nom}" ajoutée avec succès.', 'success')
            return redirect(url_for('admin_unites'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de l\'ajout : {str(e)}', 'danger')
            
    # Filtrage des unités parentes pour le dropdown selon le domaine
    if user.role in ['admin', 'gestionnaire'] and (user.unite_id or user.region_id):
        if user.unite_id:
            unites_principales = [Unite.query.get(user.unite_id)]
            regions = [] # Un provincial n'a pas besoin de choisir de région (il ne crée que des centres)
        elif user.region_id:
            unites_principales = Unite.query.filter_by(region_id=user.region_id, type='unite').all()
            regions = Region.query.filter_by(id=user.region_id).all() if user.role == 'admin' else []
        else:
            unites_principales = []
            regions = []
    else:
        unites_principales = Unite.query.filter_by(type='unite').all()
        regions = Region.query.order_by(Region.nom).all()
    return render_template('add_unite.html', unites_principales=unites_principales, regions=regions, current_user=user)

@app.route('/admin/unites/modifier/<int:unite_id>', methods=['GET', 'POST'])
@login_required
@admin_or_manager_required
def edit_unite(unite_id):
    """Modifier une unité ou un centre"""
    user = User.query.get(session['user_id'])
    unite = Unite.query.get_or_404(unite_id)
    
    # Vérification domaine de compétence pour utilisateurs restreints
    if user.role in ['admin', 'gestionnaire'] and (user.unite_id or user.region_id):
        if user.unite_id:
            competence = Unite.query.get(user.unite_id)
            if not competence or competence.type == 'centre':
                flash("Accès refusé. Le niveau centre ne peut pas gérer les structures.", 'warning')
                return redirect(url_for('dashboard'))
            
            if unite.id != user.unite_id and unite.parent_id != user.unite_id:
                flash("Vous n'avez pas l'autorisation de modifier cette structure.", 'danger')
                return redirect(url_for('admin_unites'))
        elif user.region_id:
            if unite.region_id != user.region_id:
                flash("Vous ne pouvez modifier que les structures de votre région.", 'danger')
                return redirect(url_for('admin_unites'))
    
    if request.method == 'POST':
        unite.nom = request.form.get('nom')
        unite.type = request.form.get('type')
        parent_id = request.form.get('parent_id')
        unite.region_id = request.form.get('region_id')
        
        # Bloquer le changement de type ou de parent pour un utilisateur restreint
        if user.role in ['admin', 'gestionnaire'] and (user.unite_id or user.region_id):
             if user.unite_id:
                 if unite.id == user.unite_id:
                     unite.type = 'unite'
                     unite.parent_id = None
                 else:
                     unite.type = 'centre'
                     unite.parent_id = user.unite_id
             elif user.region_id:
                 unite.region_id = user.region_id # Forcer sa région
        else:
            unite.parent_id = parent_id if parent_id and unite.type == 'centre' else None
            
        # Si c'est un centre, il hérite de la région du parent
        if unite.type == 'centre' and unite.parent_id:
             parent = Unite.query.get(unite.parent_id)
             if parent:
                 unite.region_id = parent.region_id
            
        unite.description = request.form.get('description')
        
        try:
            db.session.commit()
            flash(f'Unité "{unite.nom}" modifiée avec succès.', 'success')
            return redirect(url_for('admin_unites'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de la modification : {str(e)}', 'danger')
            
    if user.role in ['admin', 'gestionnaire'] and (user.unite_id or user.region_id):
        if user.unite_id:
            unites_principales = [Unite.query.get(user.unite_id)]
        elif user.region_id:
            unites_principales = Unite.query.filter_by(region_id=user.region_id, type='unite').all()
        else:
            unites_principales = []
    else:
        unites_principales = Unite.query.filter_by(type='unite').all()
        
    if user.role in ['admin', 'gestionnaire'] and (user.unite_id or user.region_id):
        if user.unite_id:
            regions = []
        elif user.region_id:
            regions = Region.query.filter_by(id=user.region_id).all() if user.role == 'admin' else []
        else:
            regions = []
    else:
        regions = Region.query.order_by(Region.nom).all()
    return render_template('edit_unite.html', unite=unite, unites_principales=unites_principales, regions=regions, current_user=user)

@app.route('/admin/unites/supprimer/<int:unite_id>', methods=['POST'])
@login_required
@admin_or_manager_required
def delete_unite(unite_id):
    """Supprimer une unité ou un centre"""
    user = User.query.get(session['user_id'])
    unite = Unite.query.get_or_404(unite_id)
    
    # Vérification domaine de compétence pour utilisateurs restreints
    if user.role in ['admin', 'gestionnaire'] and (user.unite_id or user.region_id):
        if user.unite_id:
            competence = Unite.query.get(user.unite_id)
            if not competence or competence.type == 'centre':
                flash("Accès refusé. Le niveau centre ne peut pas gérer les structures.", 'warning')
                return redirect(url_for('dashboard'))

            if unite.id == user.unite_id:
                flash("Vous n'avez pas l'autorisation de supprimer votre propre unité de compétence.", 'danger')
                return redirect(url_for('admin_unites'))
            if unite.parent_id != user.unite_id:
                flash("Vous n'avez l'autorisation de supprimer que les centres rattachés à votre unité.", 'danger')
                return redirect(url_for('admin_unites'))
        elif user.region_id:
            if unite.region_id != user.region_id:
                flash("Vous ne pouvez supprimer que les structures de votre région.", 'danger')
                return redirect(url_for('admin_unites'))
        else:
            flash("Accès refusé. Domaine de compétence non défini.", 'warning')
            return redirect(url_for('dashboard'))
            
    nom = unite.nom
    
    try:
        # Vérifier s'il y a des dotations liées
        if Dotation.query.filter_by(unite_id=unite_id).first():
            flash(f'Impossible de supprimer "{nom}" car des dotations y sont liées.', 'warning')
            return redirect(url_for('admin_unites'))
            
        db.session.delete(unite)
        db.session.commit()
        flash(f'Unité "{nom}" supprimée avec succès.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression : {str(e)}', 'danger')
        
    return redirect(url_for('admin_unites'))

@app.route('/admin/users/restore/<int:user_id>', methods=['POST'])
@login_required
def restore_user(user_id):
    """
    Restaurer un utilisateur supprimé - SUPER ADMIN uniquement
    """
    current_user = User.query.get(session['user_id'])
    
    # Vérifier que c'est un SUPER ADMIN
    if not current_user or not current_user.is_super_admin:
        return jsonify({'success': False, 'message': 'Accès refusé. Seul le SUPER ADMIN peut restaurer des utilisateurs.'}), 403
    
    user = User.query.get_or_404(user_id)
    
    # Vérifier que l'utilisateur est bien supprimé
    if not user.is_deleted:
        return jsonify({'success': False, 'message': 'Cet utilisateur n\'est pas supprimé.'}), 400
    
    try:
        # Restaurer l'utilisateur
        user.is_deleted = False
        user.deleted_at = None
        user.deleted_by = None
        db.session.commit()
        
        # Créer un log d'audit
        create_audit_log(
            action='restore',
            entity_type='user',
            entity_id=user.id,
            entity_name=user.name,
            description=f"Restauration de l'utilisateur {user.name} ({user.email}) par {current_user.name}"
        )
        
        return jsonify({
            'success': True, 
            'message': f'Utilisateur "{user.name}" restauré avec succès !'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Erreur lors de la restauration: {str(e)}'
        }), 500

@app.route('/admin/restore/<string:entity_type>/<int:entity_id>', methods=['POST'])
@login_required
def restore_entity(entity_type, entity_id):
    """
    Restaurer une entité supprimée (Article, Dotation, Réception, Avis)
    Accessible uniquement aux ADMIN et SUPER ADMIN
    """
    current_user = User.query.get(session['user_id'])
    
    # Vérifier que c'est un ADMIN ou SUPER ADMIN
    if not current_user or (current_user.role != 'admin' and not current_user.is_super_admin):
        return jsonify({'success': False, 'message': 'Accès refusé. Droits administrateur requis.'}), 403
    
    try:
        if entity_type == 'item':
            entity = Item.query.get_or_404(entity_id)
            entity_name = entity.name
        elif entity_type == 'dotation':
            entity = Dotation.query.get_or_404(entity_id)
            entity_name = entity.numero_dotation
            # Restaurer le coût dans le budget (re-créer la consommation)
            # Et re-soustraire du stock
            for di in entity.items:
                di.item.quantity -= di.quantite_dotee
                # Recréer le ReceptionStock ? Non, on utilise le lien existant
        elif entity_type == 'reception':
            entity = Reception.query.get_or_404(entity_id)
            entity_name = entity.item.name if entity.item else "Réception"
            # Re-créer le ReceptionStock FIFO
            new_stock = ReceptionStock(
                item_id=entity.item_id,
                reception_id=entity.id,
                quantite_initiale=entity.quantity,
                quantite_restante=entity.quantity,
                prix_unitaire_ht=entity.prix_unitaire_ht,
                prix_unitaire_ttc=entity.prix_unitaire_ttc,
                taux_tva=entity.taux_tva,
                date_reception=entity.date_reception or entity.created_at
            )
            db.session.add(new_stock)
            # Re-ajouter au stock total de l'article
            if entity.item:
                entity.item.quantity += entity.quantity
        elif entity_type == 'avis':
            entity = AvisAchat.query.get_or_404(entity_id)
            entity_name = entity.numero_avis
        else:
            return jsonify({'success': False, 'message': 'Type d\'entité inconnu.'}), 400

        if not entity.is_deleted:
            return jsonify({'success': False, 'message': 'Cet élément n\'est pas supprimé.'}), 400

        # Restauration commune
        entity.is_deleted = False
        entity.deleted_at = None
        entity.deleted_by = None
        
        db.session.commit()
        
        create_audit_log(
            action='restore',
            entity_type=entity_type,
            entity_id=entity_id,
            entity_name=entity_name,
            description=f"Restauration de {entity_type} ({entity_name}) par {current_user.name}"
        )
        
        return jsonify({
            'success': True,
            'message': f'{entity_type.capitalize()} "{entity_name}" restauré avec succès !'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Erreur lors de la restauration: {str(e)}'
        }), 500


@app.route('/admin/trash')
@login_required
def admin_trash():
    """
    Tableau de bord de restauration (Corbeille)
    Affiche tous les éléments supprimés pour restauration par les admins
    """
    current_user = User.query.get(session['user_id'])
    
    # Vérifier les droits (Admin ou Super Admin)
    if not current_user or (current_user.role != 'admin' and not current_user.is_super_admin):
        flash('❌ Accès refusé. Droits administrateur requis.', 'error')
        return redirect(url_for('dashboard'))
    
    # Récupérer les éléments supprimés
    deleted_users = User.query.filter_by(is_deleted=True).all()
    deleted_items = Item.query.filter_by(is_deleted=True).all()
    deleted_dotations = Dotation.query.filter_by(is_deleted=True).all()
    deleted_receptions = Reception.query.filter_by(is_deleted=True).all()
    deleted_avis = AvisAchat.query.filter_by(is_deleted=True).all()
    
    return render_template('admin_trash.html', 
                         users=deleted_users, 
                         items=deleted_items, 
                         dotations=deleted_dotations,
                         receptions=deleted_receptions,
                         avis=deleted_avis)

@app.route('/admin/users/delete-permanently/<int:user_id>', methods=['POST'])
@login_required
def delete_user_permanently(user_id):
    """
    Supprimer définitivement un utilisateur de la base de données - SUPER ADMIN uniquement
    ATTENTION : Cette action est IRRÉVERSIBLE !
    """
    current_user = User.query.get(session['user_id'])
    
    # Vérifier si c'est une requête AJAX
    is_ajax = request.headers.get('Content-Type') == 'application/json' or \
              request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    # Vérifier que c'est un SUPER ADMIN
    if not current_user or not current_user.is_super_admin:
        if is_ajax:
            return jsonify({'success': False, 'message': 'Accès refusé. Seul le SUPER ADMIN peut supprimer définitivement des utilisateurs.'}), 403
        flash('❌ Accès refusé. Seul le SUPER ADMIN peut supprimer définitivement des utilisateurs.', 'error')
        return redirect(url_for('admin_users_list'))
    
    user = User.query.get_or_404(user_id)
    
    # Vérifier que l'utilisateur est bien supprimé (soft delete)
    if not user.is_deleted:
        if is_ajax:
            return jsonify({'success': False, 'message': 'Cet utilisateur doit d\'abord être supprimé (soft delete) avant la suppression définitive.'}), 400
        flash('❌ Cet utilisateur doit d\'abord être supprimé avant la suppression définitive.', 'error')
        return redirect(url_for('admin_users_list'))
    
    # Vérifier qu'on ne supprime pas soi-même
    if user.id == current_user.id:
        if is_ajax:
            return jsonify({'success': False, 'message': 'Vous ne pouvez pas vous supprimer vous-même.'}), 403
        flash('❌ Vous ne pouvez pas vous supprimer vous-même.', 'error')
        return redirect(url_for('admin_users_list'))
    
    try:
        # Sauvegarder les informations pour le log avant suppression
        user_name = user.name
        user_email = user.email
        user_id_for_log = user.id
        
        # Créer un log d'audit AVANT la suppression
        create_audit_log(
            action='delete_permanently',
            entity_type='user',
            entity_id=user_id_for_log,
            entity_name=user_name,
            description=f"⚠️ SUPPRESSION DÉFINITIVE de l'utilisateur {user_name} ({user_email}) par {current_user.name}"
        )
        
        # SUPPRESSION DÉFINITIVE de la base de données
        db.session.delete(user)
        db.session.commit()
        
        if is_ajax:
            return jsonify({
                'success': True,
                'message': f'Utilisateur "{user_name}" supprimé définitivement de la base de données.'
            })
        
        flash(f'✅ Utilisateur "{user_name}" supprimé définitivement de la base de données.', 'success')
        return redirect(url_for('admin_trash'))
        
    except Exception as e:
        db.session.rollback()
        print(f"ERREUR DE SUPPRESSION DÉFINITIVE: {str(e)}")
        import traceback
        traceback.print_exc()
        
        if is_ajax:
            return jsonify({
                'success': False,
                'message': f'Erreur lors de la suppression définitive: {str(e)}'
            }), 500
        
        flash(f'❌ Erreur lors de la suppression définitive: {str(e)}', 'error')
        return redirect(url_for('admin_trash'))

@app.route('/admin/delete-permanent/<string:entity_type>/<int:entity_id>', methods=['POST'])
@login_required
def delete_permanent_entity(entity_type, entity_id):
    """
    Supprimer définitivement une entité (Article, Dotation, Réception, Avis)
    Accessible uniquement aux ADMIN et SUPER ADMIN
    """
    current_user = User.query.get(session['user_id'])
    
    # Vérifier que c'est un ADMIN ou SUPER ADMIN
    if not current_user or (current_user.role != 'admin' and not current_user.is_super_admin):
        return jsonify({'success': False, 'message': 'Accès refusé. Droits administrateur requis.'}), 403
    
    if entity_type == 'user':
        # Appeler la fonction spécifique pour les utilisateurs
        return delete_user_permanently(entity_id)
        
    try:
        if entity_type == 'item':
            entity = Item.query.get_or_404(entity_id)
            entity_name = entity.name
        elif entity_type == 'dotation':
            entity = Dotation.query.get_or_404(entity_id)
            entity_name = entity.numero_dotation
        elif entity_type == 'reception':
            entity = Reception.query.get_or_404(entity_id)
            entity_name = entity.item.name if entity.item else "Réception"
        elif entity_type == 'avis':
            entity = AvisAchat.query.get_or_404(entity_id)
            entity_name = entity.numero_avis
        else:
            return jsonify({'success': False, 'message': 'Type d\'entité inconnu.'}), 400

        if not entity.is_deleted:
            return jsonify({'success': False, 'message': 'Cet élément n\'est pas dans la corbeille.'}), 400

        # Log d'audit avant suppression
        create_audit_log(
            action='delete_permanently',
            entity_type=entity_type,
            entity_id=entity_id,
            entity_name=entity_name,
            description=f"⚠️ SUPPRESSION DÉFINITIVE de {entity_type} ({entity_name}) par {current_user.name}"
        )
        
        db.session.delete(entity)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{entity_type.capitalize()} "{entity_name}" supprimé définitivement !'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Erreur lors de la suppression définitive: {str(e)}'
        }), 500

@app.route('/admin/users/toggle-status/<int:user_id>', methods=['POST'])
@login_required
@can_manage_user
def toggle_user_status(user_id):
    """Activer ou désactiver un utilisateur"""
    current_user = User.query.get(session['user_id'])
    user = User.query.get_or_404(user_id)
    
    # Sécurité : un admin ne peut pas désactiver un Super Admin
    if user.is_super_admin:
        return jsonify({'success': False, 'message': 'Impossible de désactiver un Super Admin.'}), 403
    
    # Toggle status
    user.is_active = not getattr(user, 'is_active', True)
    db.session.commit()
    
    status_label = "activé" if user.is_active else "désactivé"
    
    create_audit_log(
        action='update',
        entity_type='user',
        entity_id=user.id,
        entity_name=user.name,
        description=f"L'utilisateur {user.name} a été {status_label} par {current_user.name}"
    )
    
    return jsonify({
        'success': True,
        'message': f'Utilisateur {status_label} avec succès !',
        'is_active': user.is_active
    })

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = (request.form.get('email') or '').strip().lower()
        password = request.form.get('password')

        if not email or not password:
            flash('Veuillez fournir un email et un mot de passe valides.', 'error')
            return render_template('login.html')

        user = User.query.filter_by(email=email, is_deleted=False).first()
        if not user or not user.verify_password(password):
            flash('Email ou mot de passe invalide.', 'error')
            return render_template('login.html')

        # Vérifier si l'utilisateur est activé
        if not getattr(user, 'is_active', True):
            flash('❌ Votre compte est désactivé. Veuillez contacter l\'administrateur.', 'error')
            return render_template('login.html')

        session.clear()
        session['user_id'] = user.id
        session['user_name'] = user.name or user.email
        session['user_role'] = user.role
        session['is_super_admin'] = user.is_super_admin

        # Mettre à jour la dernière date de connexion
        user.last_login = datetime.utcnow()
        db.session.commit()  # Sauvegarder la date de dernière connexion

        create_audit_log(
            action='login',
            entity_type='user',
            entity_id=user.id,
            entity_name=user.name or user.email,
            description='Connexion réussie au système'
        )

        flash('Connexion réussie !', 'success')
        return redirect(url_for('dashboard'))

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Vous avez été déconnecté.', 'info')
    return redirect(url_for('login'))

@app.route('/account/password', methods=['GET', 'POST'])
@login_required
def change_password():
    """Permet à l'utilisateur connecté de changer son propre mot de passe"""
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user = User.query.get(session['user_id'])
    if not user:
        session.clear()
        return redirect(url_for('login'))

    if request.method == 'POST':
        current_password = request.form.get('current_password') or ''
        new_password = request.form.get('new_password') or ''
        new_password_confirm = request.form.get('new_password_confirm') or ''

        # Vérifier le mot de passe actuel
        if not user.verify_password(current_password):
            flash('❌ Mot de passe actuel incorrect.', 'error')
            return redirect(url_for('change_password'))

        # Vérifier la longueur du nouveau mot de passe
        if len(new_password) < 8:
            flash('❌ Le nouveau mot de passe doit contenir au moins 8 caractères.', 'error')
            return redirect(url_for('change_password'))

        # Vérifier la confirmation
        if new_password != new_password_confirm:
            flash('❌ Les nouveaux mots de passe ne correspondent pas.', 'error')
            return redirect(url_for('change_password'))

        # Appliquer le nouveau mot de passe
        user.set_password(new_password)

        try:
            db.session.commit()

            create_audit_log(
                action='change_password',
                entity_type='user',
                entity_id=user.id,
                entity_name=user.name or user.email,
                description="Changement du mot de passe par l'utilisateur lui-même"
            )

            flash('✅ Votre mot de passe a été mis à jour avec succès.', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Erreur lors de la mise à jour du mot de passe: {str(e)}', 'error')
            return redirect(url_for('change_password'))

    return render_template('change_password.html')

@app.route('/dashboard')
@login_required
def dashboard():
    from datetime import datetime
    from sqlalchemy import extract
    
    # Récupérer l'année du paramètre de requête, ou utiliser l'année courante
    year_filter = request.args.get('year', type=int)
    current_year = year_filter if year_filter else datetime.now().year
    
    # Récupérer les années disponibles pour le filtre
    # Années des réceptions
    reception_years = db.session.query(
        extract('year', Reception.created_at).label('year')
    ).distinct().all()
    
    # Années des dotations
    dotation_years = db.session.query(
        extract('year', Dotation.date_dotation).label('year')
    ).distinct().all()
    
    # Combiner et trier les années uniques
    all_years = sorted(list(set([int(y[0]) for y in (reception_years + dotation_years) if y[0] is not None])), reverse=True)
    
    # Si aucune année n'est spécifiée, utiliser l'année la plus récente
    if not year_filter and all_years:
        current_year = all_years[0]
    
    # Calculer les années précédente et suivante pour la navigation
    if all_years:
        min_year = min(all_years)
        max_year = max(all_years)
        prev_year = current_year - 1 if (current_year - 1) >= min_year else None
        next_year = current_year + 1 if (current_year + 1) <= max_year else None
    else:
        min_year = max_year = current_year
        prev_year = next_year = None
    
    # Statistiques générales
    total_items = Item.query.count()
    items_in_stock = Item.query.filter(Item.quantity > Item.reorder_level).count()
    low_stock_items = Item.query.filter(Item.quantity <= Item.reorder_level, Item.quantity > 0).count()
    out_of_stock_items = Item.query.filter(Item.quantity == 0).count()
    
    # Valeur totale du stock (simulation avec prix moyen)
    total_quantity = db.session.query(db.func.sum(Item.quantity)).scalar() or 0
    
    # Réceptions récentes (filtrées par année si spécifiée)
    recent_receptions_query = Reception.query
    if year_filter:
        recent_receptions_query = recent_receptions_query.filter(
            extract('year', Reception.created_at) == current_year
        )
    recent_receptions = recent_receptions_query.order_by(Reception.created_at.desc()).limit(5).all()
    
    # Statistiques par catégorie
    categories_stats = db.session.query(
        Item.category,
        db.func.count(Item.id).label('count'),
        db.func.sum(Item.quantity).label('total_qty')
    ).group_by(Item.category).all()
    
    # Articles les plus reçus (basé sur les réceptions)
    top_received_items_query = db.session.query(
        Item.name,
        Item.sku,
        db.func.sum(Reception.quantity).label('total_received')
    ).join(Reception).filter(
        Reception.type == 'reception'
    )
    
    if year_filter:
        top_received_items_query = top_received_items_query.filter(
            extract('year', Reception.created_at) == current_year
        )
    
    top_received_items = top_received_items_query.group_by(Item.id).order_by(
        db.func.sum(Reception.quantity).desc()
    ).limit(5).all()
    
    # Articles critiques (quantité très faible)
    critical_items = Item.query.filter(
        Item.quantity <= (Item.reorder_level * 0.5),
        Item.quantity > 0
    ).limit(5).all()
    
    # Dotations récentes (filtrées par année si spécifiée)
    recent_dotations_query = Dotation.query
    if year_filter:
        recent_dotations_query = recent_dotations_query.filter(
            extract('year', Dotation.date_dotation) == current_year
        )
    recent_dotations = recent_dotations_query.order_by(
        Dotation.date_dotation.desc()
    ).limit(5).all()
    
    # Logs d'audit récents (10 dernières modifications)
    recent_audit_logs_query = AuditLog.query
    if year_filter:
        recent_audit_logs_query = recent_audit_logs_query.filter(
            extract('year', AuditLog.created_at) == current_year
        )
    recent_audit_logs = recent_audit_logs_query.order_by(AuditLog.created_at.desc()).limit(10).all()
    
    # Convertir TOUTES les dates en objets datetime
    recent_audit_logs = [convert_object_dates(log) for log in recent_audit_logs]
    recent_receptions = [convert_object_dates(r) for r in recent_receptions]
    recent_dotations = [convert_object_dates(d) for d in recent_dotations]
    
    return render_template('dashboard.html', 
                         total_items=total_items,
                         items_in_stock=items_in_stock,
                         low_stock_items=low_stock_items,
                         out_of_stock_items=out_of_stock_items,
                         total_quantity=total_quantity,
                         recent_receptions=recent_receptions,
                         categories_stats=categories_stats,
                         top_received_items=top_received_items,
                         critical_items=critical_items,
                         recent_dotations=recent_dotations,
                         recent_audit_logs=recent_audit_logs,
                         years=all_years,
                         current_year=current_year,
                         prev_year=prev_year,
                         next_year=next_year)

@app.route('/budget-details')
@app.route('/budget-details/<int:annee>')
@login_required
def budget_details(annee=None):
    """Affiche les détails du budget par Unité avec possibilité de modification individuelle"""
    from datetime import datetime
    annee_courante = annee if annee is not None else datetime.now().year
    
    # 1. Vérifier si des budgets existent pour cette année
    has_budgets = BudgetNature.query.filter_by(annee=annee_courante).first()
    if not has_budgets:
        try:
            init_budgets(annee_courante)
        except Exception as e:
            print(f"Erreur init budgets: {e}")
    
    # 2. Récupérer les unités selon la compétence de l'utilisateur
    current_user = User.query.get(session.get('user_id'))
    if not current_user:
        return redirect(url_for('login'))
        
    query = Unite.query
    
    # Filtrage par domaine de compétence
    if current_user.is_super_admin:
        # Super Admin : Tout voir (12 régions)
        unites = query.all()
    elif current_user.region_competence:
        # Compétence Régionale : Unités de la région (Provinces) + Leurs Centres
        # Note: Les centres sont liés aux provinces, qui sont liées à la région.
        # Ou directement par region_id si renseigné sur les centres.
        # Approche la plus large : tout ce qui a ce region_id
        unites = query.filter_by(region_id=current_user.region_competence.id).all()
        # Si les centres n'ont pas region_id explicitement, il faut faire une jointure ou inclure les enfants
        # Supposons que region_id est propagé. Sinon, il faut prendre les parents puis les enfants.
        # Securité supplémentaire : prendre provinces de la région + enfants de ces provinces
        provinces = query.filter_by(region_id=current_user.region_competence.id, type='unite').all()
        province_ids = [p.id for p in provinces]
        centres = query.filter(Unite.parent_id.in_(province_ids)).all()
        # Unir les deux listes sans doublons (si jamais region_id était déjà bon)
        unites = list(set(provinces + centres))
        # Trier par nom pour l'affichage
        unites.sort(key=lambda x: x.nom)
    elif current_user.unite_competence:
        # Compétence Provinciale : L'unité elle-même + Ses centres rattachés
        parent_unit = current_user.unite_competence
        children = query.filter_by(parent_id=parent_unit.id).all()
        unites = [parent_unit] + children
    else:
        # Fallback (devrait pas arriver pour un admin/gestionnaire sans compétence)
        unites = []
    
    # 3. Organiser les données par unité
    budget_par_unite = {}
    
    for unite in unites:
        budget_par_unite[unite.id] = {
            'nom': unite.nom,
            'type': unite.type,
            'categories': {}
        }
        
        # Récupérer les budgets individuels définis pour cette unité
        unite_budgets = BudgetNature.query.filter_by(unite_id=unite.id, annee=annee_courante).all()
        
        # Si aucun budget n'existe pour cette unité, on les crée par défaut (restauration)
        if not unite_budgets:
            # Configuration par défaut
            budgets_config = [
                {'nature': 'Fournitures de bureaux et documentation', 'budget_centre': 1500.00, 'budget_unite': 7000.00},
                {'nature': 'Fournitures pour le matériel informatique', 'budget_centre': 1500.00, 'budget_unite': 3000.00},
                {'nature': "Produits d'hygiène et de désinfection", 'budget_centre': 2000.00, 'budget_unite': 5000.00},
                {'nature': 'Entretien bâtiments administratifs', 'budget_centre': 1500.00, 'budget_unite': 10000.00},
                {'nature': 'Alimentation à usage Humaine', 'budget_centre': 0.00, 'budget_unite': 2000.00},
            ]
            exclusions_centres = ["Alimentation à usage Humaine", "Alimentation Humaine"]
            
            created_budgets = []
            for config in budgets_config:
                # Exclure l'alimentation humaine pour les centres
                if unite.type == 'centre' and config['nature'] in exclusions_centres:
                    continue
                    
                montant = config['budget_unite'] if unite.type == 'unite' else config['budget_centre']
                budget = BudgetNature(
                    nature=config['nature'],
                    montant_ttc=montant,
                    annee=annee_courante,
                    unite_id=unite.id
                )
                db.session.add(budget)
                created_budgets.append(budget)
            
            if created_budgets:
                db.session.commit()
                unite_budgets = created_budgets
        
        # Liste des exclusions pour les centres
        exclusions_centres = ["Alimentation à usage Humaine", "Alimentation Humaine"]
        
        for b in unite_budgets:
            # Exclure certaines natures pour les centres
            if unite.type == 'centre' and b.nature in exclusions_centres:
                continue
                
            montant_consomme = b.budget_consomme
            budget_alloue = b.montant_ttc
            budget_restant = budget_alloue - montant_consomme
            pourcentage_utilise = (montant_consomme / budget_alloue * 100) if budget_alloue > 0 else 0
            
            budget_par_unite[unite.id]['categories'][b.nature] = {
                'id': b.id,
                'budget_alloue': budget_alloue,
                'montant_consomme': montant_consomme,
                'budget_restant': budget_restant,
                'pourcentage_utilise': pourcentage_utilise,
                'est_epuise': budget_restant <= 0
            }
            
    # Liste des années disponibles
    try:
        years_budgets = [y[0] for y in db.session.query(BudgetNature.annee).distinct().order_by(BudgetNature.annee).all()]
        years_cons = [y[0] for y in db.session.query(ConsommationBudget.annee).distinct().order_by(ConsommationBudget.annee).all()]
        annees_disponibles = sorted(set(years_budgets + years_cons))
    except Exception:
        annees_disponibles = [annee_courante]

    return render_template('budget_details.html',
                         budget_par_unite=budget_par_unite,
                         annee=annee_courante,
                         annees_disponibles=annees_disponibles)

@app.route('/admin/budget/update-single', methods=['POST'])
@login_required
@modification_required
def update_single_budget():
    """Met à jour un montant budgétaire spécifique pour une ligne (Ajax)"""
    budget_id = request.form.get('budget_id', type=int)
    nouveau_montant = request.form.get('montant', type=float)
    
    current_user = User.query.get(session.get('user_id'))
    if not current_user or not current_user.is_super_admin:
        return jsonify({'success': False, 'message': 'Action non autorisée. Seul le Super Admin peut modifier les budgets.'})

    if budget_id is None or nouveau_montant is None:
        return jsonify({'success': False, 'message': 'Données invalides.'})
        
    budget = BudgetNature.query.get(budget_id)
    if not budget:
        return jsonify({'success': False, 'message': 'Budget introuvable.'})
        
    old_amount = budget.montant_ttc
    budget.montant_ttc = nouveau_montant
    
    try:
        db.session.commit()
        create_audit_log(
            action='update_budget',
            entity_type='budget_nature',
            entity_id=budget.id,
            entity_name=f"{budget.nature} ({budget.unite.nom})",
            description=f"Changement budget {budget.annee}: {old_amount} -> {nouveau_montant} DH"
        )
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/admin/budget/add-nature', methods=['POST'])
@login_required
@modification_required
def add_budget_nature():
    """Ajoute une nouvelle nature de prestation pour TOUTES les unités"""
    nature = request.form.get('nature')
    annee = request.form.get('annee', type=int)
    montant_initial = request.form.get('montant_initial', type=float, default=0.0)
    
    if not nature or not annee:
        flash('❌ Veuillez remplir tous les champs.', 'error')
        return redirect(url_for('budget_details', annee=annee))
        
    # Vérifier l'exclusion pour les centres (si on ajoute "Alimentation Humaine")
    exclusions_centres = ["Alimentation à usage Humaine", "Alimentation Humaine"]
    
    unites = Unite.query.all()
    added_count = 0
    
    for u in unites:
        # Ne pas créer pour les centres si c'est une nature exclue
        if u.type == 'centre' and nature in exclusions_centres:
            continue
            
        # Vérifier si elle existe déjà pour cette unité
        exists = BudgetNature.query.filter_by(nature=nature, annee=annee, unite_id=u.id).first()
        if not exists:
            new_b = BudgetNature(
                nature=nature,
                annee=annee,
                unite_id=u.id,
                montant_ttc=montant_initial
            )
            db.session.add(new_b)
            added_count += 1
            
    try:
        db.session.commit()
        if added_count > 0:
            flash(f'✅ Nature "{nature}" ajoutée avec succès pour {added_count} unités.', 'success')
            create_audit_log(
                action='add_budget_nature',
                entity_type='budget_nature',
                entity_name=nature,
                description=f"Nouvelle nature de prestation ajoutée pour l'année {annee}"
            )
        else:
            flash('ℹ️ Cette nature existe déjà pour toutes les unités concernées.', 'info')
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Erreur lors de l'ajout: {str(e)}", "error")
        
    return redirect(url_for('budget_details', annee=annee))

@app.route('/admin/budget/<int:annee>/update', methods=['POST'])
@login_required
@admin_or_manager_required
def update_budget_year(annee):
    """Met à jour ou ajoute des lignes de budget (BudgetNature) pour une année donnée.

    - Met à jour les montants centre/unité pour les lignes existantes.
    - Permet d'ajouter une nouvelle nature de prestation avec ses budgets.
    """
    budgets = BudgetNature.query.filter_by(annee=annee).all()

    def _parse_amount(raw_val, default_val):
        try:
            if raw_val is None:
                return default_val
            return float(str(raw_val).replace(' ', '').replace(',', '.'))
        except Exception:
            return default_val

    changed = 0
    created = 0

    # Si le bouton "Modifier" d'une ligne est utilisé, ne traiter que cette ligne
    target_id = None
    save_row = request.form.get('save_row')
    if save_row:
        try:
            target_id = int(save_row)
        except (TypeError, ValueError):
            target_id = None

    # Mise à jour des lignes existantes (montants + optionnellement renommage de la nature)
    for b in budgets:
        if target_id is not None and b.id != target_id:
            continue

        centre_key = f'budget_centre_{b.id}'
        unite_key = f'budget_unite_{b.id}'
        nature_key = f'nature_{b.id}'

        centre_raw = request.form.get(centre_key)
        unite_raw = request.form.get(unite_key)
        new_nature_raw = request.form.get(nature_key)

        # Mettre à jour les montants centre / unité
        new_centre = _parse_amount(centre_raw, b.budget_centre_ttc or 0.0)
        new_unite = _parse_amount(unite_raw, b.budget_unite_ttc or 0.0)

        if new_centre != b.budget_centre_ttc or new_unite != b.budget_unite_ttc:
            b.budget_centre_ttc = new_centre
            b.budget_unite_ttc = new_unite
            changed += 1

        # Renommage éventuel de la nature de prestation
        if new_nature_raw is not None:
            new_nature = new_nature_raw.strip()
        else:
            new_nature = b.nature

        if new_nature and new_nature != b.nature:
            # Empêcher les doublons pour la même année
            conflict = BudgetNature.query.filter(
                BudgetNature.annee == annee,
                BudgetNature.id != b.id,
                BudgetNature.nature == new_nature
            ).first()
            if conflict:
                flash(
                    f'La nature "{new_nature}" existe déjà pour {annee}. Renommage ignoré pour cette ligne.',
                    'warning'
                )
            else:
                old_nature = b.nature
                b.nature = new_nature

                # Propager le renommage dans les entités liées
                ConsommationBudget.query.filter_by(nature=old_nature, annee=annee).update({
                    'nature': new_nature
                })
                Item.query.filter_by(category=old_nature).update({
                    'category': new_nature
                })
                Dotation.query.filter_by(categorie=old_nature).update({
                    'categorie': new_nature
                })
                AvisAchat.query.filter_by(nature_prestation=old_nature).update({
                    'nature_prestation': new_nature
                })
                BonCommande.query.filter_by(nature_prestation=old_nature).update({
                    'nature_prestation': new_nature
                })

                changed += 1

    # Ajout éventuel d'une nouvelle ligne de budget
    new_nature = (request.form.get('new_nature') or '').strip()
    new_centre_raw = request.form.get('new_budget_centre')
    new_unite_raw = request.form.get('new_budget_unite')

    if new_nature and (new_centre_raw or new_unite_raw):
        new_centre = _parse_amount(new_centre_raw or 0, 0.0)
        new_unite = _parse_amount(new_unite_raw or 0, 0.0)

        existing = BudgetNature.query.filter_by(nature=new_nature, annee=annee).first()
        if existing:
            # Si la nature existe déjà pour cette année, on met à jour ses budgets
            if new_centre != existing.budget_centre_ttc or new_unite != existing.budget_unite_ttc:
                existing.budget_centre_ttc = new_centre
                existing.budget_unite_ttc = new_unite
                changed += 1
        else:
            new_budget = BudgetNature(
                nature=new_nature,
                budget_centre_ttc=new_centre,
                budget_unite_ttc=new_unite,
                annee=annee
            )
            db.session.add(new_budget)
            created += 1

    try:
        if changed > 0 or created > 0:
            db.session.commit()
            msg = f'Budgets {annee} mis à jour ({changed} ligne(s) modifiée(s), {created} créée(s)).'
            flash(msg, 'success')
            try:
                create_audit_log(
                    action='update',
                    entity_type='budget',
                    entity_id=None,
                    entity_name=f'Budgets {annee}',
                    description=msg
                )
            except Exception:
                pass
        else:
            flash('Aucune modification de budget détectée.', 'info')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la mise à jour des budgets: {str(e)}', 'error')

    return redirect(url_for('budget_details', annee=annee))


@app.route('/export/budget-excel/<int:annee>')
@login_required
@admin_or_manager_required
def export_budget_excel(annee):
    """Exporter la configuration budgétaire annuelle (BudgetNature) en Excel."""
    from datetime import datetime

    # S'assurer que les budgets de cette année existent
    budgets = BudgetNature.query.filter_by(annee=annee).order_by(BudgetNature.nature).all()
    if not budgets:
        try:
            init_budgets(annee)
            budgets = BudgetNature.query.filter_by(annee=annee).order_by(BudgetNature.nature).all()
        except Exception:
            budgets = []

    # Créer le classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Budgets {annee}"

    # En-tête principal
    ws.merge_cells('A1:H1')
    ws['A1'] = f"BUDGETS ANNUELS - {annee} - Export du {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # En-têtes des colonnes
    headers = [
        'Nature de prestation',
        'Budget centre TTC',
        'Budget unité TTC',
        'Budget total TTC',
        'Consommé TTC',
        'Restant TTC',
        '% utilisé',
        'Statut'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Contenu
    start_row = 4
    for idx, b in enumerate(budgets, start=start_row):
        budget_total = b.budget_total
        budget_consomme = b.budget_consomme
        budget_restant = b.budget_restant
        pourcentage = (budget_consomme / budget_total * 100) if budget_total > 0 else 0.0

        if b.est_verrouille:
            statut = "ÉPUISÉ"
        elif pourcentage >= 80:
            statut = "ALERTE"
        else:
            statut = "OK"

        row_data = [
            b.nature,
            round(b.budget_centre_ttc or 0, 2),
            round(b.budget_unite_ttc or 0, 2),
            round(budget_total, 2),
            round(budget_consomme, 2),
            round(budget_restant, 2),
            round(pourcentage, 1),
            statut
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col, value=value)
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    ws.freeze_panes = 'A4'

    # Sauvegarder dans un buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"budgets_{annee}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/low-stock')
@login_required
def low_stock():
    """Affiche tous les articles en stock bas avec leurs quantités"""
    page = request.args.get('page', 1, type=int)
    category_filter = request.args.get('category', '')
    
    query = Item.query.filter(Item.quantity <= Item.reorder_level)
    
    # Filtre par nature de prestation (catégorie)
    if category_filter:
        if category_filter == 'Entretien bâtiments administratifs':
            # Inclure les 3 sous-catégories d'entretien
            query = query.filter(Item.category.in_([
                'Articles de plomberies',
                'Articles électriques',
                'Articles de la peinture'
            ]))
        else:
            query = query.filter(Item.category == category_filter)
    
    low_stock_items = query.paginate(page=page, per_page=12, error_out=False)
    return render_template('low_stock.html', low_stock_items=low_stock_items, category_filter=category_filter)

@app.route('/budget-dashboard')
@app.route('/budget-dashboard/<int:annee>')
@login_required
def budget_dashboard(annee=None):
    """Redirection vers budget_details - Template fusionné"""
    if annee is None:
        annee = datetime.now().year
    return redirect(url_for('budget_details', annee=annee))

@app.route('/items')
@login_required
def items():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    category_filter = request.args.get('category', '')
    
    query = Item.query.filter_by(is_deleted=False)
    
    # Filtre par recherche textuelle
    if search:
        query = query.filter(
            (Item.name.contains(search)) | 
            (Item.sku.contains(search)) |
            (Item.category.contains(search))
        )
    
    # Filtre par nature de prestation (catégorie)
    if category_filter:
        if category_filter == 'Entretien bâtiments administratifs':
            # Inclure les 3 sous-catégories d'entretien
            query = query.filter(Item.category.in_([
                'Articles de plomberies',
                'Articles électriques',
                'Articles de la peinture'
            ]))
        else:
            query = query.filter(Item.category == category_filter)
    
    items = query.paginate(page=page, per_page=12, error_out=False)
    return render_template('items.html', items=items, search=search, category_filter=category_filter)

@app.route('/items/add', methods=['GET', 'POST'])
@login_required
@modification_required
def add_item():
    if request.method == 'POST':
        # Validation des champs obligatoires
        categorie_principale = request.form.get('categorie_principale', '').strip()
        
        # Si c'est "Entretien bâtiments administratifs", utiliser la sous-prestation
        if categorie_principale == 'Entretien bâtiments administratifs':
            category = request.form.get('category', '').strip()
            if not category:
                flash('La sous-prestation est obligatoire pour Entretien bâtiments administratifs.', 'error')
                next_sku_preview = Item.generate_next_sku()
                return render_template('add_item.html', next_sku=next_sku_preview)
        else:
            category = categorie_principale
        
        if not category:
            flash('La nature de prestation est obligatoire.', 'error')
            next_sku_preview = Item.generate_next_sku()
            return render_template('add_item.html', next_sku=next_sku_preview)
        
        # Vérification des doublons : nom insensible à la casse dans la même catégorie/sous-catégorie
        item_name = request.form['name'].strip()
        
        # Construire la requête de base pour les doublons
        query = Item.query.filter(db.func.lower(Item.name) == db.func.lower(item_name))
        
        # Liste des sous-catégories d'Entretien bâtiments administratifs
        sous_categories_entretien = [
            'Maintenance informatique',
            'Maintenance électrique',
            'Maintenance plomberie',
            'Maintenance climatisation',
            'Maintenance ascenseurs',
            'Maintenance groupes électrogènes',
            'Maintenance systèmes de sécurité'
        ]
        
        # Si c'est une sous-catégorie d'Entretien bâtiments administratifs
        if category in sous_categories_entretien:
            # Vérifier dans la sous-catégorie exacte et dans la catégorie parente
            query = query.filter(
                db.or_(
                    Item.category == category,  # Sous-catégorie exacte
                    Item.category == 'Entretien bâtiments administratifs'  # Catégorie parente
                )
            )
        else:
            # Pour les autres catégories, vérifier la catégorie exacte
            query = query.filter(Item.category == category)
        
        existing_item = query.first()
        
        if existing_item:
            flash(
                f'⚠️ Un article avec le nom "{item_name}" existe déjà dans cette catégorie '
                f'(SKU: {existing_item.sku}, Nature: {existing_item.category or "N/A"}).',
                'error'
            )
            next_sku_preview = Item.generate_next_sku()
            return render_template('add_item.html', next_sku=next_sku_preview)
        
        # Génère automatiquement le prochain SKU unique
        next_sku = Item.generate_next_sku()
        
        # Création de l'article avec quantité = 0 (sera mise à jour lors des réceptions)
        item = Item(
            sku=next_sku,
            name=request.form['name'],
            description=request.form['description'],
            category=category,
            unit=request.form['unit'],
            reorder_level=int(request.form['reorder_level'] or 0),
            quantity=0  # Quantité initialisée à 0, sera mise à jour via les réceptions
        )
        
        try:
            db.session.add(item)
            db.session.flush()  # Pour obtenir l'ID de l'item
            
            # Log d'audit pour la création d'article
            create_audit_log(
                action='create',
                entity_type='item',
                entity_id=item.id,
                entity_name=item.name,
                description=f"{item.sku} | {item.category} | Stock initial: {item.quantity} {item.unit}"
            )
            
            db.session.commit()
            # Rediriger vers la liste des articles avec paramètres pour le modal de succès
            return redirect(url_for(
                'items',
                op='item_created',
                sku=next_sku,
                name=item.name,
                category=item.category
            ))
        except Exception as e:
            db.session.rollback()
            flash('Erreur lors de l\'ajout de l\'article.', 'error')
    
    # Pré-visualise le prochain SKU pour l'utilisateur
    next_sku_preview = Item.generate_next_sku()
    return render_template('add_item.html', next_sku=next_sku_preview)

@app.route('/items/<int:item_id>/edit', methods=['GET', 'POST'])
@login_required
@modification_required
def edit_item(item_id):
    item = Item.query.get_or_404(item_id)
    
    if request.method == 'POST':
        # Le SKU ne peut pas être modifié
        # Le nom, la catégorie, la description, l'unité et le seuil de commande sont modifiables
        old_name = item.name
        old_category = item.category
        old_description = item.description
        old_unit = item.unit
        old_reorder_level = item.reorder_level
        
        new_name = request.form['name'].strip()
        
        # Gestion de la catégorie (avec ou sans sous-prestation)
        categorie_principale = request.form.get('categorie_principale', '').strip()
        
        if categorie_principale == 'Entretien bâtiments administratifs':
            new_category = request.form.get('category', '').strip()
            if not new_category:
                flash('La sous-prestation est obligatoire pour Entretien bâtiments administratifs.', 'error')
                return render_template('edit_item.html', item=item)
        else:
            new_category = categorie_principale if categorie_principale else request.form.get('category', '').strip()
        
        if not new_category:
            flash('La nature de prestation est obligatoire.', 'error')
            return render_template('edit_item.html', item=item)
        
        # Vérification des doublons : nom insensible à la casse (toutes catégories confondues, sauf article actuel)
        existing_item = Item.query.filter(
            db.func.lower(Item.name) == db.func.lower(new_name),
            Item.id != item.id  # Exclure l'article actuel
        ).first()
        
        if existing_item:
            flash(
                f'⚠️ Un article avec le nom "{new_name}" existe déjà '
                f'(SKU: {existing_item.sku}, Nature: {existing_item.category or "N/A"}). '
                f'Les noms d\'articles doivent être uniques, sans tenir compte des majuscules/minuscules.',
                'error'
            )
            return render_template('edit_item.html', item=item)
        
        item.name = new_name
        item.category = new_category
        item.description = request.form['description']
        item.unit = request.form['unit']
        item.reorder_level = int(request.form['reorder_level'] or 0)
        
        try:
            db.session.commit()
            
            # Log d'audit pour la modification d'article
            changes = []
            if old_name != item.name:
                changes.append(f"Nom: {old_name} → {item.name}")
            if old_category != item.category:
                changes.append(f"Nature de prestation: {old_category} → {item.category}")
            if old_description != item.description:
                changes.append(f"Description modifiée")
            if old_unit != item.unit:
                changes.append(f"Unité: {old_unit} → {item.unit}")
            if old_reorder_level != item.reorder_level:
                changes.append(f"Seuil: {old_reorder_level} → {item.reorder_level}")
            
            if changes:
                create_audit_log(
                    action='update',
                    entity_type='item',
                    entity_id=item.id,
                    entity_name=item.name,
                    description=f"Article modifié: {', '.join(changes)}"
                )
            
            # Message flash avec détails pour le modal
            changes_display = ' | '.join(changes) if changes else 'Aucune modification détectée'
            flash(f'✓ Article modifié avec succès | {changes_display}', 'success')
            return redirect(url_for('edit_item', item_id=item.id))
        except Exception as e:
            db.session.rollback()
            flash('Erreur lors de la mise à jour.', 'error')
    
    return render_template('edit_item.html', item=item)

@app.route('/delete-item/<int:item_id>', methods=['POST'])
@login_required
@modification_required
def delete_item(item_id):
    item = Item.query.get_or_404(item_id)
    sku = item.sku
    name = item.name
    category = item.category
    
    try:
        # Soft Delete: Marquer l'article comme supprimé au lieu de le supprimer réellement
        item.is_deleted = True
        item.deleted_at = datetime.utcnow()
        item.deleted_by = session.get('user_id')
        
        db.session.commit()
        
        # 6. Réorganiser tous les SKU pour qu'ils soient séquentiels sans trous
        Item.reorganize_all_skus()
        
        # Log d'audit pour la suppression
        create_audit_log(
            action='delete',
            entity_type='item',
            entity_id=item_id,
            entity_name=name,
            description=f"Action: Suppression Soft | SKU: {sku} | Nature: {category} | SKU réorganisés"
        )
        
        # Rediriger vers la liste avec paramètres pour le modal de succès de suppression
        return redirect(url_for(
            'items',
            op='item_deleted',
            sku=sku,
            name=name
        ))
    except Exception as e:
        db.session.rollback()
        flash(f"Erreur lors de la suppression de l'article: {str(e)}", 'error')
        return redirect(url_for('items'))

@app.route('/delete_all_items', methods=['POST'])
@login_required
@admin_or_manager_required
def delete_all_items():
    """Supprimer TOUS les articles et toutes les données liées"""
    try:
        # Récupérer tous les articles
        all_items = Item.query.all()
        nb_items = len(all_items)
        
        # Compter les réceptions et dotations qui seront supprimées
        nb_receptions = Reception.query.count()
        nb_dotations = Dotation.query.count()
        
        # Marquer TOUS les articles comme supprimés (Soft Delete)
        db.session.query(Item).update({
            Item.is_deleted: True,
            Item.deleted_at: datetime.utcnow(),
            Item.deleted_by: session.get('user_id')
        })
        
        db.session.commit()
        
        # Log d'audit pour la suppression totale
        create_audit_log(
            action='delete_all',
            entity_type='item',
            entity_id=0,
            entity_name='TOUS LES ARTICLES',
            description=f"Suppression totale : {nb_items} article(s) - {nb_receptions} réception(s) - {nb_dotations} dotation(s)"
        )
        
        return jsonify({
            'success': True,
            'message': 'Tous les articles ont été supprimés avec succès !',
            'nb_items': nb_items,
            'nb_receptions': nb_receptions,
            'nb_dotations': nb_dotations
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Erreur lors de la suppression : {str(e)}'
        }), 500

@app.route('/reset-all-quantities', methods=['POST'])
@login_required
@admin_or_manager_required
def reset_all_quantities():
    """Met à zéro la quantité de TOUS les articles (opération sensible)."""
    try:
        # Mise à jour en masse
        db.session.query(Item).update({Item.quantity: 0})
        db.session.commit()

        # Log d'audit (best-effort)
        try:
            create_audit_log(
                action='reset_quantities',
                entity_type='item',
                entity_id=0,
                entity_name='TOUS LES ARTICLES',
                description='Mise à zéro de toutes les quantités du stock'
            )
        except Exception:
            pass

        # Rediriger avec un indicateur d'opération pour affichage côté client
        return redirect(url_for('items', op='reset_quantities'))
    except Exception as e:
        db.session.rollback()
        flash(f"Erreur lors de la réinitialisation des quantités: {str(e)}", 'error')
        return redirect(url_for('items'))

@app.route('/delete-items-multiple', methods=['POST'])
@login_required
@modification_required
def delete_items_multiple():
    """Suppression multiple d'articles"""
    if request.method == 'POST':
        try:
            # Accepter JSON ou FORM
            item_ids = []
            data = None
            if request.is_json:
                data = request.get_json(silent=True) or {}
                if isinstance(data, dict):
                    item_ids = data.get('item_ids', []) or []
            if not item_ids:
                # Form fields: item_ids or item_ids[]
                item_ids = request.form.getlist('item_ids') or request.form.getlist('item_ids[]')
            # Normaliser en liste d'int
            try:
                item_ids = list({int(x) for x in item_ids})  # unique
            except Exception:
                item_ids = []

            if not item_ids:
                return jsonify({'success': False, 'error': 'Aucun article sélectionné'}), 400

            # Charger les articles pour log et vérification
            items_to_delete = Item.query.filter(Item.id.in_(item_ids)).all()
            if not items_to_delete:
                return jsonify({'success': False, 'error': 'Aucun article trouvé à supprimer'}), 404

            deleted_count = 0
            for it in items_to_delete:
                try:
                    # Soft Delete
                    it.is_deleted = True
                    it.deleted_at = datetime.utcnow()
                    it.deleted_by = session.get('user_id')
                    deleted_count += 1
                except Exception:
                    continue

            if deleted_count > 0:
                db.session.commit()
                # Réorganiser les SKU après commit de suppression
                try:
                    Item.reorganize_all_skus()
                except Exception:
                    pass

                # Audit
                try:
                    items_info = [f"{i.sku} - {i.name} (ID: {i.id})" for i in items_to_delete]
                    create_audit_log(
                        action='delete',
                        entity_type='item',
                        entity_id=None,
                        entity_name=f"{deleted_count} articles",
                        description=f"Articles supprimés : {', '.join(items_info[:5])}" + ("..." if len(items_info) > 5 else "")
                    )
                except Exception:
                    pass
                if request.is_json:
                    return jsonify({'success': True, 'message': f'{deleted_count} article(s) supprimé(s) avec succès', 'deleted_count': deleted_count})
                else:
                    return redirect(url_for('items'))

            # Aucun élément supprimé
            db.session.rollback()
            if request.is_json:
                return jsonify({'success': False, 'error': 'Aucun article supprimé'}), 400
            else:
                return redirect(url_for('items'))

        except Exception as e:
            db.session.rollback()
            if request.is_json:
                return jsonify({'success': False, 'error': str(e)}), 500
            else:
                flash(f"Erreur lors de la suppression multiple: {str(e)}", 'error')
                return redirect(url_for('items'))

@app.route('/admin/items/duplicates', methods=['GET'])
@login_required
@admin_or_manager_required
def list_item_duplicates():
    items = Item.query.all()
    by_id = {it.id: it for it in items}

    name_groups = {}
    sku_groups = {}

    for it in items:
        key_name = (it.name or '').strip().lower()
        if key_name:
            name_groups.setdefault(key_name, []).append(it.id)

        key_sku = _norm_sku_key_for_duplicates(it.sku)
        if key_sku:
            sku_groups.setdefault(key_sku, []).append(it.id)

    result = []
    seen = set()

    # Groupes par nom
    for key, ids in name_groups.items():
        if len(ids) <= 1:
            continue
        ids_sorted = sorted(set(ids))
        sig = tuple(ids_sorted)
        if sig in seen:
            continue
        seen.add(sig)
        lst_sorted = [by_id[i] for i in ids_sorted]
        kept = lst_sorted[0]
        dups = lst_sorted[1:]
        result.append({
            'group_type': 'name',
            'group_key': key,
            'kept': {
                'id': kept.id,
                'sku': kept.sku,
                'name': kept.name,
                'category': kept.category,
                'quantity': kept.quantity
            },
            'duplicates': [
                {
                    'id': d.id,
                    'sku': d.sku,
                    'name': d.name,
                    'category': d.category,
                    'quantity': d.quantity
                } for d in dups
            ]
        })

    # Groupes par SKU normalisé
    for key, ids in sku_groups.items():
        if len(ids) <= 1:
            continue
        ids_sorted = sorted(set(ids))
        sig = tuple(ids_sorted)
        if sig in seen:
            continue
        seen.add(sig)
        lst_sorted = [by_id[i] for i in ids_sorted]
        kept = lst_sorted[0]
        dups = lst_sorted[1:]
        result.append({
            'group_type': 'sku',
            'group_key': key,
            'kept': {
                'id': kept.id,
                'sku': kept.sku,
                'name': kept.name,
                'category': kept.category,
                'quantity': kept.quantity
            },
            'duplicates': [
                {
                    'id': d.id,
                    'sku': d.sku,
                    'name': d.name,
                    'category': d.category,
                    'quantity': d.quantity
                } for d in dups
            ]
        })

    return jsonify({'success': True, 'duplicate_groups': result, 'count_groups': len(result)})

def _norm_sku_key_for_duplicates(val) -> str:
    s = str(val or '').strip().upper()
    s = re.sub(r'[^A-Z0-9]', '', s)
    return s

@app.route('/admin/budget/reset', methods=['POST'])
@login_required
@admin_or_manager_required
def reset_budget_current_year():
    """Réinitialise les consommations et engagements de l'année courante.

    - Supprime les consommations budgétaires (ConsommationBudget)
    - Supprime les bons de commande (BonCommande)
    - Supprime les indemnités de déplacement (IndemnitesDeplacement)

    Les configurations de budget (BudgetNature) sont conservées.
    """
    from datetime import datetime
    annee_courante = datetime.now().year
    try:
        # Supprimer les consommations budgétaires de l'année
        ConsommationBudget.query.filter_by(annee=annee_courante).delete()

        # Supprimer les bons de commande de l'année
        BonCommande.query.filter(
            db.extract('year', BonCommande.date_engagement) == annee_courante
        ).delete()

        # Supprimer les indemnités de déplacement de l'année
        IndemnitesDeplacement.query.filter(
            db.extract('year', IndemnitesDeplacement.created_at) == annee_courante
        ).delete()

        db.session.commit()

        create_audit_log(
            action='delete',
            entity_type='budget',
            entity_id=None,
            entity_name=f'RESET {annee_courante}',
            description=f'Réinitialisation complète du budget pour {annee_courante} (consommations et engagements supprimés)'
        )

        flash(f'✅ Données budgétaires réinitialisées pour {annee_courante}. Les cartes repartiront de zéro.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la réinitialisation du budget: {str(e)}', 'error')

    return redirect(url_for('budget_details'))

@app.route('/admin/receptions/duplicates/cleanup', methods=['POST'])
@login_required
@admin_or_manager_required
def cleanup_reception_duplicates():
    print("\n=== DÉBUT DE LA FONCTION cleanup_reception_duplicates ===")
    
    try:
        # Récupération des données JSON
        data = request.get_json(silent=True) or {}
        print(f"Données reçues: {data}")
        
        confirm = bool(data.get('confirm', False))
        dry_run = bool(data.get('dry_run', False))
        print(f"Paramètres: confirm={confirm}, dry_run={dry_run}")

        # Récupération de toutes les réceptions
        print("Récupération de toutes les réceptions...")
        all_receptions = Reception.query.order_by(Reception.created_at).all()
        print(f"Nombre total de réceptions trouvées: {len(all_receptions)}")
        
        # Dictionnaire pour suivre les doublons: (item_id, date_reception, quantity, reason) -> [receptions]
        duplicates_map = {}
        
        # Compteur pour le suivi
        processed = 0
        
        print("Analyse des doublons...")
        for reception in all_receptions:
            processed += 1
            if processed % 100 == 0:  # Log toutes les 100 réceptions
                print(f"  Traité {processed}/{len(all_receptions)} réceptions...")
                
            # Création d'une clé pour identifier les doublons
            try:
                key = (
                    reception.item_id,
                    reception.date_reception.date() if reception.date_reception else None,
                    reception.quantity,
                    reception.reason or ''
                )
                
                if key not in duplicates_map:
                    duplicates_map[key] = []
                duplicates_map[key].append(reception)
                    
            except Exception as e:
                print(f"  Erreur lors de la création de la clé pour la réception {reception.id}: {str(e)}")
                print(f"  Détails de la réception: item_id={reception.item_id}, date_reception={reception.date_reception}, quantity={reception.quantity}")
                continue
        
        # Filtrer les entrées non dupliquées (uniquement celles avec plus d'une occurrence)
        duplicate_groups = {k: v for k, v in duplicates_map.items() if len(v) > 1}
        print(f"Groupes de doublons identifiés: {len(duplicate_groups)}")
        
        summary = []
        total_duplicates_removed = 0
        
        # Traitement des groupes de doublons
        for key, receptions in duplicate_groups.items():
            # Trier par date de création (la plus ancienne en premier)
            receptions_sorted = sorted(receptions, key=lambda r: r.created_at)
            kept = receptions_sorted[0]  # On garde la plus ancienne
            duplicates = receptions_sorted[1:]  # Les autres sont des doublons
            
            group_info = {
                'item_id': key[0],
                'date_reception': key[1].isoformat() if key[1] else None,
                'quantity': key[2],
                'reason': key[3] or None,
                'kept_reception_id': kept.id,
                'duplicate_ids': [r.id for r in duplicates]
            }
            
            if confirm and not dry_run:
                # Pour chaque doublon, ajuster la quantité de l'article et supprimer la réception
                item = kept.item
                for dup in duplicates:
                    # Ajuster la quantité de l'article en supprimant la quantité du doublon
                    item.quantity -= dup.quantity
                    
                    # Supprimer la réception en doublon
                    db.session.delete(dup)
                    total_duplicates_removed += 1
                    
                    # Journaliser l'action
                    create_audit_log(
                        action='delete_duplicate_reception',
                        entity_type='reception',
                        entity_id=dup.id,
                        entity_name=f"Réception doublon: {dup.item.name} x{dup.quantity}",
                        description=f"Suppression d'une réception en doublon (gardée: réception #{kept.id})"
                    )
                
                # Journaliser la réception conservée
                create_audit_log(
                    action='keep_reception',
                    entity_type='reception',
                    entity_id=kept.id,
                    entity_name=f"Réception conservée: {kept.item.name} x{kept.quantity}",
                    description=f"Conservation d'une réception après nettoyage des doublons (supprimés: {len(duplicates)} doublons)"
                )
            
            summary.append(group_info)
        
        # Validation des modifications si nécessaire
        if confirm and not dry_run and total_duplicates_removed > 0:
            print(f"Validation de la suppression de {total_duplicates_removed} doublons...")
            try:
                db.session.commit()
                print("Modifications validées avec succès dans la base de données.")
            except Exception as e:
                db.session.rollback()
                error_msg = f"Erreur lors de la validation des modifications: {str(e)}"
                print(error_msg)
                app.logger.error(f"Error cleaning up duplicate receptions: {str(e)}")
                return jsonify({
                    'success': False,
                    'message': f'Erreur lors de la validation des modifications: {str(e)}',
                    'error': str(e)
                }), 500
        else:
            print("Mode dry_run activé ou aucun doublon à supprimer - Aucune modification en base de données.")
        
        # Préparation de la réponse
        response = {
            'success': True,
            'dry_run': dry_run,
            'confirm': confirm,
            'groups_processed': len(summary),
            'duplicates_removed': total_duplicates_removed,
            'details': summary
        }
        
        print(f"=== FIN DE LA FONCTION cleanup_reception_duplicates ===\n")
        return jsonify(response)
        
    except Exception as e:
        # Gestion des erreurs générales
        error_msg = f"Erreur inattendue lors du nettoyage des doublons: {str(e)}"
        print(error_msg)
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'success': False,
            'message': 'Erreur technique lors du traitement de la requête',
            'error': str(e)
        }), 500

@app.route('/admin/items/duplicates/cleanup', methods=['POST'])
@login_required
@admin_or_manager_required
def cleanup_item_duplicates():
    data = request.get_json(silent=True) or {}
    strategy = data.get('strategy', 'keep_oldest')
    confirm = bool(data.get('confirm', False))
    dry_run = bool(data.get('dry_run', False))

    items = Item.query.all()
    by_id = {it.id: it for it in items}

    name_groups = {}
    sku_groups = {}

    for it in items:
        key_name = (it.name or '').strip().lower()
        if key_name:
            name_groups.setdefault(key_name, []).append(it.id)

        key_sku = _norm_sku_key_for_duplicates(it.sku)
        if key_sku:
            sku_groups.setdefault(key_sku, []).append(it.id)

    all_groups = []
    seen = set()

    for key, ids in name_groups.items():
        if len(ids) <= 1:
            continue
        ids_sorted = sorted(set(ids))
        sig = tuple(ids_sorted)
        if sig in seen:
            continue
        seen.add(sig)
        all_groups.append({'group_type': 'name', 'group_key': key, 'item_ids': ids_sorted})

    for key, ids in sku_groups.items():
        if len(ids) <= 1:
            continue
        ids_sorted = sorted(set(ids))
        sig = tuple(ids_sorted)
        if sig in seen:
            continue
        seen.add(sig)
        all_groups.append({'group_type': 'sku', 'group_key': key, 'item_ids': ids_sorted})

    summary = []
    total_merged = 0
    total_deleted = 0

    for grp in all_groups:
        ids = grp['item_ids']
        lst = [by_id[i] for i in ids if i in by_id]
        if len(lst) <= 1:
            continue

        lst_sorted = sorted(lst, key=lambda x: x.id)
        kept = lst_sorted[0] if strategy == 'keep_oldest' else lst_sorted[-1]
        dups = [x for x in lst_sorted if x.id != kept.id]

        group_info = {
            'group_type': grp['group_type'],
            'group_key': grp['group_key'],
            'kept': {'id': kept.id, 'sku': kept.sku, 'name': kept.name},
            'duplicates': [{'id': d.id, 'sku': d.sku, 'name': d.name} for d in dups]
        }

        if confirm and not dry_run:
            for dup in dups:
                DotationItem.query.filter_by(item_id=dup.id).update({DotationItem.item_id: kept.id}, synchronize_session=False)
                Reception.query.filter_by(item_id=dup.id).update({Reception.item_id: kept.id}, synchronize_session=False)
                ReceptionStock.query.filter_by(item_id=dup.id).update({ReceptionStock.item_id: kept.id}, synchronize_session=False)
                AvisAchatItem.query.filter_by(item_id=dup.id).update({AvisAchatItem.item_id: kept.id}, synchronize_session=False)

                kept.quantity = (kept.quantity or 0) + (dup.quantity or 0)
                kept.reorder_level = max(kept.reorder_level or 0, dup.reorder_level or 0)

                db.session.delete(dup)
                total_deleted += 1
            total_merged += 1
        summary.append(group_info)

    if confirm and not dry_run:
        db.session.commit()
        try:
            Item.reorganize_all_skus()
        except Exception:
            pass

    return jsonify({
        'success': True,
        'strategy': strategy,
        'dry_run': dry_run,
        'confirm': confirm,
        'groups_processed': len(summary),
        'items_deleted': total_deleted,
        'items_merged': total_merged,
        'details': summary
    })

# Anciennes routes de mouvements supprimées - remplacées par les réceptions

@app.route('/receptions')
@login_required
def receptions():
    """Affiche toutes les réceptions avec filtrage par catégorie, année et pagination"""
    # Récupérer les filtres depuis l'URL
    category_filter = request.args.get('category', '')
    year_filter = request.args.get('year', type=int)
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Nombre de réceptions par page

    # Construire la requête de base
    query = Reception.query.filter_by(is_deleted=False).join(Item)

    # Appliquer le filtre de catégorie si présent
    if category_filter:
        query = query.filter(Item.category == category_filter)

    # Préparer la liste des années disponibles (basée sur date_reception ou created_at)
    current_year = datetime.utcnow().year
    year_rows = db.session.query(
        func.extract('year', func.coalesce(Reception.date_reception, Reception.created_at))
    ).distinct().all()
    available_years_raw = [int(y[0]) for y in year_rows if y[0] is not None]

    if available_years_raw:
        min_year = min(available_years_raw)
        max_data_year = max(available_years_raw)
        max_year = max(max_data_year, current_year)
    else:
        min_year = current_year
        max_year = current_year

    years = list(range(max_year, min_year - 1, -1))

    # Si aucune année n'est spécifiée, utiliser par défaut l'année la plus récente
    if not year_filter and years:
        year_filter = years[0]

    # Calculer les années précédente et suivante pour la navigation (comme sur le budget)
    prev_year = year_filter - 1 if year_filter else None
    next_year = year_filter + 1 if year_filter else None

    # Appliquer le filtre d'année si présent
    if year_filter:
        query = query.filter(
            func.extract('year', func.coalesce(Reception.date_reception, Reception.created_at)) == year_filter
        )

    # Pagination
    receptions_paginated = query.order_by(Reception.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    # Récupérer toutes les catégories uniques pour le filtre
    categories = db.session.query(Item.category).distinct().order_by(Item.category).all()
    categories = [cat[0] for cat in categories if cat[0]]  # Extraire les valeurs et filtrer les None

    # Convertir les dates des réceptions
    receptions_converted = [convert_object_dates(r) for r in receptions_paginated.items]

    return render_template(
        'receptions.html',
        receptions=receptions_converted,
        pagination=receptions_paginated,
        categories=categories,
        years=years,
        current_category=category_filter,
        current_year=year_filter,
        prev_year=prev_year,
        next_year=next_year,
    )

@app.route('/add-reception', methods=['GET', 'POST'])
@login_required
@modification_required
def add_reception():
    if request.method == 'POST':
        # Validation du fournisseur obligatoire
        reason = request.form.get('reason', '').strip()
        if not reason:
            flash('Le fournisseur est obligatoire.', 'error')
            items = Item.query.order_by(Item.name).all()
            return render_template('add_reception.html', items=items)
        
        item_id = int(request.form['item_id'])
        reception_type = request.form['type']
        quantity = int(request.form['quantity'])
        
        # Date de réception
        date_reception_str = request.form.get('date_reception')
        date_reception = None
        if date_reception_str:
            try:
                date_reception = datetime.strptime(date_reception_str, '%Y-%m-%d').date()
            except ValueError:
                date_reception = datetime.utcnow().date()
        
        # NOUVEAUX CHAMPS BUDGÉTAIRES
        pu_ht = float(request.form.get('pu_ht', 0))
        taux_tva = float(request.form.get('taux_tva', 20))
        
        # Calculer les prix avec la fonction helper
        prix_calcules = calculer_prix_reception(pu_ht, quantity, taux_tva)
        
        item = Item.query.get_or_404(item_id)

        # Empêcher les doublons exacts de réception (même article, date, quantité, fournisseur)
        if reception_type == 'reception':
            dup_q = Reception.query.filter(
                Reception.item_id == item_id,
                Reception.type == 'reception',
                Reception.quantity == quantity,
                Reception.reason == reason
            )
            if date_reception is not None:
                dup_q = dup_q.filter(Reception.date_reception == date_reception)
            else:
                dup_q = dup_q.filter(Reception.date_reception.is_(None))

            if dup_q.first():
                flash("Une réception identique existe déjà pour cet article (même date, quantité et fournisseur). Aucune nouvelle ligne n'a été créée.", 'warning')
                return redirect(url_for('receptions'))

        # Create reception record avec les données budgétaires
        reception = Reception(
            item_id=item_id,
            type=reception_type,
            quantity=quantity,
            reason=reason,
            date_reception=date_reception,
            user_id=session['user_id'],
            # Champs budgétaires
            prix_unitaire_ht=pu_ht,
            taux_tva=taux_tva,
            prix_unitaire_ttc=prix_calcules['pu_ttc'],
            prix_total_ht=prix_calcules['pt_ht'],
            prix_total_ttc=prix_calcules['pt_ttc']
        )
        
        # Update item quantity
        if reception_type == 'reception':
            item.quantity += quantity
        elif reception_type == 'ajustement':
            item.quantity = quantity
        
        try:
            db.session.add(reception)
            db.session.flush()  # Pour obtenir l'ID de la réception
            
            # ========================================================================
            # CRÉER LE STOCK FIFO pour les réceptions (pas pour les ajustements)
            # ========================================================================
            if reception_type == 'reception':
                reception_stock = ReceptionStock(
                    item_id=item_id,
                    reception_id=reception.id,
                    quantite_initiale=quantity,
                    quantite_restante=quantity,  # Tout le stock est disponible
                    prix_unitaire_ht=pu_ht,
                    prix_unitaire_ttc=prix_calcules['pu_ttc'],
                    taux_tva=taux_tva,
                    date_reception=date_reception or datetime.utcnow()
                )
                db.session.add(reception_stock)
            
            # Log d'audit pour la réception
            type_text = 'Réception' if reception_type == 'reception' else 'Ajustement'
            new_stock = item.quantity
            create_audit_log(
                action='create',
                entity_type='reception',
                entity_id=reception.id,
                entity_name=item.name,
                description=f"{type_text}: +{quantity} {item.unit} | Fournisseur: {reason} | {item.sku} | Stock: {new_stock} | Montant TTC: {prix_calcules['pt_ttc']:.2f} DH"
            )
            
            db.session.commit()
            flash(f'✓ Réception enregistrée avec succès | {item.name} | SKU: {item.sku} | Quantité: +{quantity} {item.unit} | Stock actuel: {new_stock} {item.unit} | Montant: {prix_calcules["pt_ttc"]:.2f} DH TTC', 'success')
            return redirect(url_for('receptions'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de l\'enregistrement de la réception: {str(e)}', 'error')
    
    items = Item.query.order_by(Item.name).all()
    return render_template('add_reception.html', items=items)

@app.route('/import-receptions', methods=['POST'])
@login_required
@modification_required
def import_receptions():
    """Importer des réceptions depuis un fichier Excel et créer les enregistrements associés."""
    flash("L'importation Excel des réceptions a été désactivée. Veuillez saisir les réceptions manuellement.", 'error')
    return redirect(url_for('add_reception'))
    if 'file' not in request.files:
        flash('Aucun fichier sélectionné', 'error')
        return redirect(url_for('add_reception'))
    file = request.files['file']
    if not file or file.filename == '':
        flash('Aucun fichier sélectionné', 'error')
        return redirect(url_for('add_reception'))
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Format invalide. Utilisez un fichier Excel (.xlsx ou .xls)', 'error')
        return redirect(url_for('add_reception'))

    try:
        wb = load_workbook(file)
        ws = wb.active

        success = 0
        errors = 0
        details = []

        # Préindexer les articles pour des correspondances robustes
        # - SKU direct (trim, insensible à la casse, tirets et espaces ignorés)
        # - Nom normalisé (accents supprimés, espaces compressés, minuscule)
        all_items = Item.query.all()
        def _norm_name(x: str) -> str:
            """Normalise le nom en supprimant accents, espaces multiples et casse."""
            try:
                s = str(x or '').strip().lower()
                s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
                s = re.sub(r'\s+', ' ', s)
                return s
            except Exception:
                return str(x or '').strip().lower()

        def _norm_sku_key(val) -> str:
            """Normalise un SKU pour comparaison (SKU007 == SKU-007 == sku-007)."""
            s = _to_str(val).strip().upper()
            # Garder uniquement lettres et chiffres (supprime '-', espaces, etc.)
            s = re.sub(r'[^A-Z0-9]', '', s)
            return s

        # Index des SKU normalisés
        sku_index = { _norm_sku_key(it.sku): it for it in all_items if it.sku }
        name_index = { _norm_name(it.name): it for it in all_items if it.name }

        # Lecture des entêtes (compatible ancien modèle et nouveau avec Catégorie/Unité)
        headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        def _norm_key(s):
            s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode('ascii')
            return re.sub(r'[^a-z0-9]', '', s.lower())
        header_idx = {}
        if headers_row:
            for i, h in enumerate(headers_row):
                if h:
                    header_idx[_norm_key(h)] = i

        # Indices avec fallback (ancien modèle) + synonymes robustes
        def first_idx(keys, default=None):
            for k in keys:
                if k in header_idx:
                    return header_idx[k]
            return default
        idx_sku = first_idx(['sku', 'code', 'codearticle'], 0)
        idx_name = first_idx(['nomarticle', 'article', 'designation', 'nom'], 1)
        idx_categorie = first_idx(['categorie', 'nature', 'categoriearticle'], None)
        idx_unite = first_idx(['unite', 'unitee', 'u'], None)
        idx_type = first_idx(['type', 'operation', 'mouvement'], 2)
        idx_qty = first_idx(['quantite', 'qte', 'qty'], 3)
        idx_date = first_idx(['datereception', 'date', 'dateentree'], 4)
        idx_fourn = first_idx(['fournisseur', 'fourn', 'supplier'], 5)
        idx_puht = first_idx(['puht', 'prixunitaireht', 'prixht', 'pu'], 6)
        idx_tva = first_idx(['tauxtva', 'tva', 'taux'], 7)

        # Helpers conversion de cellule et génération SKU unique
        def _to_str(val):
            if val is None:
                return ''
            if isinstance(val, float):
                try:
                    if val.is_integer():
                        return str(int(val))
                except Exception:
                    pass
                return str(val).strip()
            return str(val).strip()

        # Ensemble de tous les SKU connus (clé normalisée)
        existing_skus = set(sku_index.keys())
        def _generate_next_sku(existing_norm_keys):
            """Génère un nouveau SKU canonique "SKU-XXX" sans collision sur la clé normalisée."""
            max_n = 0
            for s in list(existing_norm_keys):
                if not s:
                    continue
                # Sur les clés normalisées, on attend des formes du type "SKU007"
                m = re.match(r'^SKU0*(\d+)$', s)
                if m:
                    try:
                        max_n = max(max_n, int(m.group(1)))
                    except Exception:
                        pass
            n = max_n + 1
            while True:
                number = f"{n:03d}"
                candidate_norm = f"SKU{number}"
                if candidate_norm not in existing_norm_keys:
                    # SKU stocké en base au format canonique avec tiret
                    return f"SKU-{number}"
                n += 1

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                def _get(idx, default=None):
                    return row[idx] if (idx is not None and idx < len(row)) else default
                sku = _to_str(_get(idx_sku, ''))
                name = _to_str(_get(idx_name, ''))
                _type_raw = _get(idx_type, 'reception')
                type_val = 'reception'
                try:
                    _type_str = _to_str(_type_raw).lower()
                    _type_norm = unicodedata.normalize('NFKD', _type_str).encode('ascii', 'ignore').decode('ascii')
                    _type_norm = re.sub(r'[^a-z]', '', _type_norm)
                    if _type_norm.startswith('recept') or _type_norm in ('reception',):
                        type_val = 'reception'
                    elif _type_norm.startswith('ajust') or _type_norm in ('ajustement', 'adjustement'):
                        type_val = 'ajustement'
                except Exception:
                    type_val = 'reception'
                quantite_raw = _get(idx_qty, None)
                date_raw = _get(idx_date, None)
                fournisseur = (_get(idx_fourn, '') or '').strip()
                pu_ht_raw = _get(idx_puht, 0)
                tva_raw = _get(idx_tva, 20)
                categorie_val = (_get(idx_categorie, '') or '').strip() if idx_categorie is not None else ''
                unite_val = (_get(idx_unite, '') or '').strip() if idx_unite is not None else ''

                # Ignorer les lignes totalement vides (aucun identifiant d'article et aucune donnée utile)
                if (not sku and not name) and (quantite_raw in [None, '']) and (not fournisseur):
                    continue

                # Valider type
                if type_val not in ['reception', 'ajustement']:
                    raise ValueError(f"Type invalide: {type_val} (utilisez 'reception' ou 'ajustement')")

                # Valider quantité
                if quantite_raw is None:
                    raise ValueError('Quantité manquante')
                try:
                    q_str = str(quantite_raw).replace(',', '.')
                    quantity = int(float(q_str))
                except Exception:
                    raise ValueError(f"Quantité invalide: {quantite_raw}")
                if quantity <= 0:
                    raise ValueError('Quantité doit être > 0')

                # PU HT
                try:
                    pu_ht = float(str(pu_ht_raw or 0).replace(',', '.'))
                except Exception:
                    raise ValueError(f"PU HT invalide: {pu_ht_raw}")

                # TVA (%), accepter format '20', '20%', '20,0'
                try:
                    tva_str = str(tva_raw).replace('%', '').replace(',', '.')
                    taux_tva = float(tva_str)
                except Exception:
                    raise ValueError(f"TVA invalide: {tva_raw}")

                # Date
                from datetime import datetime as _dt
                date_reception = None
                if date_raw:
                    if hasattr(date_raw, 'strftime'):
                        # datetime/date Excel
                        date_reception = date_raw.date() if hasattr(date_raw, 'date') else date_raw
                    else:
                        try:
                            date_reception = _dt.strptime(str(date_raw), '%Y-%m-%d').date()
                        except Exception:
                            try:
                                date_reception = _dt.strptime(str(date_raw), '%d/%m/%Y').date()
                            except Exception:
                                raise ValueError(f"Date invalide: {date_raw} (utilisez YYYY-MM-DD)")
                else:
                    date_reception = _dt.utcnow().date()

                # Fournisseur (optionnel → valeur par défaut si vide)
                if not fournisseur:
                    fournisseur = 'Inconnu'

                # Trouver l'article (SKU prioritaire, sinon par nom normalisé)
                item = None
                sku_norm = _norm_sku_key(sku)
                if sku_norm:
                    # Si le SKU existe déjà (peu importe casse / tirets), on réutilise l'article
                    item = sku_index.get(sku_norm)
                if not item and name:
                    item = name_index.get(_norm_name(name))
                # Si introuvable, créer l'article automatiquement pour importer TOUTES les lignes
                if not item:
                    if not name:
                        raise ValueError("Article introuvable et 'Nom Article' manquant pour création")
                    try:
                        # Préserver le SKU fourni s'il est unique (après normalisation),
                        # sinon générer un nouveau SKU vraiment unique.
                        candidate = _to_str(sku).strip()
                        candidate_norm = _norm_sku_key(candidate)
                        if (not candidate_norm) or (candidate_norm in existing_skus):
                            candidate = _generate_next_sku(existing_skus)
                            candidate_norm = _norm_sku_key(candidate)
                        new_item = Item(
                            sku=candidate,
                            name=name.strip(),
                            category=(categorie_val or None),
                            unit=(unite_val or 'Unité'),
                            quantity=0
                        )
                        db.session.add(new_item)
                        db.session.flush()  # obtenir l'ID et rendre disponible dans la session
                        # Mettre à jour les index en mémoire pour les lignes suivantes
                        sku_index[candidate_norm] = new_item
                        name_index[_norm_name(new_item.name)] = new_item
                        existing_skus.add(candidate_norm)
                        item = new_item
                    except Exception as create_err:
                        raise ValueError(f"Création article échouée: {create_err}")

                # Empêcher les doublons exacts de réception (même article, date, quantité, fournisseur)
                if type_val == 'reception':
                    dup_q = Reception.query.filter(
                        Reception.item_id == item.id,
                        Reception.type == 'reception',
                        Reception.quantity == quantity,
                        Reception.reason == fournisseur
                    )
                    if date_reception is not None:
                        dup_q = dup_q.filter(Reception.date_reception == date_reception)
                    else:
                        dup_q = dup_q.filter(Reception.date_reception.is_(None))

                    if dup_q.first():
                        # Ignorer cette ligne d'import : elle correspond à une réception déjà enregistrée
                        continue

                # Calculs budgétaires
                prix = calculer_prix_reception(pu_ht, quantity, taux_tva)

                # Créer la réception
                reception = Reception(
                    item_id=item.id,
                    type=type_val,
                    quantity=quantity,
                    reason=fournisseur,
                    date_reception=date_reception,
                    user_id=session['user_id'],
                    prix_unitaire_ht=pu_ht,
                    taux_tva=taux_tva,
                    prix_unitaire_ttc=prix['pu_ttc'],
                    prix_total_ht=prix['pt_ht'],
                    prix_total_ttc=prix['pt_ttc']
                )

                # Mettre à jour le stock
                if type_val == 'reception':
                    item.quantity = (item.quantity or 0) + quantity
                else:
                    item.quantity = quantity

                db.session.add(reception)
                db.session.flush()

                # Créer le stock FIFO pour les réceptions
                if type_val == 'reception':
                    reception_stock = ReceptionStock(
                        item_id=item.id,
                        reception_id=reception.id,
                        quantite_initiale=quantity,
                        quantite_restante=quantity,
                        prix_unitaire_ht=pu_ht,
                        prix_unitaire_ttc=prix['pu_ttc'],
                        taux_tva=taux_tva,
                        date_reception=date_reception or _dt.utcnow()
                    )
                    db.session.add(reception_stock)

                success += 1

            except Exception as e:
                errors += 1
                details.append({'row': row_idx, 'error': str(e)})
                continue

        db.session.commit()
        # Stocker les détails d'erreurs dans la session (limités) pour affichage optionnel
        try:
            session['import_receptions_result'] = {
                'success': success,
                'errors': errors,
                'details': details[:10]  # limiter à 10 pour l'UI
            }
        except Exception:
            pass
        return redirect(url_for('receptions', op='import', success=success, errors=errors))

    except Exception as e:
        db.session.rollback()
        flash(f"Erreur lors de l'import des réceptions: {str(e)}", 'error')
        return redirect(url_for('add_reception'))

@app.route('/download-template-receptions')
@login_required
def download_template_receptions():
    """Télécharger le modèle Excel pour l'importation des réceptions."""
    flash("Le modèle Excel d'importation des réceptions n'est plus disponible.", 'error')
    return redirect(url_for('receptions'))
    wb = Workbook()
    ws = wb.active
    ws.title = 'Receptions'

    headers = ['SKU', 'Nom Article', 'Catégorie', 'Unité', 'Type', 'Quantité', 'Date Réception', 'Fournisseur', 'PU HT', 'Taux TVA']
    ws.append(headers)

    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    for col, _h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Largeurs colonnes
    widths = [15, 30, 20, 12, 14, 10, 18, 28, 12, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Feuille d'instructions
    ws2 = wb.create_sheet('Instructions')
    instructions = [
        ["INSTRUCTIONS D'IMPORT RÉCEPTIONS"],
        [''],
        ["Colonnes et exigences:"],
        ["- SKU: Code article (prioritaire). Si vide, 'Nom Article' est utilisé"],
        ["- Nom Article: Utilisé si SKU est vide"],
        ["- Catégorie: Optionnelle. Utilisée uniquement lors de la création d'un nouvel article"],
        ["- Unité: Optionnelle. Utilisée uniquement lors de la création d'un nouvel article (ex: Unité, Paquet, Boite)"],
        ["- Type: 'reception' (ajoute) ou 'ajustement' (écrase) — par défaut: reception"],
        ["- Quantité: Nombre entier > 0"],
        ["- Date Réception: YYYY-MM-DD ou DD/MM/YYYY (laisser vide pour aujourd'hui)"],
        ["- Fournisseur: Nom du fournisseur (obligatoire)"],
        ["- PU HT: Prix unitaire HT (ex: 12.5)"],
        ["- Taux TVA: en % (ex: 20 ou 20%)"],
        [''],
        ["Note: Les colonnes sont reconnues par leur titre (l'ordre peut varier)."],
        ["Note: Les montants TTC/HT sont calculés automatiquement."],
    ]
    for r in instructions:
        ws2.append(r)
    ws2['A1'].font = Font(bold=True, size=14, color='2E7D32')
    ws2.column_dimensions['A'].width = 70

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='modele_importation_receptions.xlsx'
    )

@app.route('/receptions/<int:reception_id>/edit', methods=['GET', 'POST'])
@login_required
@modification_required
def edit_reception(reception_id):
    """Modifier une réception et rectifier les données budgétaires"""
    reception = Reception.query.get_or_404(reception_id)
    
    if request.method == 'POST':
        # Sauvegarder les anciennes valeurs pour l'audit
        old_quantity = reception.quantity
        old_reason = reception.reason
        old_date_reception = reception.date_reception
        old_pu_ht = reception.prix_unitaire_ht
        old_tva = reception.taux_tva
        old_pt_ttc = reception.prix_total_ttc
        
        # Récupérer les nouvelles valeurs
        new_quantity = int(request.form['quantity'])
        new_reason = request.form.get('reason', '').strip()
        new_pu_ht = float(request.form.get('pu_ht', 0))
        new_taux_tva = float(request.form.get('taux_tva', 20))
        
        # Date de réception
        date_reception_str = request.form.get('date_reception')
        new_date_reception = None
        if date_reception_str:
            try:
                new_date_reception = datetime.strptime(date_reception_str, '%Y-%m-%d').date()
            except ValueError:
                new_date_reception = datetime.utcnow().date()
        
        if not new_reason:
            flash('Le fournisseur est obligatoire.', 'error')
            items = Item.query.order_by(Item.name).all()
            return render_template('edit_reception.html', reception=reception, items=items)
        
        # Calculer les nouveaux prix
        prix_calcules = calculer_prix_reception(new_pu_ht, new_quantity, new_taux_tva)
        
        # Calculer la différence de quantité pour ajuster le stock
        quantity_diff = new_quantity - old_quantity
        
        try:
            # Mettre à jour la réception
            reception.quantity = new_quantity
            reception.reason = new_reason
            reception.date_reception = new_date_reception
            reception.prix_unitaire_ht = new_pu_ht
            reception.taux_tva = new_taux_tva
            reception.prix_unitaire_ttc = prix_calcules['pu_ttc']
            reception.prix_total_ht = prix_calcules['pt_ht']
            reception.prix_total_ttc = prix_calcules['pt_ttc']
            
            # Ajuster le stock de l'article
            if reception.type == 'reception':
                reception.item.quantity += quantity_diff
            elif reception.type == 'ajustement':
                reception.item.quantity = new_quantity
            
            # Log d'audit
            changes = []
            if old_quantity != new_quantity:
                changes.append(f"Quantité: {old_quantity} → {new_quantity}")
            if old_reason != new_reason:
                changes.append(f"Fournisseur: {old_reason} → {new_reason}")
            if old_date_reception != new_date_reception:
                old_date_str = old_date_reception.strftime('%d/%m/%Y') if old_date_reception else 'Non définie'
                new_date_str = new_date_reception.strftime('%d/%m/%Y') if new_date_reception else 'Non définie'
                changes.append(f"Date réception: {old_date_str} → {new_date_str}")
            if old_pu_ht != new_pu_ht:
                changes.append(f"PU HT: {old_pu_ht:.2f} → {new_pu_ht:.2f} DH")
            if old_tva != new_taux_tva:
                changes.append(f"TVA: {old_tva}% → {new_taux_tva}%")
            if old_pt_ttc != prix_calcules['pt_ttc']:
                changes.append(f"Total TTC: {old_pt_ttc:.2f} → {prix_calcules['pt_ttc']:.2f} DH")
            
            if changes:
                create_audit_log(
                    action='update',
                    entity_type='reception',
                    entity_id=reception.id,
                    entity_name=reception.item.name,
                    description=f"Réception modifiée: {' | '.join(changes)} | Stock actuel: {reception.item.quantity}"
                )
            
            db.session.commit()
            
            # Message flash avec détails pour le modal
            flash(f'✓ Réception modifiée avec succès | {reception.item.name} | Stock actuel: {reception.item.quantity} {reception.item.unit} | Montant: {prix_calcules["pt_ttc"]:.2f} DH TTC', 'success')
            return redirect(url_for('edit_reception', reception_id=reception.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de la modification: {str(e)}', 'error')
    
    items = Item.query.order_by(Item.name).all()
    return render_template('edit_reception.html', reception=reception, items=items)

@app.route('/receptions/<int:reception_id>/delete', methods=['POST'])
@login_required
@modification_required
def delete_reception(reception_id):
    """Supprimer une réception (avec option de suppression de l'article)"""
    reception = Reception.query.get_or_404(reception_id)
    item = Item.query.get(reception.item_id) if reception.item_id else None
    
    # Paramètre pour supprimer l'article aussi
    delete_item_too = request.args.get('delete_item') == 'true'

    # Interdire la suppression des ajustements
    if reception.type == 'ajustement' and not delete_item_too:
        flash('Suppression impossible pour un ajustement de stock. Modifiez l\'enregistrement si nécessaire.', 'error')
        return redirect(url_for('receptions'))

    # Sauvegarder les infos pour le log d'audit
    rec_qty = reception.quantity
    item_id = item.id if item else None
    item_name = item.name if item else 'Réception'
    item_sku = item.sku if item else ''

    # Vérifier l'utilisation FIFO (sauf si on supprime tout l'article)
    if not delete_item_too:
        stocks = ReceptionStock.query.filter_by(reception_id=reception.id).all()
        for stock in stocks:
            used = (stock.quantite_restante is not None and stock.quantite_initiale is not None and stock.quantite_restante < stock.quantite_initiale)
            if used or (hasattr(stock, 'dotation_items') and stock.dotation_items):
                dotations_count = len(stock.dotation_items) if hasattr(stock, 'dotation_items') else 0
                flash(f'Impossible de supprimer : le stock de cette réception ({rec_qty}) a déjà été utilisé dans {dotations_count} dotation(s).', 'error')
                return redirect(url_for('receptions'))

    try:
        if delete_item_too and item:
            # Utiliser la logique de suppression complète de l'article
            # On vérifie quand même s'il y a des dotations (sécurité)
            dots_count = DotationItem.query.filter_by(item_id=item_id).count()
            if dots_count > 0:
                flash(f'Impossible de supprimer l\'article : il est lié à {dots_count} dotation(s). Supprimez d\'abord les dotations.', 'error')
                return redirect(url_for('receptions'))
                
            # Supprimer toutes les dépendances
            DotationItem.query.filter_by(item_id=item_id).delete()
            ReceptionStock.query.filter_by(item_id=item_id).delete()
            Reception.query.filter_by(item_id=item_id).delete()
            db.session.delete(item)
            db.session.commit()
            
            # Réorganiser SKUs
            Item.reorganize_all_skus()
            
            create_audit_log(
                action='delete',
                entity_type='item',
                entity_id=item_id,
                entity_name=item_name,
                description=f"Article et toutes ses réceptions supprimés (SKU: {item_sku})"
            )
            flash(f'✓ L\'article "{item_name}" et toutes ses réceptions ont été supprimés.', 'success')
        else:
            # Suppression simple de la réception
            if item and reception.type == 'reception':
                qty_before = item.quantity or 0
                item.quantity = max(0, qty_before - (rec_qty or 0))
                item.updated_at = datetime.utcnow()

            # Soft Delete: Marquer la réception comme supprimée au lieu de la supprimer réellement
            reception.is_deleted = True
            reception.deleted_at = datetime.utcnow()
            reception.deleted_by = session.get('user_id')
            
            # Supprimer les stocks FIFO liés (On les supprimera réellement car ils seront recréés à la restauration)
            ReceptionStock.query.filter_by(reception_id=reception.id).delete()
            
            db.session.commit()

            create_audit_log(
                action='delete',
                entity_type='reception',
                entity_id=reception_id,
                entity_name=item_name,
                description=f"SKU: {item_sku} | Quantité: {rec_qty}"
            )
            flash('✓ Réception supprimée avec succès', 'success')
            
    except Exception as e:
        db.session.rollback()
        flash(f"Erreur lors de la suppression: {str(e)}", 'error')

    return redirect(url_for('receptions'))

@app.route('/receptions/delete-multiple', methods=['POST'])
@login_required
@modification_required
def delete_receptions_multiple():
    """Suppression multiple des réceptions sélectionnées avec rapport détaillé."""
    delete_items_too = request.args.get('delete_items') == 'true'
    raw_ids = request.form.getlist('reception_ids[]') or request.form.getlist('reception_ids')
    ids = []
    for x in raw_ids:
        try:
            ids.append(int(x))
        except:
            continue
            
    if not ids:
        flash('Aucune réception sélectionnée.', 'warning')
        return redirect(url_for('receptions'))

    deleted = 0
    blocked = 0
    for rid in ids:
        reception = Reception.query.get(rid)
        if not reception:
            continue

        # Bloquer les ajustements
        if reception.type == 'ajustement':
            blocked += 1
            continue

        stocks = ReceptionStock.query.filter_by(reception_id=reception.id).all()
        used = False
        for stock in stocks:
            if (stock.quantite_restante is not None and stock.quantite_initiale is not None and stock.quantite_restante < stock.quantite_initiale) or \
               (hasattr(stock, 'dotation_items') and stock.dotation_items):
                used = True
                break
        if used:
            blocked += 1
            continue

        item = reception.item
        item_id = item.id if item else None

        # Réajuster le stock
        if item and reception.type == 'reception':
            try:
                current_qty = item.quantity or 0
                item.quantity = max(0, current_qty - (reception.quantity or 0))
                item.updated_at = datetime.utcnow()
            except Exception:
                pass

        # Soft Delete: Marquer la réception comme supprimée au lieu de la supprimer réellement
        reception.is_deleted = True
        reception.deleted_at = datetime.utcnow()
        reception.deleted_by = session.get('user_id')
        
        # Supprimer les stocks FIFO liés
        for s in stocks:
            db.session.delete(s)
        deleted += 1
        
        # Optionnel: Supprimer l'article si orphelin
        if delete_items_too and item_id:
            try:
                db.session.flush() 
                other_recs = Reception.query.filter_by(item_id=item_id).first()
                other_dots = DotationItem.query.filter_by(item_id=item_id).first()
                if not other_recs and not other_dots:
                    item_to_del = Item.query.get(item_id)
                    if item_to_del:
                        db.session.delete(item_to_del)
            except:
                pass

    if deleted > 0:
        db.session.commit()
        create_audit_log(
            action='delete',
            entity_type='reception',
            entity_id=0,
            entity_name='SUPPRESSION MULTIPLE',
            description=f'{deleted} réception(s) supprimée(s) (Nettoyage articles: {delete_items_too})'
        )
        
        if blocked == 0:
            flash(f'✓ {deleted} réception(s) supprimée(s) avec succès.', 'success')
        else:
            flash(f'✓ {deleted} réception(s) supprimée(s). {blocked} réception(s) bloquée(s) car le stock est déjà utilisé ou ce sont des ajustements.', 'warning')
    return redirect(url_for('receptions'))

@app.route('/dotations')
@login_required
def dotations():
    """Affiche toutes les dotations avec filtrage par année"""
    page = request.args.get('page', 1, type=int)
    year_filter = request.args.get('year', type=int)

    # Préparer la liste des années disponibles à partir des dates de dotation
    current_year = datetime.utcnow().year
    year_rows = db.session.query(
        func.extract('year', Dotation.date_dotation)
    ).distinct().all()
    available_years_raw = [int(y[0]) for y in year_rows if y[0] is not None]

    if available_years_raw:
        min_year = min(available_years_raw)
        max_data_year = max(available_years_raw)
        max_year = max(max_data_year, current_year)
    else:
        min_year = current_year
        max_year = current_year

    years = list(range(max_year, min_year - 1, -1))

    # Si aucune année n'est spécifiée, utiliser par défaut l'année la plus récente
    if not year_filter and years:
        year_filter = years[0]

    # Calculer les années précédente et suivante pour la navigation (comme sur le budget)
    prev_year = year_filter - 1 if year_filter else None
    next_year = year_filter + 1 if year_filter else None

    query = Dotation.query.filter_by(is_deleted=False)

    if year_filter:
        query = query.filter(
            func.extract('year', Dotation.date_dotation) == year_filter
        )

    dotations_paginated = query.order_by(Dotation.numero_dotation.asc()).paginate(
        page=page, per_page=10, error_out=False
    )

    # Convertir les dates
    dotations_paginated.items = [convert_object_dates(d) for d in dotations_paginated.items]

    return render_template(
        'dotations.html',
        dotations=dotations_paginated,
        years=years,
        current_year=year_filter,
        prev_year=prev_year,
        next_year=next_year,
    )

@app.route('/dotations/add', methods=['GET', 'POST'])
@login_required
@modification_required
def add_dotation():
    """Ajouter une nouvelle dotation avec plusieurs articles"""
    if request.method == 'POST':
        # Nouveau système : récupérer les natures sélectionnées (sélection multiple)
        selected_natures = request.form.getlist('natures[]')
        
        if not selected_natures:
            return jsonify({
                'success': False,
                'message': 'Veuillez sélectionner au moins une nature de prestation.'
            })
        
        # Récupérer l'unité destinataire
        unite_id_str = request.form.get('unite_id', '').strip()
        if not unite_id_str:
            return jsonify({
                'success': False,
                'message': 'L\'unité destinataire est obligatoire.'
            })
        
        unite_id = int(unite_id_str)
        service_id = request.form.get('service_id', '').strip()
        note = request.form.get('note', '').strip()
        
        # Validation : le service est obligatoire UNIQUEMENT pour Sidi Kacem
        unite = Unite.query.get(unite_id)
        if unite and 'SIDI KACEM' in unite.nom.upper():
            if not service_id:
                return jsonify({
                    'success': False,
                    'message': 'Le service est obligatoire pour l\'unité de Sidi Kacem.'
                })
        
        # Gérer le cas spécial des centres de secours
        if service_id == 'centre_secours':
            service_id = None  # Pour les centres de secours
        
        # La catégorie sélectionnée dans le formulaire sera ignorée
        # car nous allons créer des dotations automatiquement par catégorie d'articles
        
        try:
            # Nouveau système : récupérer les quantités et natures des articles
            quantities = request.form.to_dict()
            natures_map = request.form.to_dict()
            
            articles_by_category = {}
            
            # Parcourir tous les champs quantities[item_id]
            for key, value in quantities.items():
                if key.startswith('quantities[') and key.endswith(']'):
                    item_id_str = key[11:-1]  # Extraire l'ID entre 'quantities[' et ']'
                    try:
                        item_id = int(item_id_str)
                        quantite = int(value)
                        
                        if quantite > 0:
                            item = Item.query.get(item_id)
                            if item and item.quantity >= quantite:
                                # Récupérer la nature de l'article depuis natures_map
                                nature_key = f'natures_map[{item_id}]'
                                category = natures_map.get(nature_key, item.category)
                                
                                if category not in articles_by_category:
                                    articles_by_category[category] = []
                                articles_by_category[category].append({
                                    'item': item,
                                    'quantite': quantite
                                })
                            elif item:
                                flash(f'Stock insuffisant pour {item.name}', 'warning')
                            else:
                                flash('Article introuvable', 'error')
                    except (ValueError, KeyError):
                        continue
            
            if not articles_by_category:
                return jsonify({
                    'success': False,
                    'message': 'Aucun article valide sélectionné.'
                })
            
            # Créer une dotation pour chaque catégorie
            dotations_creees = []
            total_items_added = 0
            
            for category, items_list in articles_by_category.items():
                # ============================================
                # VÉRIFICATION BUDGÉTAIRE
                # ============================================
                
                # 1. Calculer le montant TTC de la dotation
                articles_pour_calcul = [(item_data['item'].id, item_data['quantite']) for item_data in items_list]
                montant_dotation_ttc = calculer_montant_dotation_ttc(articles_pour_calcul)
                
                # 2. Vérifier le budget disponible
                annee_courante = datetime.now().year
                verification_budget = verifier_budget_disponible(
                    nature=category,
                    centre_id=unite_id,  # L'unité destinataire
                    montant_ttc=montant_dotation_ttc,
                    annee=annee_courante
                )
                
                # 3. Bloquer si budget insuffisant
                if not verification_budget['disponible']:
                    db.session.rollback()
                    return jsonify({
                        'success': False,
                        'message': f'⚠️ BUDGET INSUFFISANT pour "{category}" !\n\n'
                                   f'Budget alloué: {verification_budget["budget_alloue"]:.2f} DH\n'
                                   f'Déjà consommé: {verification_budget["consommation"]:.2f} DH\n'
                                   f'Reste disponible: {verification_budget["reste"]:.2f} DH\n'
                                   f'Montant demandé: {montant_dotation_ttc:.2f} DH\n\n'
                                   f'Dépassement: {montant_dotation_ttc - verification_budget["reste"]:.2f} DH'
                    })
                
                # ============================================
                # CRÉATION DE LA DOTATION (Budget OK)
                # ============================================
                
                # Générer un numéro de dotation unique pour chaque catégorie
                numero_dotation_cat = Dotation.generate_numero_dotation()
                
                # Créer la dotation pour cette catégorie
                # Convertir service_id en int seulement si c'est un nombre
                service_id_final = None
                service_nom_dotation = None
                if service_id:
                    try:
                        service_id_final = int(service_id)
                    except ValueError:
                        # Si ce n'est pas un nombre (ex: 'bureau_cdt'), c'est Sidi Kacem
                        service_id_final = None
                        # Récupérer le nom du service pour Sidi Kacem
                        services_sidi_kacem = {
                            'bureau_cdt': 'Bureau Cdt',
                            'sce_secretariat': 'Sce Secrétariat',
                            'sce_prevention': 'Sce Prévention',
                            'sce_technique': 'Sce Technique',
                            'sce_intervention': 'Sce Intervention',
                            'standardiste': 'Standardiste'
                        }
                        service_nom_dotation = services_sidi_kacem.get(service_id, service_id)
                
                # Construire la note avec le service Sidi Kacem si applicable
                note_finale = f"Nature de prestation: {category}"
                if service_nom_dotation:
                    note_finale = f"Service: {service_nom_dotation} | {note_finale}"
                if note:
                    note_finale = f"{note} | {note_finale}"
                
                dotation = Dotation(
                    numero_dotation=numero_dotation_cat,
                    unite_id=unite_id,
                    service_id=service_id_final,
                    categorie=category,
                    notes=note_finale,
                    dotee_par=session['user_id']
                )
                
                db.session.add(dotation)
                db.session.flush()  # Pour obtenir l'ID de la dotation
                
                # Ajouter les articles de cette catégorie avec FIFO
                items_added_for_category = 0
                montant_total_dotation_ht = 0.0
                montant_total_dotation_ttc = 0.0
                
                for item_data in items_list:
                    item = item_data['item']
                    quantite = item_data['quantite']
                    
                    # ========================================================================
                    # APPLIQUER LA LOGIQUE FIFO
                    # ========================================================================
                    resultat_fifo = apply_dotation_fifo(item.id, quantite, dotation.id)
                    
                    if not resultat_fifo['success']:
                        # Si FIFO échoue (stock insuffisant), rollback et retourner erreur
                        db.session.rollback()
                        return jsonify({
                            'success': False,
                            'message': f'Erreur FIFO pour {item.name}: {resultat_fifo["message"]}'
                        })
                    
                    # Accumuler les montants
                    montant_total_dotation_ht += resultat_fifo['montant_total_ht']
                    montant_total_dotation_ttc += resultat_fifo['montant_total_ttc']
                    
                    # Diminuer le stock de l'article
                    item.quantity -= quantite
                    item.updated_at = datetime.utcnow()
                    items_added_for_category += 1
                    total_items_added += 1
                
                if items_added_for_category > 0:
                    dotations_creees.append({
                        'numero': numero_dotation_cat,
                        'categorie': category,
                        'nb_articles': items_added_for_category
                    })
                    
                    # Log d'audit pour la dotation créée
                    unite = Unite.query.get(unite_id)
                    
                    # Gérer le nom du service (DB ou Sidi Kacem)
                    service_nom = None
                    if service_id:
                        # Essayer de récupérer depuis la DB
                        service_obj = Service.query.get(service_id) if service_id else None
                        if service_obj:
                            service_nom = service_obj.nom
                    elif request.form.get('service_id'):
                        # Si c'est un service de Sidi Kacem (string), utiliser le texte du select
                        service_value = request.form.get('service_id')
                        services_sidi_kacem = {
                            'bureau_cdt': 'Bureau Cdt',
                            'sce_secretariat': 'Sce Secrétariat',
                            'sce_prevention': 'Sce Prévention',
                            'sce_technique': 'Sce Technique',
                            'sce_intervention': 'Sce Intervention',
                            'standardiste': 'Standardiste'
                        }
                        service_nom = services_sidi_kacem.get(service_value, service_value)
                    
                    destinataire = f"{unite.nom if unite else 'Unité'}"
                    if service_nom:
                        destinataire += f" - {service_nom}"
                    
                    create_audit_log(
                        action='create',
                        entity_type='dotation',
                        entity_id=dotation.id,
                        entity_name=numero_dotation_cat,
                        description=f"{category} | {items_added_for_category} article(s) | {destinataire} | Montant: {montant_dotation_ttc:.2f} DH TTC"
                    )
                    
                    # ============================================
                    # CRÉER LA CONSOMMATION BUDGÉTAIRE
                    # ============================================
                    consommation = ConsommationBudget(
                        nature=category,
                        centre_id=unite_id,  # L'unité destinataire
                        centre_nom=unite.nom if unite else 'Unité inconnue',
                        montant_ttc=montant_dotation_ttc,
                        dotation_id=dotation.id,
                        annee=annee_courante
                    )
                    db.session.add(consommation)
            
            if total_items_added > 0:
                db.session.commit()
                
                # Récupérer l'unité et le service pour le message
                unite = Unite.query.get(unite_id)
                
                # Gérer le nom du service (DB ou Sidi Kacem)
                service_nom_msg = None
                if service_id:
                    service_obj = Service.query.get(service_id)
                    if service_obj:
                        service_nom_msg = service_obj.nom
                elif request.form.get('service_id'):
                    service_value = request.form.get('service_id')
                    services_sidi_kacem = {
                        'bureau_cdt': 'Bureau Cdt',
                        'sce_secretariat': 'Sce Secrétariat',
                        'sce_prevention': 'Sce Prévention',
                        'sce_technique': 'Sce Technique',
                        'sce_intervention': 'Sce Intervention',
                        'standardiste': 'Standardiste'
                    }
                    service_nom_msg = services_sidi_kacem.get(service_value, service_value)
                
                destinataire_msg = f"{unite.nom if unite else 'Unité'}"
                if service_nom_msg:
                    destinataire_msg += f" - {service_nom_msg}"
                
                # Retourner JSON pour le modal de succès
                return jsonify({
                    'success': True,
                    'dotations': dotations_creees,
                    'unite': destinataire_msg,
                    'message': f'{len(dotations_creees)} dotation(s) créée(s) avec succès'
                })
            else:
                db.session.rollback()
                return jsonify({
                    'success': False,
                    'message': 'Aucun article n\'a pu être doté. Vérifiez les stocks disponibles.'
                })
                
        except Exception as e:
            db.session.rollback()
            print(f"ERREUR DOTATION: {str(e)}")  # Log console
            import traceback
            traceback.print_exc()  # Stack trace complet
            return jsonify({
                'success': False,
                'message': f'Erreur lors de la création de la dotation: {str(e)}'
            })
    
    # Récupérer les données pour le formulaire (GET)
    unites = Unite.query.all()
    
    return render_template('add_dotation.html', unites=unites)

@app.route('/dotations/delete-all', methods=['GET', 'POST'])
@login_required
@modification_required
def delete_all_dotations():
    """Route pour supprimer TOUTES les dotations (Danger)"""
    user = User.query.get(session['user_id'])
    if not user or not user.has_permission('admin'):
        flash('Action non autorisée.', 'error')
        return redirect(url_for('dotations'))
        
    try:
        # Récupérer toutes les dotations
        all_dotations = Dotation.query.all()
        count = len(all_dotations)
        
        if count == 0:
            flash('Aucune dotation à supprimer.', 'info')
            return redirect(url_for('dotations'))
            
        # Supprimer les dépendances d'abord si nécessaire (bien que cascade devrait gérer)
        # Supprimer consommations budgétaires liées aux dotations
        ConsommationBudget.query.filter(ConsommationBudget.dotation_id.isnot(None)).delete(synchronize_session=False)
        
        # Supprimer les items de dotation
        DotationItem.query.delete()
        
        # Supprimer les dotations
        Dotation.query.delete()
        
        db.session.commit()
        
        create_audit_log(
            action='delete_all',
            entity_type='dotation',
            entity_id=None,
            entity_name='ALL',
            description=f"Suppression forcée de toutes les dotations ({count})"
        )
        
        flash(f'Toutes les dotations ({count}) ont été supprimées avec succès.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression globale: {str(e)}', 'error')
        print(f"Error deleting all dotations: {e}")
        
    return redirect(url_for('dotations'))

@app.route('/dotations/delete/<int:dotation_id>', methods=['POST'])
@login_required
@modification_required
def delete_dotation(dotation_id):
    """Supprimer une dotation et remettre les quantités en stock"""
    dotation = Dotation.query.get_or_404(dotation_id)
    
    # Vérifier si la dotation est livrée
    if dotation.statut == 'livree':
        flash('Impossible de supprimer une dotation livrée. La décharge a déjà été générée.', 'error')
        return redirect(url_for('dotations'))
    
    try:
        # Sauvegarder les infos pour le log avant suppression
        numero_dotation = dotation.numero_dotation
        unite_nom = dotation.unite.nom if dotation.unite else "Unité inconnue"
        nb_articles = len(dotation.items)
        
        # Remettre toutes les quantités en stock
        for dotation_item in dotation.items:
            item = dotation_item.item
            item.quantity += dotation_item.quantite_dotee
            item.updated_at = datetime.utcnow()
        
        # Soft Delete: Marquer la dotation comme supprimée au lieu de la supprimer réellement
        dotation.is_deleted = True
        dotation.deleted_at = datetime.utcnow()
        dotation.deleted_by = session.get('user_id')
        
        # Supprimer également la consommation budgétaire liée à cette dotation (On la supprimera réellement car c'est un record de transaction)
        ConsommationBudget.query.filter_by(dotation_id=dotation.id).delete()
        
        db.session.commit()
        
        # Log d'audit pour la suppression de dotation
        create_audit_log(
            action='delete',
            entity_type='dotation',
            entity_id=dotation_id,
            entity_name=numero_dotation,
            description=f"Dotation supprimée: {numero_dotation} - {unite_nom} ({nb_articles} article(s)) - Quantités remises en stock"
        )
        
        flash(f'✓ Dotation supprimée avec succès | {numero_dotation} | {nb_articles} article(s) remis en stock | Unité: {unite_nom}', 'delete_success')
    except Exception as e:
        db.session.rollback()
        flash('Erreur lors de la suppression de la dotation.', 'error')
    
    return redirect(url_for('dotations'))



@app.route('/dotations/<int:dotation_id>')
@login_required
def view_dotation(dotation_id):
    """Voir les détails d'une dotation"""
    dotation = Dotation.query.get_or_404(dotation_id)
    return render_template('view_dotation.html', dotation=dotation)

@app.route('/dotations/<int:dotation_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_dotation(dotation_id):
    """Modifier une dotation (quantités et prix)"""
    dotation = Dotation.query.get_or_404(dotation_id)
    
    # Empêcher la modification des dotations livrées
    if dotation.statut == 'livree':
        flash('Impossible de modifier une dotation déjà livrée.', 'error')
        return redirect(url_for('dotations'))
    
    if request.method == 'POST':
        try:
            # Mettre à jour les quantités et prix de chaque article
            for dotation_item in dotation.items:
                qty_key = f'quantity_{dotation_item.id}'
                price_key = f'price_{dotation_item.id}'
                
                if qty_key in request.form:
                    new_quantity = int(request.form[qty_key])
                    dotation_item.quantite_dotee = new_quantity
                
                if price_key in request.form:
                    new_price = float(request.form[price_key])
                    dotation_item.prix_unitaire = new_price
            
            # Mettre à jour la consommation budgétaire associée à cette dotation
            consommation = ConsommationBudget.query.filter_by(dotation_id=dotation.id).first()
            if consommation:
                consommation.montant_ttc = dotation.cout_total_dotation
            
            db.session.commit()
            
            # Log d'audit
            create_audit_log(
                action='update',
                entity_type='dotation',
                entity_id=dotation.id,
                entity_name=dotation.numero_dotation,
                description=f'Dotation modifiée | Coût total: {dotation.cout_total_dotation:.2f} MAD'
            )
            
            flash(f'✓ Dotation {dotation.numero_dotation} modifiée avec succès | Coût total: {dotation.cout_total_dotation:.2f} MAD', 'success')
            return redirect(url_for('dotations'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de la modification: {str(e)}', 'error')
    
    return render_template('edit_dotation.html', dotation=dotation)

@app.route('/api/dotation/<int:dotation_id>/status', methods=['POST'])
@login_required
def update_dotation_status(dotation_id):
    """Mettre à jour le statut d'une dotation"""
    dotation = Dotation.query.get_or_404(dotation_id)
    data = request.get_json()
    
    try:
        old_statut = dotation.statut
        new_statut = data['statut']
        dotation.statut = new_statut
        db.session.commit()
        
        # Log d'audit pour la modification du statut
        statut_labels = {
            'en_cours': 'En cours',
            'validee': 'Validée',
            'livree': 'Livrée'
        }
        create_audit_log(
            action='update',
            entity_type='dotation',
            entity_id=dotation.id,
            entity_name=dotation.numero_dotation,
            description=f"Statut modifié: {statut_labels.get(old_statut, old_statut)} → {statut_labels.get(new_statut, new_statut)}"
        )
        
        return jsonify({'success': True, 'message': 'Statut mis à jour'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Erreur lors de la mise à jour'}), 400

@app.route('/api/dotation/<int:dotation_id>/articles')
@login_required
def api_dotation_articles(dotation_id):
    """API pour récupérer les articles d'une dotation"""
    dotation = Dotation.query.get_or_404(dotation_id)
    
    return jsonify([{
        'name': item.item.name,
        'sku': item.item.sku,
        'quantite_dotee': item.quantite_dotee,
        'unit': item.item.unit
    } for item in dotation.items])

@app.route('/dotations/<int:dotation_id>/decharge')
@login_required
def generate_decharge(dotation_id):
    """Générer le bon de décharge PDF pour une dotation livrée"""
    dotation = Dotation.query.get_or_404(dotation_id)
    
    # Vérifier que la dotation est livrée
    if dotation.statut != 'livree':
        flash('Le bon de décharge ne peut être généré que pour les dotations livrées.', 'error')
        return redirect(url_for('view_dotation', dotation_id=dotation_id))
    
    # Vérifier que tous les articles appartiennent à la même catégorie
    categories = set()
    for dotation_item in dotation.items:
        if dotation_item.item.category:
            categories.add(dotation_item.item.category)
    
    if len(categories) > 1:
        categories_list = ', '.join(sorted(categories))
        flash(f'Erreur : La décharge ne peut pas contenir des articles de catégories différentes. Catégories trouvées : {categories_list}', 'error')
        return redirect(url_for('view_dotation', dotation_id=dotation_id))
    
    if len(categories) == 0:
        flash('Erreur : Aucune catégorie définie pour les articles de cette dotation.', 'error')
        return redirect(url_for('view_dotation', dotation_id=dotation_id))
    
    # Préparer les données pour notre nouvelle fonction
    dotation_category = list(categories)[0]  # La catégorie unique
    
    # Créer les données du tableau
    table_data = [['N°', 'Désignations', 'Unité de mesure', 'Quantité']]
    
    for i, dotation_item in enumerate(dotation.items, 1):
        table_data.append([
            f'{i:02d}',
            dotation_item.item.name.upper(),
            dotation_item.item.unit,
            f'{dotation_item.quantite_dotee:02d}'
        ])
    
    # Logique conditionnelle pour le service
    service_destinataire = ""
    if "CENTRE DE SECOURS" in dotation.unite.nom.upper():
        service_destinataire = dotation.unite.nom
    elif dotation.service:
        service_destinataire = f"{dotation.unite.nom} - {dotation.service.nom}"
    
    # Utiliser notre nouvelle fonction améliorée
    try:
        file_name = create_decharge_pdf(dotation_category, table_data, 
                                      unite_destinataire=dotation.unite.nom,
                                      numero_dotation=dotation.numero_dotation,
                                      service_destinataire=service_destinataire)
        
        # Lire le fichier généré et le retourner
        with open(file_name, 'rb') as f:
            pdf_data = f.read()
        
        # Supprimer le fichier temporaire
        os.remove(file_name)
        
        # Créer un buffer avec les données PDF
        buffer = BytesIO(pdf_data)
        buffer.seek(0)
        
        # Nom du fichier de téléchargement
        filename = f"decharge_{dotation.numero_dotation}_{datetime.now().strftime('%Y%m%d')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        flash(f'Erreur lors de la génération du PDF : {str(e)}', 'error')
        return redirect(url_for('view_dotation', dotation_id=dotation_id))

@app.route('/dotations/<int:dotation_id>/import-decharge', methods=['POST'])
@login_required
def import_decharge_signee(dotation_id):
    """Importer un fichier PDF de décharge signée pour une dotation"""
    dotation = Dotation.query.get_or_404(dotation_id)
    
    # Vérifier que la dotation est livrée
    if dotation.statut != 'livree':
        if request.is_json:
            return jsonify({'success': False, 'message': 'Le PDF signé ne peut être importé que pour les dotations livrées.'}), 400
        flash('Le PDF signé ne peut être importé que pour les dotations livrées.', 'error')
        return redirect(url_for('view_dotation', dotation_id=dotation_id))
    
    # Vérifier si un fichier a été envoyé
    if 'decharge_pdf' not in request.files:
        if request.is_json:
            return jsonify({'success': False, 'message': 'Aucun fichier n\'a été sélectionné.'}), 400
        flash('Aucun fichier n\'a été sélectionné.', 'error')
        return redirect(url_for('view_dotation', dotation_id=dotation_id))
    
    file = request.files['decharge_pdf']
    
    # Vérifier si le fichier a un nom
    if file.filename == '':
        if request.is_json:
            return jsonify({'success': False, 'message': 'Aucun fichier sélectionné.'}), 400
        flash('Aucun fichier sélectionné.', 'error')
        return redirect(url_for('view_dotation', dotation_id=dotation_id))
    
    # Vérifier l'extension du fichier
    if not file.filename.lower().endswith('.pdf'):
        if request.is_json:
            return jsonify({'success': False, 'message': 'Le fichier doit être au format PDF.'}), 400
        flash('Le fichier doit être au format PDF.', 'error')
        return redirect(url_for('view_dotation', dotation_id=dotation_id))
    
    try:
        # Créer le répertoire uploads s'il n'existe pas
        upload_dir = os.path.join(app.root_path, 'uploads', 'decharges_signees')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Générer un nom de fichier unique
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        original_filename = secure_filename(file.filename)
        filename = f"decharge_{dotation.numero_dotation}_{timestamp}_{original_filename}"
        file_path = os.path.join(upload_dir, filename)
        
        # Sauvegarder le fichier
        file.save(file_path)
        
        # Mettre à jour la dotation avec les informations du fichier
        dotation.decharge_signee_path = file_path
        dotation.decharge_signee_filename = original_filename
        dotation.decharge_signee_date = datetime.utcnow()
        
        db.session.commit()
        
        # Log d'audit
        create_audit_log(
            action='update',
            entity_type='dotation',
            entity_id=dotation.id,
            entity_name=dotation.numero_dotation,
            description=f'PDF signé importé | Fichier: {original_filename}'
        )
        
        if request.is_json:
            return jsonify({
                'success': True, 
                'message': f'✓ PDF signé importé avec succès | Fichier: {original_filename}',
                'filename': original_filename,
                'date': dotation.decharge_signee_date.strftime('%d/%m/%Y %H:%M')
            })
        
        flash(f'✓ PDF signé importé avec succès | Fichier: {original_filename}', 'success')
        return redirect(url_for('view_dotation', dotation_id=dotation_id))
        
    except Exception as e:
        db.session.rollback()
        if request.is_json:
            return jsonify({'success': False, 'message': f'Erreur lors de l\'import: {str(e)}'}), 500
        flash(f'Erreur lors de l\'import: {str(e)}', 'error')
        return redirect(url_for('view_dotation', dotation_id=dotation_id))

@app.route('/dotations/<int:dotation_id>/telecharger-decharge')
@login_required
def telecharger_decharge_signee(dotation_id):
    """Télécharger le fichier PDF de décharge signée"""
    dotation = Dotation.query.get_or_404(dotation_id)
    
    # Vérifier si un PDF signé existe
    if not dotation.decharge_signee_path or not os.path.exists(dotation.decharge_signee_path):
        flash('Aucun PDF signé n\'est disponible pour cette dotation.', 'error')
        return redirect(url_for('view_dotation', dotation_id=dotation_id))
    
    try:
        # Créer un buffer avec les données du fichier
        with open(dotation.decharge_signee_path, 'rb') as f:
            pdf_data = f.read()
        
        buffer = BytesIO(pdf_data)
        buffer.seek(0)
        
        # Nom du fichier de téléchargement
        filename = f"decharge_signee_{dotation.numero_dotation}_{dotation.decharge_signee_filename}"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        flash(f'Erreur lors du téléchargement: {str(e)}', 'error')
        return redirect(url_for('view_dotation', dotation_id=dotation_id))

@app.route('/api/services/<int:unite_id>')
@login_required
def api_services(unite_id):
    """API pour récupérer les services d'une unité"""
    services = Service.query.filter_by(unite_id=unite_id).all()
    return jsonify([{
        'id': service.id,
        'nom': service.nom
    } for service in services])


# API endpoints for AJAX
@app.route('/api/items')
@login_required
def api_items():
    items = Item.query.all()
    return jsonify([{
        'id': item.id,
        'name': item.name,
        'sku': item.sku,
        'quantity': item.quantity
    } for item in items])

@app.route('/api/items/<int:item_id>/update', methods=['POST'])
@login_required
def api_update_item(item_id):
    item = Item.query.get_or_404(item_id)
    data = request.get_json()
    
    try:
        # Seuls la description et le seuil de commande sont modifiables
        # Le SKU, nom, catégorie et unité ne peuvent plus être modifiés
        if 'description' in data:
            item.description = data['description']
        if 'reorder_level' in data:
            item.reorder_level = int(data['reorder_level'])
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Article mis à jour'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Erreur lors de la mise à jour'}), 400

@app.route('/api/check-duplicate', methods=['POST'])
@login_required
def check_duplicate():
    """API pour vérifier si un article existe déjà (nom + catégorie)"""
    data = request.get_json()
    
    if not data or 'name' not in data or 'category' not in data:
        return jsonify({'error': 'Données invalides'}), 400
    
    item_name = data['name'].strip()
    category = data['category']
    item_id = data.get('item_id')  # Pour l'édition, exclure l'article actuel
    
    # Liste des sous-catégories d'Entretien bâtiments administratifs
    sous_categories_entretien = [
        'Maintenance informatique',
        'Maintenance électrique',
        'Maintenance plomberie',
        'Maintenance climatisation',
        'Maintenance ascenseurs',
        'Maintenance groupes électrogènes',
        'Maintenance systèmes de sécurité'
    ]
    
    # Si c'est une sous-catégorie d'entretien, on vérifie dans toutes les sous-catégories
    if category in sous_categories_entretien:
        query = Item.query.filter(
            db.func.lower(Item.name) == db.func.lower(item_name),
            db.or_(
                Item.category == category,  # La sous-catégorie exacte
                Item.category == 'Entretien bâtiments administratifs'  # Ou la catégorie parente
            )
        )
    else:
        # Pour les autres catégories, vérification normale
        query = Item.query.filter(
            db.func.lower(Item.name) == db.func.lower(item_name),
            Item.category == category
        )
    
    # Si c'est une édition, exclure l'article actuel
    if item_id:
        query = query.filter(Item.id != item_id)
    
    existing_item = query.first()
    
    if existing_item:
        return jsonify({
            'exists': True,
            'sku': existing_item.sku,
            'name': existing_item.name,
            'category': existing_item.category,
            'is_duplicate': True
        })
    else:
        return jsonify({'exists': False})



@app.route('/export/stock-excel-form')
@login_required
def export_stock_excel_form():
    """Formulaire de sélection pour l'export Excel"""
    return render_template('export_form.html')

@app.route('/export/stock-excel', methods=['GET', 'POST'])
@login_required
def export_stock_excel():
    """Exporter l'état du stock en Excel selon les critères sélectionnés"""
    
    # Récupérer les paramètres de filtrage
    if request.method == 'POST':
        date_debut = request.form.get('date_debut')
        date_fin = request.form.get('date_fin')
        type_etat = request.form.get('type_etat', 'global')
        inclure_stats = 'inclure_stats' in request.form
        inclure_dotations = 'inclure_dotations' in request.form
    else:
        # Valeurs par défaut pour GET
        date_debut = None
        date_fin = None
        type_etat = request.args.get('type_etat', 'global')
        inclure_stats = request.args.get('inclure_stats') == 'true'
        inclure_dotations = request.args.get('inclure_dotations') == 'true'
    
    try:
        # Créer un nouveau classeur Excel
        wb = Workbook()
        
        # === FEUILLE 1: ÉTAT DU STOCK ===
        ws_stock = wb.active
        
        # Titre selon le type d'état sélectionné
        if type_etat == 'global':
            ws_stock.title = "État Global"
            titre = "RAPPORT GLOBAL DU STOCK"
        elif type_etat == 'ruptures':
            ws_stock.title = "Ruptures Stock"
            titre = "ARTICLES EN RUPTURE DE STOCK"
        elif type_etat == 'stock_bas':
            ws_stock.title = "Stock Bas"
            titre = "ARTICLES EN STOCK BAS"
        else:
            ws_stock.title = "État du Stock"
            titre = "RAPPORT DU STOCK"
        
        # En-tête principal
        ws_stock.merge_cells('A1:I1')
        ws_stock['A1'] = f"ÉTAT DU STOCK - {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws_stock['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws_stock['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws_stock['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws_stock.row_dimensions[1].height = 30
        
        # En-têtes des colonnes
        headers = ['SKU', 'Nom Article', 'Description', 'Catégorie', 
                  'Quantité', 'Unité', 'Seuil', 'Statut', 'Dernière MAJ']
        
        for col, header in enumerate(headers, 1):
            cell = ws_stock.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Filtrer les articles selon le type d'état sélectionné
        if type_etat == 'ruptures':
            items = Item.query.filter(Item.quantity == 0).order_by(Item.sku).all()
        elif type_etat == 'stock_bas':
            items = Item.query.filter(Item.quantity <= Item.reorder_level, Item.quantity > 0).order_by(Item.sku).all()
        else:  # global
            items = Item.query.order_by(Item.sku).all()
        
        # Filtrer par date si spécifiée (pour les articles créés/modifiés)
        if date_debut:
            try:
                date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d')
                items = [item for item in items if item.created_at >= date_debut_obj]
            except ValueError:
                pass
        
        if date_fin:
            try:
                date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d') + timedelta(days=1)
                items = [item for item in items if item.created_at < date_fin_obj]
            except ValueError:
                pass
        
        # Remplir les données
        for row, item in enumerate(items, 4):
            # Déterminer le statut
            if item.quantity == 0:
                statut = "RUPTURE"
                statut_color = "E74C3C"
            elif item.quantity <= item.reorder_level:
                statut = "STOCK BAS"
                statut_color = "F39C12"
            else:
                statut = "EN STOCK"
                statut_color = "27AE60"
            
            # Données de la ligne
            row_data = [
                item.sku,
                item.name,
                item.description[:50] + '...' if item.description and len(item.description) > 50 else (item.description or ''),
                item.category or '',
                item.quantity,
                item.unit,
                item.reorder_level,
                statut,
                item.updated_at.strftime('%d/%m/%Y %H:%M') if item.updated_at else ''
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws_stock.cell(row=row, column=col, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Coloration du statut
                if col == 9:  # Colonne Statut
                    cell.fill = PatternFill(start_color=statut_color, end_color=statut_color, fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center")
                
                # Alignement des nombres
                if col in [6, 8]:  # Quantité et Seuil
                    cell.alignment = Alignment(horizontal="center")
        
        # Ajuster la largeur des colonnes
        column_widths = [12, 25, 35, 15, 10, 8, 8, 12, 18]
        for col, width in enumerate(column_widths, 1):
            ws_stock.column_dimensions[get_column_letter(col)].width = width
        
        # === FEUILLE 2: STATISTIQUES (Conditionnelle) ===
        if inclure_stats:
            ws_stats = wb.create_sheet(title="Statistiques")
            
            # Calculer les statistiques
            total_items = len(items)
            items_en_stock = len([i for i in items if i.quantity > i.reorder_level])
            items_stock_bas = len([i for i in items if 0 < i.quantity <= i.reorder_level])
            items_rupture = len([i for i in items if i.quantity == 0])
            
            categories_ids = request.form.getlist('categories')
            categories_objets = Categorie.query.filter(Categorie.id.in_(categories_ids)).all()
            
            # Categories
            categories = {}
            for item in items:
                cat = item.category or 'Non classé'
                if cat not in categories:
                    categories[cat] = {'total': 0, 'quantite': 0}
                categories[cat]['total'] += 1
                categories[cat]['quantite'] += item.quantity
            
            # En-tête statistiques
            ws_stats.merge_cells('A1:D1')
            ws_stats['A1'] = "STATISTIQUES DU STOCK"
            ws_stats['A1'].font = Font(size=14, bold=True, color="FFFFFF")
            ws_stats['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            ws_stats['A1'].alignment = Alignment(horizontal="center")
        
            # Résumé général
            stats_data = [
                ['', 'RÉSUMÉ GÉNÉRAL', '', ''],
                ['Total Articles', total_items, '', ''],
                ['Articles en Stock', items_en_stock, f'{(items_en_stock/total_items*100):.1f}%' if total_items > 0 else '0%', ''],
                ['Stock Bas', items_stock_bas, f'{(items_stock_bas/total_items*100):.1f}%' if total_items > 0 else '0%', ''],
                ['Rupture de Stock', items_rupture, f'{(items_rupture/total_items*100):.1f}%' if total_items > 0 else '0%', ''],
                ['', '', '', ''],
                ['', 'PAR CATÉGORIE', '', ''],
                ['Catégorie', 'Nb Articles', 'Quantité Total', 'Moyenne']
            ]
            
            # Ajouter les données de catégories
            for cat, data in categories.items():
                moyenne = data['quantite'] / data['total'] if data['total'] > 0 else 0
                stats_data.append([cat, data['total'], data['quantite'], f'{moyenne:.1f}'])
            
            # Remplir les statistiques
            for row, row_data in enumerate(stats_data, 3):
                for col, value in enumerate(row_data, 1):
                    cell = ws_stats.cell(row=row, column=col, value=value)
                    
                    # Style pour les en-têtes de section
                    if row_data[1] in ['RÉSUMÉ GÉNÉRAL', 'PAR CATÉGORIE']:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                    
                    # Style pour les en-têtes de colonnes
                    if row_data[0] == 'Catégorie':
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
            
            # Ajuster les colonnes des statistiques
            ws_stats.column_dimensions['A'].width = 20
            ws_stats.column_dimensions['B'].width = 15
            ws_stats.column_dimensions['C'].width = 15
            ws_stats.column_dimensions['D'].width = 12
        
        # === FEUILLE 3: DOTATIONS (Conditionnelle) ===
        if inclure_dotations:
            ws_dotations = wb.create_sheet(title="Dotations")
        
            # En-tête dotations
            ws_dotations.merge_cells('A1:H1')
            ws_dotations['A1'] = "DOTATIONS PAR UNITÉ"
            ws_dotations['A1'].font = Font(size=14, bold=True, color="FFFFFF")
            ws_dotations['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            ws_dotations['A1'].alignment = Alignment(horizontal="center")
            
            # En-têtes des dotations
            dot_headers = ['Numéro', 'Unité', 'Service', 'Catégorie', 'Nb Articles', 'Date', 'Statut', 'Dotée par']
            for col, header in enumerate(dot_headers, 1):
                cell = ws_dotations.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Récupérer les dotations
            dotations = Dotation.query.order_by(Dotation.date_dotation.desc()).all()
            
            for row, dot in enumerate(dotations, 4):
                dot_data = [
                    dot.numero_dotation,
                    dot.unite.nom,
                    dot.service.nom if dot.service else 'Toute l\'unité',
                    dot.categorie,
                    len(dot.items),
                    dot.date_dotation.strftime('%d/%m/%Y'),
                    dot.statut.replace('_', ' ').title(),
                    dot.user.name if dot.user else 'Système'
                ]
                
                for col, value in enumerate(dot_data, 1):
                    ws_dotations.cell(row=row, column=col, value=value)
            
            # Ajuster les colonnes des dotations
            dot_widths = [20, 25, 20, 20, 12, 12, 15, 20]
            for col, width in enumerate(dot_widths, 1):
                ws_dotations.column_dimensions[get_column_letter(col)].width = width
        
        # Sauvegarder en mémoire
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nom du fichier avec timestamp et type d'état
        type_suffix = {
            'global': 'global',
            'ruptures': 'ruptures',
            'stock_bas': 'stock_bas'
        }.get(type_etat, 'export')
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"stock_{type_suffix}_{timestamp}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Erreur lors de l\'export: {str(e)}', 'error')
        # API des services : Cette fonction utilise l'API de Flask pour envoyer un fichier Excel en réponse à une requête HTTP
        return redirect(url_for('items'))

# ==================== ROUTES AVIS D'ACHAT ====================

@app.route('/avis-achats')
@login_required
def avis_achats():
    """Liste des avis d'achat avec filtrage par nature et année"""
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    # Filtrer par nature de prestation si spécifié
    nature_filter = request.args.get('nature', '')
    year_filter = request.args.get('year', type=int)
    
    query = AvisAchat.query.filter_by(is_deleted=False)
    if nature_filter:
        query = query.filter_by(nature_prestation=nature_filter)
    
    # Préparer la liste des années disponibles
    current_year = datetime.utcnow().year
    year_rows = db.session.query(
        func.extract('year', AvisAchat.date_creation)
    ).distinct().all()
    available_years_raw = [int(y[0]) for y in year_rows if y[0] is not None]

    if available_years_raw:
        min_year = min(available_years_raw)
        max_data_year = max(available_years_raw)
        max_year = max(max_data_year, current_year)
    else:
        min_year = current_year
        max_year = current_year

    years = list(range(max_year, min_year - 1, -1))

    # Si aucune année n'est spécifiée, utiliser par défaut l'année la plus récente
    if not year_filter and years:
        year_filter = years[0]

    # Calculer les années précédente et suivante pour la navigation
    prev_year = year_filter - 1 if year_filter and (year_filter - 1) >= min_year else None
    next_year = year_filter + 1 if year_filter and (year_filter + 1) <= max_year else None

    # Appliquer le filtre d'année si présent
    if year_filter:
        query = query.filter(
            func.extract('year', AvisAchat.date_creation) == year_filter
        )
    
    avis_list = query.order_by(AvisAchat.date_creation.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Récupérer les catégories uniques pour le filtre
    categories = db.session.query(Item.category).distinct().filter(Item.category.isnot(None)).all()
    categories = [c[0] for c in categories if c[0]]
    
    return render_template('avis_achats.html', 
                         avis_list=avis_list, 
                         categories=categories,
                         nature_filter=nature_filter,
                         years=years,
                         current_year=year_filter,
                         prev_year=prev_year,
                         next_year=next_year)

@app.route('/avis-achats/nouveau', methods=['GET', 'POST'])
@login_required
def add_avis_achat():
    """Créer un ou plusieurs avis d'achat (sélection multiple)"""
    if request.method == 'POST':
        try:
            # Récupérer les natures sélectionnées (NOUVEAU - sélection multiple)
            natures_selectionnees = request.form.getlist('natures[]')
            
            if not natures_selectionnees:
                flash('Veuillez sélectionner au moins une nature de prestation.', 'error')
                return redirect(url_for('add_avis_achat'))
            
            # Récupérer tous les articles sélectionnés
            selected_items = request.form.getlist('selected_items[]')
            
            if not selected_items:
                flash('Veuillez sélectionner au moins un article.', 'error')
                return redirect(url_for('add_avis_achat'))
            
            # Récupérer les informations complémentaires (partagées entre tous les avis)
            delai_execution = request.form.get('delai_execution', '').strip()
            date_limite_reception = request.form.get('date_limite_reception', '').strip()
            
            # Grouper les articles par nature
            articles_par_nature = {}
            for item_id in selected_items:
                nature = request.form.get(f'nature_{item_id}')
                if nature and nature in natures_selectionnees:
                    if nature not in articles_par_nature:
                        articles_par_nature[nature] = []
                    articles_par_nature[nature].append(item_id)
            
            # Créer un avis d'achat pour chaque nature
            avis_crees = []
            
            for nature, item_ids in articles_par_nature.items():
                # Générer le numéro d'avis
                numero_avis = AvisAchat.generate_numero_avis()
                
                # Auto-générer l'objet de la prestation
                if nature in ['Articles de plomberies', 'Articles électriques', 'Articles de la peinture']:
                    objet_prestation = f"Achat d'{nature} pour l'entretien et réparation des bâtiments administratifs des services du Commandement Provincial de la Protection Civile de Sidi Kacem."
                elif nature == "Alimentation à usage Humaine":
                    objet_prestation = "Achat de denrées alimentaires pour les pauses café à l'occasion des sessions de formation pour les services de la Protection Civile."
                else:
                    objet_prestation = f"Achat de {nature} pour les services du Commandement Provincial de la Protection Civile de Sidi Kacem."
                
                # Créer l'avis
                avis = AvisAchat(
                    numero_avis=numero_avis,
                    nature_prestation=nature,
                    objet_prestation=objet_prestation,
                    delai_execution=delai_execution,
                    date_limite_reception=date_limite_reception,
                    created_by=session.get('user_id'),
                    statut='brouillon'
                )
                db.session.add(avis)
                db.session.flush()
                
                # Ajouter les articles
                for item_id in item_ids:
                    quantite = request.form.get(f'quantite_{item_id}', 0)
                    caracteristiques = request.form.get(f'caracteristiques_{item_id}', '')
                    garanties = request.form.get(f'garanties_{item_id}', '')
                    
                    if int(quantite) > 0:
                        avis_item = AvisAchatItem(
                            avis_achat_id=avis.id,
                            item_id=int(item_id),
                            quantite=int(quantite),
                            caracteristiques=caracteristiques,
                            garanties=garanties
                        )
                        db.session.add(avis_item)
                
                # Log d'audit
                create_audit_log(
                    action='create',
                    entity_type='avis_achat',
                    entity_id=avis.id,
                    entity_name=numero_avis,
                    description=f"Nature: {nature} | Articles: {len(item_ids)}"
                )
                
                avis_crees.append({
                    'numero': numero_avis,
                    'nature': nature,
                    'articles': len(item_ids)
                })
            
            db.session.commit()
            
            # Message de succès avec récapitulatif
            if len(avis_crees) == 1:
                avis = avis_crees[0]
                flash(f'✓ Avis d\'achat créé avec succès | {avis["numero"]} | Nature: {avis["nature"]} | {avis["articles"]} article(s)', 'success')
            else:
                recap = ' | '.join([f"{a['numero']} ({a['nature']}: {a['articles']} art.)" for a in avis_crees])
                flash(f'✓ {len(avis_crees)} avis d\'achat créés avec succès | {recap}', 'success')
            
            return redirect(url_for('avis_achats'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de la création des avis d\'achat: {str(e)}', 'error')
            return redirect(url_for('add_avis_achat'))
    
    # GET - Afficher le formulaire
    return render_template('add_avis_achat.html')

@app.route('/api/items-by-category/<category>')
@login_required
def get_items_by_category(category):
    """Récupère les articles d'une catégorie spécifique"""
    try:
        items = Item.query.filter_by(category=category).order_by(Item.name).all()
        return jsonify({
            'success': True,
            'items': [{
                'id': item.id,
                'name': item.name,
                'sku': item.sku,
                'unit': item.unit or 'Unité',
                'quantity': item.quantity,
                'reorder_level': item.reorder_level,
                'category': item.category
            } for item in items]
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erreur lors de la récupération des articles: {str(e)}'
        }), 500

@app.route('/avis-achats/<int:avis_id>/pdf')
@login_required
def generate_avis_achat_pdf(avis_id):
    """Générer le PDF de l'avis d'achat"""
    avis = AvisAchat.query.get_or_404(avis_id)
    
    try:
        from reportlab.platypus import PageTemplate, Frame
        from reportlab.platypus.doctemplate import BaseDocTemplate
        
        # Créer le PDF avec marges optimisées (augmenter bottomMargin pour le footer)
        buffer = BytesIO()
        
        # Fonction pour ajouter le footer sur chaque page
        def add_footer(canvas, doc):
            canvas.saveState()
            # Position du footer (bas de page)
            footer_y = 50
            
            # Style pour le footer
            canvas.setFont('Helvetica', 7)
            
            # Ligne de séparation plus courte
            canvas.line(30, footer_y + 25, 250, footer_y + 25)
            
            # Note (2) avec exposant - interligne réduit
            canvas.drawString(30, footer_y + 12, "(2) À compléter pour chaque prestation, lorsque la garantie est exigée.")
            
            # Note (3) avec exposant - interligne réduit
            canvas.drawString(30, footer_y, "(3) Sous réserve de l'application des dispositions de l'arrêté pris pour l'application de l'article 135 du décret n° 2-22-431 du 8")
            canvas.drawString(30, footer_y - 8, "mars 2023 relatif aux marchés publics.")
            
            canvas.restoreState()
        
        # Utiliser BaseDocTemplate pour personnaliser le footer
        doc = BaseDocTemplate(buffer, pagesize=A4, 
                            rightMargin=42.52, leftMargin=10, 
                            topMargin=10, bottomMargin=80)
        
        # Définir le frame et le template de page
        frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='normal')
        template = PageTemplate(id='main', frames=frame, onPage=add_footer)
        doc.addPageTemplates([template])
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Styles personnalisés alignés avec le décharge
        styles.add(ParagraphStyle(
            name='CustomTitle',
            alignment=TA_CENTER,
            fontSize=12,
            spaceAfter=15,
            leading=14,
            fontName='Helvetica-Bold'
        ))
        
        styles.add(ParagraphStyle(
            name='CustomNormal',
            alignment=TA_LEFT,
            fontSize=10,
            spaceAfter=6,
            leading=12,
            fontName='Helvetica'
        ))
        
        styles.add(ParagraphStyle(
            name='WrappedCell',
            alignment=TA_LEFT,
            fontSize=9,
            leading=11,
            wordWrap='CJK'
        ))
        
        styles.add(ParagraphStyle(
            name='WrappedCellHeader',
            alignment=TA_CENTER,
            fontSize=8,
            leading=10,
            wordWrap='CJK',
            fontName='Helvetica-Bold'
        ))
        
        styles.add(ParagraphStyle(
            name='WrappedCellCenter',
            alignment=TA_CENTER,
            fontSize=9,
            leading=11,
            wordWrap='CJK'
        ))
        
        # Style pour caractéristiques centré
        styles.add(ParagraphStyle(
            name='WrappedCellCenterBoth',
            alignment=TA_CENTER,
            fontSize=9,
            leading=11,
            wordWrap='CJK'
        ))
        
        # Style pour texte justifié avec indentation
        styles.add(ParagraphStyle(
            name='CustomJustified',
            alignment=TA_JUSTIFY,
            fontSize=10,
            spaceAfter=6,
            leading=12,
            fontName='Helvetica',
            leftIndent=1*cm  # Débute à partir de la colonne Désignations
        ))
        
        # Styles pour sections indentées
        styles.add(ParagraphStyle(
            name='CustomIndented',
            alignment=TA_LEFT,
            fontSize=10,
            spaceAfter=6,
            leading=12,
            fontName='Helvetica',
            leftIndent=1*cm
        ))

        # --- CONFIGURATION DES POLICES ET LANGUES ---
        try:
            pdfmetrics.registerFont(TTFont('Calibri', 'C:\\Windows\\Fonts\\calibri.ttf'))
            pdfmetrics.registerFont(TTFont('Calibri-Bold', 'C:\\Windows\\Fonts\\calibrib.ttf'))
            pdfmetrics.registerFont(TTFont('Ebrima', 'C:\\Windows\\Fonts\\ebrima.ttf')) # Pour Tifinagh
        except Exception as e:
            print(f"Warning: Could not load system fonts: {e}")

        def process_arabic(text):
            if not text: return ""
            reshaped_text = arabic_reshaper.reshape(text)
            bidi_text = get_display(reshaped_text)
            return bidi_text

        # Dictionnaire de traduction (à enrichir ou mettre en DB)
        AR_TRANSLATIONS = {
            'RABAT-SALÉ-KENITRA': 'الرباط سلا القنيطرة',
            'SIDI KACEM': 'سيدي قاسم',
            'SIDI-KACEM': 'سيدي قاسم',
            'Sidi Kacem': 'سيدي قاسم',
            'Unité de SIDI KACEM': 'سيدي قاسم'
        }

        # Styles pour l'en-tête
        styles.add(ParagraphStyle(name='HeaderFR', fontName='Helvetica-Bold', fontSize=8, alignment=TA_LEFT, leading=10))
        styles.add(ParagraphStyle(name='HeaderAR', fontName='Calibri-Bold', fontSize=9, alignment=TA_RIGHT, leading=11))
        styles.add(ParagraphStyle(name='HeaderTifinagh', fontName='Ebrima', fontSize=8, alignment=TA_RIGHT, leading=10))

        # --- CONSTRUCTION DE L'EN-TÊTE DYNAMIQUE ---
        
        # Données du domaine de compétence
        current_user = User.query.get(session.get('user_id'))
        if not current_user:
             return redirect(url_for('login'))
        
        nom_region_fr = "RABAT-SALÉ-KENITRA" # Par défaut
        nom_unite_fr = "SIDI KACEM" # Par défaut
        is_provincial = True 

        if current_user.unite_competence:
            # Niveau Provincial
            u_nom = current_user.unite_competence.nom
            # Nettoyage "Unité de " ou "Unité Provinciale de "
            u_clean = u_nom.replace('Unité de ', '').replace('Unité Provinciale de ', '').strip()
            nom_unite_fr = u_clean
            
            if current_user.unite_competence.region:
                 nom_region_fr = current_user.unite_competence.region.nom
            is_provincial = True
        elif current_user.region_competence:
            # Niveau Régional
            nom_region_fr = current_user.region_competence.nom
            nom_unite_fr = "" # Pas d'unité spécifique
            is_provincial = False
        
        # Traductions Arabe
        nom_region_ar = AR_TRANSLATIONS.get(nom_region_fr, nom_region_fr)
        nom_unite_ar = AR_TRANSLATIONS.get(nom_unite_fr, nom_unite_fr)

        # Contenu Colonne Gauche (Français)
        header_fr_text = [
            "ROYAUME DU MAROC",
            "MINISTÈRE DE L'INTÉRIEUR",
            "DIRECTION GÉNÉRALE DE LA",
            "PROTECTION CIVILE",
            "COMMANDEMENT RÉGIONAL",
            nom_region_fr
        ]
        if is_provincial:
            header_fr_text.append("COMMANDEMENT PROVINCIAL")
            header_fr_text.append(nom_unite_fr)
        
        col_left = []
        for line in header_fr_text:
            col_left.append(Paragraph(line, styles['HeaderFR']))

        # Contenu Colonne Droite (Arabe/Tifinagh)
        # Ordre visuel : Haut vers Bas
        col_right = []
        col_right.append(Paragraph(process_arabic("المملكة المغربية"), styles['HeaderAR']))
        col_right.append(Paragraph("ⵜⴰⴳⵍⴷⵉⵜ ⵏ ⵍⵎⵖⵔⵉⴱ", styles['HeaderTifinagh']))
        
        col_right.append(Paragraph(process_arabic("وزارة الداخلية"), styles['HeaderAR']))
        col_right.append(Paragraph("ⵜⴰⵎⴰⵡⴰⵙⵜ ⵏ ⵓⴳⵏⵙ", styles['HeaderTifinagh']))
        
        col_right.append(Paragraph(process_arabic("المديرية العامة للوقاية المدنية"), styles['HeaderAR']))
        col_right.append(Paragraph("ⵜⴰⵎⵀⵍⴰ ⵜⴰⵎⴰⵜⴰⵢⵜ ⵏ ⵡⴰⵔⴰⵢ ⵓⵖⵔⵉⵎ", styles['HeaderTifinagh']))
        
        col_right.append(Paragraph(process_arabic("القيادة الجهوية للوقاية المدنية"), styles['HeaderAR']))
        col_right.append(Paragraph(process_arabic(nom_region_ar), styles['HeaderAR']))
        
        if is_provincial:
             col_right.append(Paragraph(process_arabic("القيادة الإقليمية للوقاية المدنية"), styles['HeaderAR']))
             col_right.append(Paragraph(process_arabic(nom_unite_ar), styles['HeaderAR']))

        # Logo au centre
        img_logo_path = os.path.join(app.root_path, 'static', 'images', 'logo.png')
        if os.path.exists(img_logo_path):
            img_logo = Image(img_logo_path, width=2.5*cm, height=2.5*cm) # Logo un peu plus grand
        else:
            img_logo = Paragraph("LOGO", styles['HeaderFR'])

        # Table En-tête (A4 Width = 595.27, Left Margin=10, Right Margin=42.52)
        total_width = A4[0] - 10 - 42.52
        col_w = total_width / 3
        
        header_table_data = [[col_left, img_logo, col_right]]
        header_table = Table(header_table_data, colWidths=[col_w*1.25, col_w*0.75, col_w*1.0])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),  # FR Left align (standard western)
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'), # AR Right align for clean margin
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (0, 0), 1*cm), # Alignment with "Le Gouverneur" (1cm indent)
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(header_table)
        
        elements.append(Spacer(1, 15))
        
        # Titre de l'avis (numéro formaté sur 2 chiffres)
        numero_parts = (avis.numero_avis or '').split('/')
        display_num = avis.numero_avis
        try:
            n_int = int(numero_parts[0])
            display_num = f"{n_int:02d}/" + "/".join(numero_parts[1:])
        except Exception:
            pass
        title = f"Avis d'achat sur bon de commande n° {display_num}"
        elements.append(Paragraph(title, styles['CustomTitle']))
        elements.append(Spacer(1, 10))
        
        # Texte d'introduction DYNAMIQUE
        intro_text = f"Le Gouverneur de la Province de {nom_unite_fr} procède à l'achat sur bon de commande des prestations objet du présent avis, en application des dispositions de l'article 91 du décret n° 2-22-431 relatif aux marchés publics."
        elements.append(Paragraph(intro_text, styles['CustomJustified']))
        elements.append(Spacer(1, 10))
        
        # Objet de la prestation DYNAMIQUE
        # "Objet de la prestation : Achat de Fournitures... pour les services du [Nom Unité]"
        # "Commandement Provincial de la Protection Civile de [Nom Unité]"
        
        # Construction de l'objet complet
        objet_complet = avis.objet_prestation
        # Si c'est l'objet auto-généré par défaut, on le reformate dynamiquement si besoin
        # Ou on force le format demandé :
        # "Achat de [NATURE] pour les services du [UNIT] Commandement Provincial..."
        
        # On peut reconstruire l'objet basé sur la nature pour être sûr
        if avis.nature_prestation:
            objet_construit = f"Achat de {avis.nature_prestation} pour les services du {nom_unite_fr}<br/>Commandement Provincial de la Protection Civile de {nom_unite_fr}"
        else:
            objet_construit = avis.objet_prestation

        # Créer un tableau avec 2 colonnes : label + contenu
        objet_data = [[
            Paragraph("<b><u>Objet de la prestation :</u></b>", styles['CustomIndented']),
            Paragraph(objet_construit, styles['CustomNormal'])
        ]]
        objet_table = Table(objet_data, colWidths=[5*cm, 12*cm], hAlign='LEFT')
        objet_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(objet_table)
        elements.append(Spacer(1, 6))
        
        # Titre du tableau
        elements.append(Paragraph("<b>Consistance de la prestation :</b>", styles['CustomIndented']))
        elements.append(Spacer(1, 5))
        
        # Créer le tableau des articles avec Paragraphs pour le wrapping
        table_data_wrapped = []
        
        # En-tête
        header_row = [
            Paragraph('N°', styles['WrappedCellHeader']),
            Paragraph('Désignations', styles['WrappedCellHeader']),
            Paragraph('Caractéristiques<br/>et spécifications', styles['WrappedCellHeader']),
            Paragraph('Unité de<br/>mesure', styles['WrappedCellHeader']),
            Paragraph('Quantité', styles['WrappedCellHeader']),
            Paragraph('Garanties<br/>exigées(2)', styles['WrappedCellHeader'])
        ]
        table_data_wrapped.append(header_row)
        
        # Données
        for idx, avis_item in enumerate(avis.items, 1):
            item = avis_item.item
            row = [
                Paragraph(str(idx), styles['WrappedCellCenter']),
                Paragraph(item.name, styles['WrappedCell']),  # Aligné à gauche
                Paragraph(avis_item.caracteristiques or 'Voir désignation', styles['WrappedCellCenterBoth']),  # Centré
                Paragraph(item.unit, styles['WrappedCellCenter']),
                Paragraph(str(avis_item.quantite), styles['WrappedCellCenter']),
                Paragraph(avis_item.garanties or '', styles['WrappedCellCenter'])
            ]
            table_data_wrapped.append(row)
        
        # Largeurs de colonnes optimisées (Désignations récupère l'espace)
        # N°: 1cm, Désignations: calculé, Caract: 3cm, Unité: 1.7cm, Quantité: 1.7cm, Garanties: 1.9cm
        col_widths = [1*cm, 7.5*cm, 3*cm, 1.7*cm, 1.7*cm, 1.9*cm]
        
        table = Table(table_data_wrapped, repeatRows=1, hAlign='CENTER', colWidths=col_widths)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.7, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),  # Centrage vertical pour toutes les cellules
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            # Centrage pour les colonnes N°, Unité, Quantité, Garanties
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('VALIGN', (0, 1), (0, -1), 'MIDDLE'),
            # Désignations aligné à gauche
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('VALIGN', (1, 1), (1, -1), 'MIDDLE'),
            # Caractéristiques centré horizontal et vertical
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),
            ('VALIGN', (2, 1), (2, -1), 'MIDDLE'),
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),
            ('VALIGN', (3, 1), (3, -1), 'MIDDLE'),
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),
            ('VALIGN', (4, 1), (4, -1), 'MIDDLE'),
            ('ALIGN', (5, 1), (5, -1), 'CENTER'),
            ('VALIGN', (5, 1), (5, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 8))
        
        # Lieu d'exécution DYNAMIQUE
        # "Caserne de la Protection Civile de [Nom Unité]"
        if is_provincial:
            lieu_text = f"Lieu d'exécution : Caserne de la Protection Civile de {nom_unite_fr}"
        else:
            lieu_text = f"Lieu d'exécution : {avis.lieu_execution}" # Fallback si pas provincial
            
        elements.append(Paragraph(lieu_text, styles['CustomIndented']))
        elements.append(Spacer(1, 6))
        
        # Délai d'exécution (indenté à partir de la colonne Désignations)
        if avis.delai_execution:
            delai_text = f"Délai d'exécution ou date de livraison de la prestation : {avis.delai_execution} jours."
            elements.append(Paragraph(delai_text, styles['CustomIndented']))
            elements.append(Spacer(1, 6))
        
        # Date limite de réception (indenté à partir de la colonne Désignations)
        if avis.date_limite_reception:
            # Format avec espaces pour renseigner la date et l'heure manuellement
            # Date: ~4.5cm (30 points), Heure: ~0.5cm (4 points)
            date_text = f"Date et heure limites de réception des devis des concurrents : ............................. à .... H."
            elements.append(Paragraph(date_text, styles['CustomIndented']))
            elements.append(Spacer(1, 6))
        
        # Texte de clôture (indenté à partir de la colonne Désignations avec exposant (3))
        cloture_text = """Les plis des concurrents sont déposés par voie électronique dans le portail des marchés publics accessible à l'adresse www.marchespublics.gov.ma.(3)"""
        elements.append(Paragraph(cloture_text, styles['CustomIndented']))
        
        # Construire le PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Mettre à jour le statut si c'était un brouillon
        if avis.statut == 'brouillon':
            avis.statut = 'publie'
            db.session.commit()
        
        # Nom du fichier
        filename = f"avis_achat_{avis.numero_avis.replace('/', '_')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        flash(f'Erreur lors de la génération du PDF: {str(e)}', 'error')
        return redirect(url_for('avis_achats'))

@app.route('/avis-achats/<int:avis_id>/publish', methods=['POST'])
@login_required
@modification_required
def publish_avis_achat(avis_id):
    avis = AvisAchat.query.get_or_404(avis_id)
    try:
        if avis.statut != 'publie':
            old_status = avis.statut
            avis.statut = 'publie'
            db.session.commit()
            create_audit_log(
                action='update',
                entity_type='avis_achat',
                entity_id=avis.id,
                entity_name=avis.numero_avis,
                description=f"Statut: {old_status} → publie"
            )
            flash("✓ Avis d'achat publié.", 'success')
        else:
            flash("Avis d'achat déjà publié.", 'info')
    except Exception as e:
        db.session.rollback()
        flash(f"Erreur lors de la publication: {str(e)}", 'error')
    return redirect(url_for('avis_achats'))

@app.route('/avis-achats/<int:avis_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_avis_achat(avis_id):
    """Modifier un avis d'achat existant"""
    avis = AvisAchat.query.get_or_404(avis_id)
    
    if request.method == 'POST':
        try:
            # Récupérer les données modifiables
            delai_execution = request.form.get('delai_execution', '').strip()
            date_limite_reception = request.form.get('date_limite_reception', '').strip()
            lieu_execution = request.form.get('lieu_execution', '').strip()
            
            # Validation
            if not delai_execution or not date_limite_reception or not lieu_execution:
                flash('Tous les champs obligatoires doivent être renseignés.', 'error')
                return redirect(url_for('edit_avis_achat', avis_id=avis_id))
            
            # Mettre à jour les informations de l'avis
            avis.delai_execution = delai_execution
            avis.date_limite_reception = date_limite_reception
            avis.lieu_execution = lieu_execution
            
            # Passer le statut en BROUILLON lors de la modification
            avis.statut = 'brouillon'
            
            # Auto-générer l'objet de la prestation
            if avis.nature_prestation in ['Fournitures de bureaux et documentation', 
                                          'Produits d\'hygiène et de désinfection', 
                                          'Fournitures pour le matériel informatique']:
                avis.objet_prestation = f"Achat de {avis.nature_prestation} pour les services du Commandement Provincial de la Protection Civile de Sidi Kacem."
            elif avis.nature_prestation in ['Articles de plomberies', 'Articles électriques', 'Articles de la peinture']:
                # Pour la quatrième prestation (sous-prestations d'entretien)
                avis.objet_prestation = f"Achat d'{avis.nature_prestation} pour l'entretien et réparation des bâtiments administratifs des services du Commandement Provincial de la Protection Civile de Sidi Kacem."
            else:
                # Pour les autres prestations
                avis.objet_prestation = f"Achat de {avis.nature_prestation} pour les services du Commandement Provincial de la Protection Civile de Sidi Kacem."
            
            # Mettre à jour les articles (quantités, caractéristiques, garanties)
            item_ids = request.form.getlist('item_ids[]')
            quantites = request.form.getlist('quantites[]')
            caracteristiques_list = request.form.getlist('caracteristiques[]')
            garanties_list = request.form.getlist('garanties[]')
            
            for idx, item_id in enumerate(item_ids):
                avis_item = AvisAchatItem.query.get(int(item_id))
                if avis_item and avis_item.avis_achat_id == avis.id:
                    avis_item.quantite = int(quantites[idx])
                    avis_item.caracteristiques = caracteristiques_list[idx] if idx < len(caracteristiques_list) else ''
                    avis_item.garanties = garanties_list[idx] if idx < len(garanties_list) else ''
            
            db.session.commit()
            
            # Log d'audit
            create_audit_log(
                action='update',
                entity_type='avis_achat',
                entity_id=avis.id,
                entity_name=avis.numero_avis,
                description=f"Modification: Délai={delai_execution}j, Date limite={date_limite_reception}"
            )
            
            flash(f'✓ Avis d\'achat modifié avec succès | {avis.numero_avis} | Statut: BROUILLON | Articles: {len(item_ids)}', 'success')
            return redirect(url_for('avis_achats'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de la modification: {str(e)}', 'error')
            return redirect(url_for('edit_avis_achat', avis_id=avis_id))
    
    # GET - Afficher le formulaire d'édition
    return render_template('edit_avis_achat.html', avis=avis)

@app.route('/avis-achats/<int:avis_id>/delete', methods=['POST'])
@login_required
def delete_avis_achat(avis_id):
    """Supprimer un avis d'achat"""
    avis = AvisAchat.query.get_or_404(avis_id)
    numero = avis.numero_avis
    nature = avis.nature_prestation
    
    try:
        # Soft Delete: Marquer l'avis comme supprimé au lieu de le supprimer réellement
        avis.is_deleted = True
        avis.deleted_at = datetime.utcnow()
        avis.deleted_by = session.get('user_id')
        
        db.session.commit()
        
        # Log d'audit
        create_audit_log(
            action='delete',
            entity_type='avis_achat',
            entity_id=avis_id,
            entity_name=numero,
            description=f"Nature: {nature}"
        )
        
        flash(f'✓ Avis d\'achat supprimé avec succès | {numero} | Nature: {nature}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression: {str(e)}', 'error')
    
    return redirect(url_for('avis_achats'))

def create_admin_user():
    """Create default admin user if it doesn't exist"""
    admin = User.query.filter_by(email='anas_prev@protection.com').first()
    if not admin:
        admin = User(
            email='anas_prev@protection.com',
            password_hash=generate_password_hash('$@SK$@2025PC'),
            name='Anas',
            role='admin'
        )
        db.session.add(admin)
        db.session.commit()
        print("Anas (admin) created successfully!")

def create_default_unites_services():
    """Créer les unités et services par défaut"""
    # Vérifier si les données existent déjà
    if Unite.query.count() > 0:
        return
    
    # Créer l'unité de SIDI KACEM avec ses services
    unite_sidi_kacem = Unite(
        nom='Unité de SIDI KACEM',
        description='Unité principale de Sidi Kacem'
    )
    db.session.add(unite_sidi_kacem)
    db.session.flush()  # Pour obtenir l'ID
    
    # Services de l'unité SIDI KACEM
    services_sidi_kacem = [
        Service(nom='Sce Secrétariat', unite_id=unite_sidi_kacem.id),
        Service(nom='Sce Prévention', unite_id=unite_sidi_kacem.id),
        Service(nom='Sce Technique', unite_id=unite_sidi_kacem.id),
        Service(nom='Standardiste', unite_id=unite_sidi_kacem.id),
        Service(nom='Bureau du Cdt Provincial', unite_id=unite_sidi_kacem.id)
    ]
    
    for service in services_sidi_kacem:
        db.session.add(service)
    
    # Centres de secours (sans services spécifiques)
    centres = [
        Unite(nom='CENTRE DE SECOURS JORF EL MELHA', description='Centre de secours Jorf El Melha'),
        Unite(nom='CENTRE DE SECOURS M BEL KSIRI', description='Centre de secours M Bel Ksiri'),
        Unite(nom='CENTRE DE SECOURS HAD KOURT', description='Centre de secours Had Kourt')
    ]
    
    for centre in centres:
        db.session.add(centre)
    
    db.session.commit()
    print("Unités et services créés avec succès!")


# ==================== IMPORTATION EXCEL ====================

@app.route('/import-items', methods=['GET', 'POST'])
@login_required
def import_items():
    """Page d'importation des articles depuis Excel"""
    import_results = None
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Aucun fichier sélectionné', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        
        if file.filename == '':
            flash('Aucun fichier sélectionné', 'error')
            return redirect(request.url)
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Format de fichier invalide. Veuillez utiliser un fichier Excel (.xlsx ou .xls)', 'error')
            return redirect(request.url)
        
        skip_duplicates = request.form.get('skip_duplicates') == 'on'
        
        try:
            # Charger le fichier Excel
            wb = load_workbook(file)
            ws = wb.active
            
            success_count = 0
            skipped_count = 0
            error_count = 0
            error_details = []
            
            # Catégories principales et sous-catégories valides
            valid_categories = [
                'Fournitures de bureaux et documentation',
                'Produits d\'hygiène et de désinfection',
                'Fournitures pour le matériel informatique',
                'Alimentation à usage Humaine',  # Catégorie principale
                'Entretien bâtiments administratifs',  # Catégorie principale
                'Articles de plomberies',  # Sous-catégorie d'Entretien
                'Articles électriques',  # Sous-catégorie d'Entretien
                'Articles de la peinture'  # Sous-catégorie d'Entretien
            ]
            valid_units = ['Unité', 'Lot', 'Paquet', 'Boîte', 'Ramette', 'Kg', 'Litre']
            
            # Mapping pour normaliser les unités (insensible à la casse)
            unit_mapping = {
                'unite': 'Unité',
                'unité': 'Unité',
                'lot': 'Lot',
                'paquet': 'Paquet',
                'boite': 'Boîte',
                'boîte': 'Boîte',
                'ramette': 'Ramette',
                'kg': 'Kg',
                'kilo': 'Kg',
                'kilogramme': 'Kg',
                'litre': 'Litre',
                'l': 'Litre'
            }
            
            # Parcourir les lignes (ignorer l'en-tête)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extraire les données
                    sku = row[0] if row[0] else None
                    name = row[1]
                    category = row[2]
                    description = row[3] if len(row) > 3 else ''
                    unit_raw = row[4] if len(row) > 4 else 'Unité'
                    quantity = int(row[5]) if len(row) > 5 and row[5] else 0
                    reorder_level = int(row[6]) if len(row) > 6 and row[6] else 0
                    
                    # Normaliser l'unité (insensible à la casse)
                    if unit_raw:
                        unit_normalized = unit_mapping.get(str(unit_raw).strip().lower())
                        if unit_normalized:
                            unit = unit_normalized
                        else:
                            unit = unit_raw
                    else:
                        unit = 'Unité'
                    
                    # Validation
                    if not name:
                        error_details.append({'row': row_idx, 'message': 'Nom manquant'})
                        error_count += 1
                        continue
                    
                    if not category or category not in valid_categories:
                        error_details.append({'row': row_idx, 'message': f'Catégorie invalide: {category}'})
                        error_count += 1
                        continue
                    
                    if unit not in valid_units:
                        error_details.append({'row': row_idx, 'message': f'Unité invalide: {unit_raw} (attendu: Unité, Lot, Paquet, Boîte, Ramette, Kg ou Litre)'})
                        error_count += 1
                        continue
                    
                    # Vérifier si le SKU existe déjà
                    if sku:
                        existing_item = Item.query.filter_by(sku=sku).first()
                        if existing_item:
                            if skip_duplicates:
                                skipped_count += 1
                                continue
                            else:
                                error_details.append({'row': row_idx, 'message': f'SKU déjà existant: {sku}'})
                                error_count += 1
                                continue
                    
                    # Générer un SKU si nécessaire (utilise le même système que l'ajout manuel)
                    if not sku:
                        sku = Item.generate_next_sku()
                    
                    # Créer l'article
                    new_item = Item(
                        sku=sku,
                        name=name,
                        category=category,
                        description=description,
                        unit=unit,
                        quantity=quantity,
                        reorder_level=reorder_level
                    )
                    
                    db.session.add(new_item)
                    success_count += 1
                    
                except Exception as e:
                    error_details.append({'row': row_idx, 'message': str(e)})
                    error_count += 1
                    continue
            
            db.session.commit()
            
            # Créer un log d'audit
            create_audit_log(
                action='import',
                entity_type='item',
                entity_id=0,
                entity_name='Import Excel',
                description=f'{success_count} article(s) importé(s) | {skipped_count} ignoré(s) | {error_count} erreur(s)'
            )
            
            import_results = {
                'success': success_count,
                'skipped': skipped_count,
                'errors': error_count,
                'error_details': error_details[:10]  # Limiter à 10 erreurs affichées
            }
            
            if success_count > 0:
                flash(f'✓ {success_count} article(s) importé(s) avec succès', 'success')
            if skipped_count > 0:
                flash(f'⚠ {skipped_count} article(s) ignoré(s) (doublons)', 'warning')
            if error_count > 0:
                flash(f'✗ {error_count} erreur(s) détectée(s)', 'error')
                
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de l\'importation: {str(e)}', 'error')
    
    return render_template('import_items.html', import_results=import_results)


@app.route('/download-template')
@login_required
def download_template():
    """Télécharger le modèle Excel pour l'importation"""
    # Créer un nouveau classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "Articles"
    
    # En-têtes
    headers = ['SKU', 'Nom', 'Catégorie', 'Description', 'Unité', 'Quantité', 'Seuil de Commande']
    ws.append(headers)
    
    # Styliser les en-têtes
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajuster la largeur des colonnes
    column_widths = [15, 25, 40, 35, 12, 12, 18]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Ajouter une feuille avec les instructions
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ['INSTRUCTIONS D\'IMPORTATION'],
        [''],
        ['Colonnes obligatoires:'],
        ['- Nom: Nom de l\'article (obligatoire)'],
        ['- Catégorie: Catégorie de l\'article (obligatoire)'],
        ['- Unité: Unité de mesure (obligatoire)'],
        ['- Quantité: Quantité initiale en stock (obligatoire)'],
        ['- Seuil de Commande: Seuil de réapprovisionnement (obligatoire)'],
        [''],
        ['Colonnes optionnelles:'],
        ['- SKU: Code unique (sera généré automatiquement si vide)'],
        ['- Description: Description détaillée de l\'article'],
        [''],
        ['Catégories valides:'],
        ['CATÉGORIES PRINCIPALES:'],
        ['- Fournitures de bureaux et documentation'],
        ['- Produits d\'hygiène et de désinfection'],
        ['- Fournitures pour le matériel informatique'],
        ['- Alimentation à usage Humaine'],
        [''],
        ['CATÉGORIE AVEC SOUS-PRESTATIONS:'],
        ['- Entretien bâtiments administratifs (ou ses sous-catégories):'],
        ['  → Articles de plomberies'],
        ['  → Articles électriques'],
        ['  → Articles de la peinture'],
        [''],
        ['NOTE: Pour Entretien, vous pouvez utiliser soit la catégorie'],
        ['principale, soit directement une sous-catégorie.'],
        [''],
        ['Unités valides (insensibles à la casse):'],
        ['- Unité (ou UNITE, unite, Unité)'],
        ['- Lot (ou LOT, lot)'],
        ['- Paquet (ou PAQUET, paquet)'],
        ['- Boîte (ou BOITE, boite, Boîte)'],
        ['- Ramette (ou RAMETTE, ramette)'],
        ['- Kg (ou KG, kg, kilo, kilogramme)'],
        ['- Litre (ou LITRE, litre, L, l)'],
        [''],
        ['NOTE: Les unités sont automatiquement normalisées.'],
        ['Vous pouvez écrire en majuscules, minuscules ou avec/sans accent.'],
    ]
    
    for row in instructions:
        ws_instructions.append(row)
    
    # Styliser le titre
    ws_instructions['A1'].font = Font(bold=True, size=14, color='4472C4')
    ws_instructions.column_dimensions['A'].width = 60
    
    # Sauvegarder dans un buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='modele_importation_articles.xlsx'
    )

# Fonction pour initialiser les budgets par nature de prestation
def init_budgets(annee=None):
    """Initialises les budgets pour chaque nature de prestation.

    - Si une année est fournie, initialise pour cette année.
    - Sinon, utilise l'année courante.
    - Si l'année précédente possède déjà une configuration, on duplique
      ces lignes et montants (pour garder les budgets actuels).
    - Sinon, on utilise une configuration par défaut codée en dur.
    """
    from datetime import datetime
    annee_courante = annee if annee is not None else datetime.now().year

    # Ne rien faire si des budgets existent déjà pour cette année
    if BudgetNature.query.filter_by(annee=annee_courante).first():
        return

    # 1) Essayer de copier l'année précédente si elle existe
    annee_precedente = annee_courante - 1
    prev_budgets = BudgetNature.query.filter_by(annee=annee_precedente).all()
    
    # Liste des exclusions pour les centres
    exclusions_centres = ["Alimentation à usage Humaine", "Alimentation Humaine"]
    
    if prev_budgets:
        for b in prev_budgets:
            unite = Unite.query.get(b.unite_id) if b.unite_id else None
            
            # Respecter l'exclusion même lors de la copie si besoin
            if unite and unite.type == 'centre' and b.nature in exclusions_centres:
                continue
                
            clone = BudgetNature(
                nature=b.nature,
                montant_ttc=b.montant_ttc,
                annee=annee_courante,
                unite_id=b.unite_id
            )
            db.session.add(clone)
        db.session.commit()
        print(f"✅ Budgets {annee_courante} initialisés par copie de {annee_precedente}")
        return

    # 2) Sinon, utiliser la configuration par défaut pour TOUTES les unités et centres
    budgets_config = [
        {'nature': 'Fournitures de bureaux et documentation', 'budget_centre': 1500.00, 'budget_unite': 7000.00},
        {'nature': 'Fournitures pour le matériel informatique', 'budget_centre': 1500.00, 'budget_unite': 3000.00},
        {'nature': "Produits d'hygiène et de désinfection", 'budget_centre': 2000.00, 'budget_unite': 5000.00},
        {'nature': 'Entretien bâtiments administratifs', 'budget_centre': 1500.00, 'budget_unite': 10000.00},
        {'nature': 'Alimentation à usage Humaine', 'budget_centre': 0.00, 'budget_unite': 2000.00},
    ]

    unites = Unite.query.all()
    for config in budgets_config:
        for u in unites:
            # Exclure l'alimentation humaine pour les centres
            if u.type == 'centre' and config['nature'] in exclusions_centres:
                continue
                
            existing = BudgetNature.query.filter_by(nature=config['nature'], annee=annee_courante, unite_id=u.id).first()
            if not existing:
                montant = config['budget_unite'] if u.type == 'unite' else config['budget_centre']
                budget = BudgetNature(
                    nature=config['nature'],
                    montant_ttc=montant,
                    annee=annee_courante,
                    unite_id=u.id
                )
                db.session.add(budget)

    db.session.commit()
    print(f"✅ Budgets initialisés individuellement pour {annee_courante}")

def ensure_budget_individual_migration():
    """Ajoute la colonne 'montant_ttc' et migre les données existantes"""
    try:
        from sqlalchemy import text
        res = db.session.execute(text("PRAGMA table_info('budget_natures')"))
        existing_cols = {row[1] for row in res}
        
        if 'montant_ttc' not in existing_cols:
            db.session.execute(text("ALTER TABLE budget_natures ADD COLUMN montant_ttc FLOAT DEFAULT 0.0"))
            db.session.commit()
            print("✅ Colonne 'montant_ttc' ajoutée à 'budget_natures'.")
        
        # Migration effective des données legacy
        # On ne traite que les records où montant_ttc est encore 0
        legacy_budgets = BudgetNature.query.filter_by(montant_ttc=0.0).all()
        for b in legacy_budgets:
            unite = Unite.query.get(b.unite_id) if b.unite_id else None
            # Si rattaché à une unité provinciale, on affecte le budget unité
            if unite and unite.type == 'unite':
                b.montant_ttc = b.budget_unite_ttc
                
                # Et on crée les budgets pour les centres rattachés s'ils n'existent pas
                for center in unite.children:
                    exists = BudgetNature.query.filter_by(nature=b.nature, annee=b.annee, unite_id=center.id).first()
                    if not exists:
                        new_b = BudgetNature(
                            nature=b.nature,
                            annee=b.annee,
                            unite_id=center.id,
                            montant_ttc=b.budget_centre_ttc
                        )
                        db.session.add(new_b)
            elif unite and unite.type == 'centre':
                b.montant_ttc = b.budget_centre_ttc
            
        db.session.commit()
    except Exception as e:
        print(f"⚠️ Erreur migration budgets individuels: {e}")
        db.session.rollback()

# Routes pour les engagements
@app.route('/engagements')
@login_required
def engagements():
    """Page principale des engagements avec statistiques détaillées et filtrage par année."""
    from sqlalchemy import func, case
    from datetime import datetime
    
    type_depense = request.args.get('type', 'bon_commande')
    nature_filter = request.args.get('nature', '')
    year_filter = request.args.get('year', type=int)
    
    # Récupérer les années disponibles pour le filtre
    # Pour les bons de commande
    bc_years = db.session.query(
        func.extract('year', BonCommande.date_engagement)
    ).distinct().all()
    
    # Pour les indemnités
    indemnite_years = db.session.query(
        func.extract('year', IndemnitesDeplacement.created_at)
    ).distinct().all()
    
    # Combiner et trier les années uniques
    all_years = sorted(list(set([int(y[0]) for y in (bc_years + indemnite_years) if y[0] is not None])), reverse=True)
    
    # Si aucune année n'est spécifiée, utiliser l'année courante
    annee_courante = year_filter if year_filter else datetime.now().year
    
    # Calculer les années précédente et suivante pour la navigation
    if all_years:
        min_year = min(all_years)
        max_year = max(all_years)
        prev_year = annee_courante - 1 if (annee_courante - 1) >= min_year else None
        next_year = annee_courante + 1 if (annee_courante + 1) <= max_year else None
    else:
        min_year = max_year = annee_courante
        prev_year = next_year = None

    # 1. Budgets par nature de prestation pour l'année sélectionnée
    budgets = BudgetNature.query.filter_by(annee=annee_courante).filter(
        ~BudgetNature.nature.in_([
            'Articles de plomberies', 'Articles électriques', 
            'Articles de la peinture', 'Alimentation'
        ])
    ).all()
    
    # Si pas de budgets pour l'année sélectionnée, essayer de récupérer ceux de l'année courante
    if not budgets and year_filter and year_filter != datetime.now().year:
        annee_courante = datetime.now().year
        budgets = BudgetNature.query.filter_by(annee=annee_courante).filter(
            ~BudgetNature.nature.in_([
                'Articles de plomberies', 'Articles électriques', 
                'Articles de la peinture', 'Alimentation'
            ])
        ).all()

    # 2. Statistiques globales des engagements (Bons de Commande) pour l'année sélectionnée
    stats_bc = db.session.query(
        func.sum(BonCommande.montant_engage).label('total_engage'),
        func.sum(case((BonCommande.decision_controle == 'Visa', BonCommande.montant_engage), else_=0)).label('total_vise'),
        func.sum(case((BonCommande.date_reglement != None, BonCommande.montant_engage), else_=0)).label('total_regle')
    ).filter(func.extract('year', BonCommande.date_engagement) == annee_courante).first()

    stats = {
        'total_engage': stats_bc.total_engage or 0,
        'total_vise': stats_bc.total_vise or 0,
        'total_regle': stats_bc.total_regle or 0,
        'total_en_attente': (stats_bc.total_engage or 0) - (stats_bc.total_vise or 0)
    }

    # 3. Récupérer les engagements à afficher dans le tableau
    query_engagements = BonCommande.query.filter(func.extract('year', BonCommande.date_engagement) == annee_courante)
    if nature_filter:
        query_engagements = query_engagements.join(BudgetNature).filter(BudgetNature.nature == nature_filter)
    
    if type_depense == 'bon_commande':
        engagements = query_engagements.order_by(BonCommande.date_engagement.desc()).all()
    else:  # indemnites
        engagements = IndemnitesDeplacement.query.filter(
            func.extract('year', IndemnitesDeplacement.created_at) == annee_courante
        ).order_by(IndemnitesDeplacement.created_at.desc()).all()

    engagements = [convert_object_dates(e) for e in engagements]

    return render_template('engagements.html',
                         budgets=budgets,
                         engagements=engagements,
                         type_depense=type_depense,
                         nature_filter=nature_filter,
                         stats=stats,
                         years=all_years,
                         current_year=annee_courante,
                         prev_year=prev_year,
                         next_year=next_year)

@app.route('/add-bon-commande', methods=['GET', 'POST'])
@login_required
def add_bon_commande():
    """Ajouter un bon de commande avec nouveau système budgétaire"""
    if request.method == 'POST':
        nature_prestation = request.form['nature_prestation']
        montant_engage = float(request.form['montant_engage'])
        annee_courante = datetime.now().year
        
        # Vérifier le budget
        budget = BudgetNature.query.filter_by(nature=nature_prestation, annee=annee_courante).first()
        
        if not budget:
            flash('❌ Nature de prestation introuvable', 'danger')
            return redirect(url_for('add_bon_commande'))
        
        if budget.est_verrouille:
            flash('🔒 Budget verrouillé ! Cette nature de prestation a épuisé son budget.', 'danger')
            return redirect(url_for('add_bon_commande'))
        
        if montant_engage > budget.budget_restant:
            flash(f'⚠️ Budget insuffisant ! Montant demandé: {montant_engage:.2f} DH | Budget restant: {budget.budget_restant:.2f} DH', 'warning')
            return redirect(url_for('add_bon_commande'))
        
        # Créer le bon de commande
        bon = BonCommande(
            numero_bc=request.form['numero_bc'],
            loi_finance=request.form['loi_finance'],
            nature_prestation=nature_prestation,
            sous_categorie=request.form.get('sous_categorie'),
            date_engagement=datetime.strptime(request.form['date_engagement'], '%Y-%m-%d').date(),
            numero_fiche_navette=request.form['numero_fiche_navette'],
            montant_engage=montant_engage,
            beneficiaire=request.form['beneficiaire'],
            decision_controle=request.form['decision_controle'],
            observations_controle=request.form.get('observations_controle'),
            budget_nature_id=budget.id,
            user_id=session.get('user_id')
        )
        
        # Champs conditionnels si Visa
        if request.form['decision_controle'] == 'Visa':
            bon.numero_bordereau_emission = request.form.get('numero_bordereau_emission')
            if request.form.get('date_service_fait'):
                bon.date_service_fait = datetime.strptime(request.form['date_service_fait'], '%Y-%m-%d').date()
            bon.controle_tresorier = request.form.get('controle_tresorier')
            bon.observations_tresorier = request.form.get('observations_tresorier')
            
            if request.form.get('controle_tresorier') == 'Visa':
                if request.form.get('date_reglement'):
                    bon.date_reglement = datetime.strptime(request.form['date_reglement'], '%Y-%m-%d').date()
                bon.mode_reglement = 'Virement'
        
        db.session.add(bon)
        db.session.commit()
        
        # Enregistrer la consommation budgétaire
        consommation = ConsommationBudget(
            nature=nature_prestation,
            centre_id=None,  # Bon de commande = Unité principale
            centre_nom=None,
            montant_ttc=montant_engage,
            annee=annee_courante
        )
        db.session.add(consommation)
        db.session.commit()
        
        # Log d'audit
        create_audit_log(
            action='create',
            entity_type='bon_commande',
            entity_id=bon.id,
            entity_name=bon.numero_bc,
            description=f'Nature: {nature_prestation} | Montant: {montant_engage:.2f} DH'
        )
        
        flash(f'✅ Bon de commande {bon.numero_bc} ajouté avec succès | Budget restant: {budget.budget_restant:.2f} DH', 'success')
        
        # Alerte si budget faible (< 20%)
        if budget.budget_restant < (budget.budget_total * 0.2):
            flash(f'⚠️ Attention ! Le budget de "{nature_prestation}" est presque épuisé ({budget.budget_restant:.2f} DH restants)', 'warning')
        
        return redirect(url_for('engagements'))
    
    # GET - Afficher le formulaire
    annee_courante = datetime.now().year
    budgets = BudgetNature.query.filter_by(annee=annee_courante).filter(
        ~BudgetNature.nature.in_([
            'Articles de plomberies',
            'Articles électriques',
            'Articles de la peinture',
            'Alimentation'
        ])
    ).all()
    return render_template('add_bon_commande.html', budgets=budgets)

@app.route('/edit-bon-commande/<int:bc_id>', methods=['GET', 'POST'])
@login_required
def edit_bon_commande(bc_id):
    """Modifier un bon de commande existant"""
    bon = BonCommande.query.get_or_404(bc_id)
    
    if request.method == 'POST':
        nature_prestation = request.form['nature_prestation']
        montant_engage = float(request.form['montant_engage'])
        annee_courante = datetime.now().year
        
        # Vérifier le budget (en tenant compte du montant déjà engagé)
        budget = BudgetNature.query.filter_by(nature=nature_prestation, annee=annee_courante).first()
        
        if not budget:
            flash('❌ Nature de prestation introuvable', 'danger')
            return redirect(url_for('edit_bon_commande', bc_id=bc_id))
        
        # Calculer la différence de montant
        difference_montant = montant_engage - bon.montant_engage
        
        if difference_montant > 0 and difference_montant > budget.budget_restant:
            flash(f'⚠️ Budget insuffisant ! Augmentation demandée: {difference_montant:.2f} DH | Budget restant: {budget.budget_restant:.2f} DH', 'warning')
            return redirect(url_for('edit_bon_commande', bc_id=bc_id))
        
        # Mettre à jour le bon de commande
        bon.numero_bc = request.form['numero_bc']
        bon.loi_finance = request.form['loi_finance']
        bon.nature_prestation = nature_prestation
        bon.sous_categorie = request.form.get('sous_categorie')
        bon.date_engagement = datetime.strptime(request.form['date_engagement'], '%Y-%m-%d').date()
        bon.numero_fiche_navette = request.form['numero_fiche_navette']
        bon.montant_engage = montant_engage
        bon.beneficiaire = request.form['beneficiaire']
        bon.decision_controle = request.form['decision_controle']
        bon.observations_controle = request.form.get('observations_controle')
        
        # Nouveaux champs
        if request.form.get('date_visa'):
            bon.date_visa = datetime.strptime(request.form['date_visa'], '%Y-%m-%d').date()
        
        # Champs conditionnels si Visa
        if request.form['decision_controle'] == 'Visa':
            bon.numero_bordereau_emission = request.form.get('numero_bordereau_emission')
            if request.form.get('date_service_fait'):
                bon.date_service_fait = datetime.strptime(request.form['date_service_fait'], '%Y-%m-%d').date()
            
            # Nouveaux champs Ordonnancement
            bon.decision_controle_ordo = request.form.get('decision_controle_ordo')
            if request.form.get('date_visa_ordo'):
                bon.date_visa_ordo = datetime.strptime(request.form['date_visa_ordo'], '%Y-%m-%d').date()
            bon.etat_reglement = request.form.get('etat_reglement')
            
            if request.form.get('etat_reglement') == 'Réglé' and request.form.get('date_reglement'):
                bon.date_reglement = datetime.strptime(request.form['date_reglement'], '%Y-%m-%d').date()
                bon.mode_reglement = 'Virement'
        
        # Mettre à jour la consommation budgétaire si le montant a changé
        if difference_montant != 0:
            consommation = ConsommationBudget.query.filter_by(
                nature=bon.nature_prestation,
                centre_id=None,
                annee=annee_courante
            ).first()
            
            if consommation:
                consommation.montant_ttc += difference_montant
            else:
                # Créer une nouvelle consommation si elle n'existe pas
                consommation = ConsommationBudget(
                    nature=nature_prestation,
                    centre_id=None,
                    centre_nom=None,
                    montant_ttc=montant_engage,
                    annee=annee_courante
                )
                db.session.add(consommation)
        
        db.session.commit()
        
        # Log d'audit
        create_audit_log(
            action='update',
            entity_type='bon_commande',
            entity_id=bon.id,
            entity_name=bon.numero_bc,
            description=f'Nature: {nature_prestation} | Montant: {montant_engage:.2f} DH'
        )
        
        flash(f'✅ Bon de commande {bon.numero_bc} modifié avec succès', 'success')
        return redirect(url_for('engagements'))
    
    # GET - Afficher le formulaire avec les données existantes
    annee_courante = datetime.now().year
    budgets = BudgetNature.query.filter_by(annee=annee_courante).filter(
        ~BudgetNature.nature.in_([
            'Articles de plomberies',
            'Articles électriques',
            'Articles de la peinture',
            'Alimentation'
        ])
    ).all()
    return render_template('edit_bon_commande.html', bon=bon, budgets=budgets)

@app.route('/add-indemnites', methods=['GET', 'POST'])
@login_required
def add_indemnites():
    """Ajouter des indemnités de déplacement"""
    if request.method == 'POST':
        # Créer l'indemnité
        indemnite = IndemnitesDeplacement(
            objet=request.form['objet'],
            periode_deplacement=request.form['periode_deplacement'],
            loi_finance=request.form['loi_finance'],
            montant_engage=float(request.form['montant_engage']),
            type_engagement=request.form['type_engagement'],
            numero_engagement=request.form['numero_engagement'],
            user_id=session.get('user_id')
        )
        
        db.session.add(indemnite)
        db.session.flush()  # Pour obtenir l'ID
        
        # Ajouter les bénéficiaires (tableau dynamique)
        beneficiaires_data = request.form.getlist('beneficiaires[]')
        for ben_json in beneficiaires_data:
            import json
            ben = json.loads(ben_json)
            beneficiaire = BeneficiaireIndemnite(
                indemnite_id=indemnite.id,
                nom=ben['nom'],
                prenom=ben['prenom'],
                grade=ben['grade'],
                montant=float(ben['montant'])
            )
            db.session.add(beneficiaire)
        
        db.session.commit()
        
        # Log d'audit
        create_audit_log(
            action='create',
            entity_type='indemnites_deplacement',
            entity_id=indemnite.id,
            entity_name=indemnite.objet,
            description=f'Période: {indemnite.periode_deplacement} | Montant: {indemnite.montant_engage:.2f} DH'
        )
        
        flash(f'✅ Indemnités de déplacement ajoutées avec succès', 'success')
        return redirect(url_for('engagements', type='indemnites'))
    
    # GET
    return render_template('add_indemnites.html')

@app.route('/delete-bon-commande/<int:bc_id>', methods=['POST'])
@login_required
def delete_bon_commande(bc_id):
    """Supprimer un bon de commande et restituer le budget"""
    bon = BonCommande.query.get_or_404(bc_id)
    budget = bon.budget_nature
    montant = bon.montant_engage
    numero = bon.numero_bc
    annee_courante = datetime.now().year
    
    # Supprimer la consommation budgétaire associée
    consommation = ConsommationBudget.query.filter_by(
        nature=bon.nature_prestation,
        montant_ttc=montant,
        annee=annee_courante,
        centre_id=None  # Bon de commande = Unité
    ).first()
    
    if consommation:
        db.session.delete(consommation)
    
    # Supprimer le bon
    db.session.delete(bon)
    db.session.commit()
    
    # Log d'audit
    create_audit_log(
        action='delete',
        entity_type='bon_commande',
        entity_id=bc_id,
        entity_name=numero,
        description=f'Budget restitué: {montant:.2f} DH'
    )
    
    flash(f'✅ Bon de commande {numero} supprimé | Budget restitué: {montant:.2f} DH', 'success')
    return redirect(url_for('engagements'))

@app.route('/historique-budget/<int:budget_id>')
@login_required
def historique_budget(budget_id):
    """Afficher l'historique des consommations d'un budget"""
    budget = BudgetNature.query.get_or_404(budget_id)
    bons = BonCommande.query.filter_by(budget_nature_id=budget_id).order_by(BonCommande.date_engagement.desc()).all()
    
    return render_template('historique_budget.html', budget=budget, bons=bons)

# ========================= Modules additionnels =========================
@app.route('/fournisseurs')
@login_required
def fournisseurs():
    # Sécuriser la présence des nouvelles colonnes avant toute requête
    try:
        ensure_supplier_extra_columns()
    except Exception:
        pass
    fournisseurs = Fournisseur.query.order_by(Fournisseur.nom.asc()).all()
    return render_template('suppliers.html', fournisseurs=fournisseurs)

@app.route('/fournisseurs/ajouter', methods=['GET', 'POST'])
@login_required
def add_fournisseur():
    if request.method == 'POST':
        form = request.form
        
        # --- Validation des champs requis ---
        required_fields = ['nom', 'adresse', 'telephone', 'email', 'ville', 'registre_commerce', 'identification_fonciere', 'compte_bancaire_ribe']
        for field in required_fields:
            if not form.get(field):
                flash(f'Le champ "{field.replace("_", " ").title()}" est obligatoire.', 'danger')
                return redirect(url_for('add_fournisseur'))

        # --- Vérification de l'unicité ---
        if Fournisseur.query.filter_by(nom=form['nom']).first():
            flash('Un fournisseur avec cette raison sociale existe déjà.', 'warning')
            return redirect(url_for('add_fournisseur'))
        
        # --- Création du fournisseur ---
        nouveau_fournisseur = Fournisseur(
            nom=form['nom'],
            adresse=form['adresse'],
            telephone=form['telephone'],
            fax=form.get('fax'),
            email=form['email'],
            ville=form['ville'],
            registre_commerce=form['registre_commerce'],
            identification_fonciere=form['identification_fonciere'],
            compte_bancaire_ribe=form['compte_bancaire_ribe'],
            statut='En attente'  # Statut par défaut
        )
        db.session.add(nouveau_fournisseur)
        db.session.flush() # Pour obtenir l'ID du nouveau fournisseur avant le commit

        # --- Traitement des catégories ---
        category_ids = request.form.getlist('categories')
        if category_ids:
            categories = Categorie.query.filter(Categorie.id.in_(category_ids)).all()
            nouveau_fournisseur.categories.extend(categories)

        # --- Traitement dynamique des contacts ---
        contact_index = 0
        while f'contact_nom_{contact_index}' in form:
            nom = form.get(f'contact_nom_{contact_index}')
            if nom: # On ajoute le contact seulement si un nom est fourni
                nouveau_contact = Contact(
                    nom=nom,
                    poste=form.get(f'contact_poste_{contact_index}'),
                    email=form.get(f'contact_email_{contact_index}'),
                    telephone=form.get(f'contact_telephone_{contact_index}'),
                    fournisseur_id=nouveau_fournisseur.id
                )
                db.session.add(nouveau_contact)
            contact_index += 1

        try:
            db.session.commit()
            flash('Fournisseur ajouté avec succès, incluant les contacts et catégories.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de l\'ajout du fournisseur : {e}', 'danger')

        return redirect(url_for('fournisseurs'))

    # --- Méthode GET ---
    toutes_les_categories = Categorie.query.order_by(Categorie.nom).all()
    return render_template('add_supplier.html', toutes_les_categories=toutes_les_categories)

@app.route('/fournisseurs/<int:supplier_id>/supprimer', methods=['POST'])
@login_required
def delete_fournisseur(supplier_id):
    fournisseur = Fournisseur.query.get_or_404(supplier_id)
    try:
        db.session.delete(fournisseur)
        db.session.commit()
        flash('Fournisseur supprimé avec succès.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression du fournisseur: {str(e)}', 'danger')
    # Retourner vers la liste des fournisseurs (endpoint existant)
    return redirect(url_for('fournisseurs'))

@app.route('/fournisseurs/<int:supplier_id>/modifier', methods=['GET', 'POST'])
@login_required
def edit_fournisseur(supplier_id):
    """Modifier un fournisseur existant (informations + catégories)."""
    fournisseur = Fournisseur.query.get_or_404(supplier_id)
    toutes_les_categories = Categorie.query.order_by(Categorie.nom).all()

    if request.method == 'POST':
        form = request.form

        # Validation des champs requis (même logique que pour add_fournisseur)
        required_fields = ['nom', 'adresse', 'telephone', 'email', 'ville',
                           'registre_commerce', 'identification_fonciere', 'compte_bancaire_ribe']
        for field in required_fields:
            if not form.get(field):
                flash(f'Le champ "{field.replace("_", " ").title()}" est obligatoire.', 'danger')
                return redirect(url_for('edit_fournisseur', supplier_id=supplier_id))

        # Mise à jour des informations principales
        fournisseur.nom = form['nom']
        fournisseur.adresse = form['adresse']
        fournisseur.ville = form['ville']
        fournisseur.email = form['email']
        fournisseur.telephone = form['telephone']
        fournisseur.fax = form.get('fax')
        fournisseur.registre_commerce = form['registre_commerce']
        fournisseur.identification_fonciere = form['identification_fonciere']
        fournisseur.compte_bancaire_ribe = form['compte_bancaire_ribe']
        fournisseur.statut = form.get('statut', fournisseur.statut or 'En attente')
        fournisseur.notes = form.get('notes')

        # Mise à jour des catégories associées
        category_ids = form.getlist('categories')
        if category_ids:
            categories = Categorie.query.filter(Categorie.id.in_(category_ids)).all()
            fournisseur.categories = categories
        else:
            fournisseur.categories = []

        try:
            db.session.commit()
            flash('Fournisseur mis à jour avec succès.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f"Erreur lors de la mise à jour du fournisseur : {e}", 'danger')

        return redirect(url_for('edit_fournisseur', supplier_id=supplier_id))

    # GET - Afficher le formulaire pré-rempli
    selected_category_ids = {cat.id for cat in fournisseur.categories}
    return render_template(
        'edit_supplier.html',
        fournisseur=fournisseur,
        toutes_les_categories=toutes_les_categories,
        selected_category_ids=selected_category_ids
    )

@app.route('/fournisseur/<int:fournisseur_id>/ajouter_contact', methods=['POST'])
@login_required
def ajouter_contact(fournisseur_id):
    fournisseur = Fournisseur.query.get_or_404(fournisseur_id)
    # ... (le reste du code reste inchangé)
    nouveau_contact = Contact(
        nom=request.form.get('nom'),
        poste=request.form.get('poste'),
        telephone=request.form.get('telephone'),
        email=request.form.get('email'),
        fournisseur_id=fournisseur.id
    )
    
    db.session.add(nouveau_contact)
    db.session.commit()
    
    flash('Contact ajouté avec succès', 'success')
    return redirect(url_for('edit_fournisseur', supplier_id=fournisseur_id))

@app.route('/contact/<int:id>/supprimer', methods=['GET', 'POST'])
@login_required
def supprimer_contact(id):
    contact = Contact.query.get_or_404(id)
    fournisseur_id = contact.fournisseur_id
    db.session.delete(contact)
    db.session.commit()
    flash('Contact supprimé avec succès', 'success')
    return redirect(url_for('fournisseurs'))

#  

# Désactiver le cache en mode développement
@app.after_request
def add_no_cache_headers(response):
    """Ajoute des headers pour désactiver le cache en développement"""
    if app.debug:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '-1'
    return response

def seed_categories():
    """Peuple la table des catégories avec des données initiales."""
    if Categorie.query.first():
        print("La table 'categorie' contient déjà des données. Le peuplement est ignoré.")
        return

    print("Peuplement de la table 'categorie'...")
    
    # Catégories principales
    cat_entretien = Categorie(nom='ENTRETIEN DES BÂTIMENTS ADMINISTRATIFS')
    autres_cats = [
        Categorie(nom='FOURNITURES DE BUREAU ET DOCUMENTATION'),
        Categorie(nom="PRODUITS D'HYGIÈNE ET DE DÉSINFECTION"),
        Categorie(nom='FOURNITURES POUR MATÉRIEL INFORMATIQUE')
    ]
    
    db.session.add(cat_entretien)
    db.session.add_all(autres_cats)
    db.session.commit()

    # Sous-catégories pour Entretien
    sous_cats_entretien = [
        Categorie(nom='Articles de plomberie', parent_id=cat_entretien.id),
        Categorie(nom='Articles électriques', parent_id=cat_entretien.id),
        Categorie(nom='Articles de la peinture', parent_id=cat_entretien.id)
    ]
    db.session.add_all(sous_cats_entretien)
    db.session.commit()
    print("Peuplement de la table 'categorie' terminé.")

@app.cli.command('init-db')
def init_db_command():
    """Initialise la base de données et peuple les catégories."""
    db.create_all()
    print('Base de données initialisée.')
    seed_categories()

def fix_sidi_slimane_unite():
    """Corrige l'unité SIDI SLIMANE qui doit être une Unité Provinciale indépendante"""
    try:
        from sqlalchemy import text
        # Chercher l'unité par nom
        sidi_slimane = Unite.query.filter(Unite.nom.like('%SIDI SLIMANE%')).first()
        if sidi_slimane:
            if sidi_slimane.type != 'unite' or sidi_slimane.parent_id is not None:
                sidi_slimane.type = 'unite'
                sidi_slimane.parent_id = None
                db.session.commit()
                print("✅ Correction appliquée : UNITE SIDI SLIMANE est maintenant une Unité Provinciale indépendante.")
    except Exception as e:
        db.session.rollback()
        print(f"❌ Erreur lors de la correction de SIDI SLIMANE : {str(e)}")

with app.app_context():
    try:
        db.create_all()
        ensure_soft_delete_columns()
        ensure_supplier_extra_columns()
        ensure_unite_hierarchy_columns()
        ensure_budget_individual_migration()
        fix_sidi_slimane_unite() # Correction SIDI SLIMANE
        
        # S'assurer que les colonnes 'is_super_admin' et 'is_active' existent dans la table 'users'
        res = db.session.execute(text("PRAGMA table_info('users')"))
        existing_cols = {row[1] for row in res}
        if 'is_super_admin' not in existing_cols:
            db.session.execute(text("ALTER TABLE users ADD COLUMN is_super_admin BOOLEAN DEFAULT 0"))
            db.session.commit()
        if 'is_active' not in existing_cols:
            db.session.execute(text("ALTER TABLE users ADD COLUMN is_active BOOLEAN DEFAULT 1"))
            db.session.commit()
            
        # PROMOTION AUTOMATIQUE (Rescue): Désactivé à la demande de l'utilisateur
        # agent_user = User.query.filter_by(email='agent@protectioncivile.ma').first()
        # if agent_user:
        #     if not agent_user.is_super_admin or agent_user.role != 'admin':
        #         agent_user.is_super_admin = True
        #         agent_user.role = 'admin'
        #         agent_user.is_active = True
        #         db.session.commit()
        #         print("🚀 Rescue: 'agent@protectioncivile.ma' a été promu Super Admin.")
        # else:
        #     # Créer le compte s'il n'existe pas du tout (Premier démarrage ou DB vide)
        #     try:
        #         new_admin = User(
        #             email='agent@protectioncivile.ma',
        #             name='Administrateur Principal',
        #             role='admin',
        #             is_super_admin=True,
        #             is_active=True
        #         )
        #         new_admin.set_password('Admin123!') # Mot de passe par défaut sécurisé à changer
        #         db.session.add(new_admin)
        #         db.session.commit()
        #         print("🚀 Rescue: Compte 'agent@protectioncivile.ma' créé avec succès (MDP: Admin123!).")
        #     except Exception as creation_err:
        #         print(f"⚠️ Erreur lors de la création du compte rescue: {creation_err}")
        #         db.session.rollback()

    except Exception as e:
        print(f"⚠️ Erreur lors de l'initialisation/migration: {e}")

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
